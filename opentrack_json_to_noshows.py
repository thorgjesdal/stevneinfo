# -*- coding: utf-8 -*-

# TODO: 
#       + clean up/more modular
#       + PARA categories
#       + sorting order
#       + ties
#       + foreign and non-default teams
#       + best valid attempt (wind)
#       + rounds on different days
#       + series in multi round field events
#
import sys
import os
import json
import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import colors as xlcolors
from openpyxl.styles import Font, Color
import requests
import random
from collections import defaultdict
import argparse

from stevneinfo import clubs, events, opentrack as ot

from pprint import pformat, pprint


def is_relay(event_code):
    # from athlib/codes.py
    PAT_RELAYS = re.compile(r"^(?:(\d{1,2})[xX]((\d+(\.\d+)?)[hHMK]?|[rR][eE][lL][aA][yY]|[sS]?[dDsS][mM][rR]|[sS][wW][rR]))$") # 4x100, 4x400, 4xReLAy, 4xDMR, 4xSMR, 12x200H
    match = re.search(PAT_RELAYS, event_code)
    return match is not None


#---------------------------------------
parser = argparse.ArgumentParser()
parser.add_argument('url')
args = parser.parse_args()
   
url = args.url
print(url)

j = ot.fetch_json(url)
print( j.keys() )

meetname = j['fullName']
slug = j['slug']
if j.get('venue') == None: 
    venue = ''
else:
    venue = j['venue']['formalName']

d  = j['date']
d2 = j['finishDate']
isodateformat = "%Y-%m-%d"
date0 = datetime.datetime.strptime(d, isodateformat)
date1 = datetime.datetime.strptime(d2, isodateformat)

no_shows = []
for c in j['competitors']:
    fn = ''; ln = ''; dob= ''; t=''
    bib  = c['competitorId']
    if 'firstName' in c.keys():
        fn   = c['firstName']
    if 'lastName' in c.keys():
        ln   = c['lastName']
    if 'category' in c.keys():
        cat   = c['category']
    if 'teamName' in c.keys():
        t    = c['teamName']
    for e in c['eventsEntered']:
        if not e['timeCheckedIn']:
            if not is_relay(e['eventCode']):
                no_shows.append( (fn, ln, t, cat, e['eventId']) )
events = {}
for e in j['events']:
    id = e['eventId']
    event = e['name']
    events[id] = event
    


#... write no shows to xlsx workbook
wb = Workbook()

ws = wb.active
    
ws.title = "No shows"
    
ws['a1'] = meetname
ws['a2'] = venue
ws['a3'] = date0.strftime('%d.%m.%Y'); ws['b3'] = date1.strftime('%d.%m.%Y')

ws['a5'] = 'Påmeldte som ikke svarte på opprop'

row_counter = 7 

for line in no_shows:
    ws[f'a{row_counter}'] = f'{line[0]} {line[1]}'
    ws[f'b{row_counter}'] = line[2]
    ws[f'c{row_counter}'] = events[line[4]]
    row_counter += 1


xlname = slug + '-NoShows.xlsx'
wb.save(xlname)
