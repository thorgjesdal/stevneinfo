# -*- coding: utf-8 -*-

# TODO: 
#       + combined events results
#       + clean up/more modular
#       + different sorting critera (age, cats in json, predefined cats)
#       + PARA categories
#       + sorting order
#
import sys
import json
import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import colors as xlcolors
from openpyxl.styles import Font, Color
import requests
import random
from collections import defaultdict

import pprint

noplace = int(1.e10)-1
def get_category(birthdate, eventdate, gender):
    birthyear = birthdate.year
    eventyear = eventdate.year
    age = int(eventyear)-int(birthyear)

    g = {'F' : 'J', 'M' : 'G' }
    if age > 19:
        g = {'F' : 'K', 'M' : 'M' }
        if age < 35:
           a = 'S'
        else:
           a = 'V' + '%d'%(5*int(age/5))
    elif age in (18,19):
        a = '18/19'
    else:
        a = '%d'%(age)

    cat = g[gender]+a
    return cat

   
def event_name(code):
    event_names = {
            '60'     : '60 meter'          , 
            '80'     : '80 meter'          , 
            '100'    : '100 meter'         , 
            '150'    : '150 meter'         , 
            '200'    : '200 meter'         , 
            '300'    : '300 meter'         , 
            '400'    : '400 meter'         , 
            '600'    : '600 meter'         , 
            '800'    : '800 meter'         , 
            '1000'   : '1000 meter'        , 
            '2000'   : '2000 meter'        , 
            '1500'   : '1500 meter'        , 
            '3000'   : '3000 meter'        , 
            '5000'   : '5000 meter'        , 
            '10000'  : '10000 meter'       , 
            '60H'    : '60 meter hekk'     , 
            '80H'    : '80 meter hekk'     , 
            '100H'   : '100 meter hekk'    , 
            '110H'   : '110 meter hekk'    , 
            '200H'   : '200 meter hekk'    ,
            '300H'   : '300 meter hekk'    , 
            '400H'   : '400 meter hekk'    , 
            '1500SC' : '1500 meter hinder' , 
            '2000SC' : '2000 meter hinder' , 
            '3000SC' : '3000 meter hinder' , 
            '1000W'  : 'Kappgang 1000 meter'        , 
            '2000W'  : 'Kappgang 2000 meter'        , 
            '3000W'  : 'Kappgang 3000 meter'        , 
            'HJ'     : 'Høyde'             , 
            'PV'     : 'Stav'              , 
            'LJ'     : 'Lengde'            , 
            'TJ'     : 'Tresteg'           , 
            'SP'     : 'Kule'              , 
            'DT'     : 'Diskos'            , 
            'HT'     : 'Slegge'            , 
            'JT'     : 'Spyd'              , 
            'OT'     : 'Liten ball'              , 
            'BT'     : 'Liten ball'              , 
            'DEC'    : 'Tikamp'            , 
            'HEP'    : 'Sjukamp'           ,
            'PEN'    : 'Femkamp'           ,
            'SHJ'    : 'Høyde uten tilløp' ,
            'SLJ'    : 'Lengde uten tilløp'           ,
            'STJ'    : 'Tresteg uten tilløp'           
            }
    return event_names[code]

def event_spec(event, klasse):
    # 18.05.2020 rewrite based om implements.py form athlib
    gender = 'F'
    if klasse[0] in ('M', 'G'):
        gender = 'M'

    weight = ''
    if event == 'SP' or event == 'HT':
        if gender == 'F':
           if klasse in ('J10', 'J11', 'J12', 'J13'):
              weight = '2,0kg'
           elif klasse in ('J14', 'J15', 'J15', 'J16', 'J17'):
               weight = '3,0kg'
           elif klasse in ('J18/19', 'KU20', 'KS', 'KV35', 'KV40', 'KV45'):
               weight = '4,0kg'
           elif klasse >= 'KV50':
               weight = '3,0kg'
        elif gender == 'M':
           if klasse in ('G10', 'G11' ):
              weight = '2,0kg'
           elif klasse in ('G12', 'G13' ):
               weight = '3,0kg'
           elif klasse in ('G14', 'G15', 'MV70', 'MV75' ):
               weight = '4,0kg'
           elif klasse in ('G16', 'G17', 'MV60', 'MV65' ):
               weight = '5,0kg'
           elif klasse in ('G18/19', 'MU20', 'MV50', 'MV55' ):
               weight = '6,0kg'
           elif klasse in ('MS', 'MU23', 'MV35', 'MV40', 'MV45' ):
               weight = '7,26kg'
           elif klasse >= 'MV80':
               weight = '3,0kg'
    elif event == 'DT' :
        if gender == 'F':
           if klasse in ('J10', 'J11', 'J12', 'J13'):
              weight = '0,6kg'
           elif klasse in ('J14', 'J15'):
               weight = '0,75kg'
           elif klasse >= 'KV80':
               weight = '0,75kg'
           else:
               weight = '1,0kg'
        elif gender == 'M':
           if klasse in ('G10', 'G11' ):
              weight = '0,6kg'
           elif klasse in ('G12', 'G13' ):
               weight = '0,75kg'
           elif klasse in ('G14', 'G15' ):
               weight = '1,0kg'
           elif klasse in ('G16', 'G17', 'MV50', 'MV55' ):
               weight = '1,5kg'
           elif klasse in ('G18/19', 'MU20' ):
               weight = '1,75kg'
           elif klasse in ('MS', 'MU23', 'MV35', 'MV40', 'MV45' ):
               weight = '2,0kg'
           elif klasse >= 'MV60':
               weight = '1,0kg'
    elif event == 'JT' :
        if gender == 'F':
           if klasse in ('J10', 'J11', 'J12', 'J13', 'J14'):
              weight = '400g'
           elif klasse in ('J15', 'J16', 'J17', 'KV50', 'KV55'):
               weight = '500g'
           elif klasse in ('J18/19', 'KU20', 'KU23', 'KS', 'KV35', 'KV40', 'KV45'):
               weight = '600g'
           elif klasse >= 'KV60':
               weight = '400g'
        elif gender == 'M':
           if klasse in ('G10', 'G11', 'G13', ):
              weight = '400g'
           elif klasse in ('G14', 'G15', 'MV60', 'MV65' ):
               weight = '600g'
           elif klasse in ('G16', 'G17', 'MV50', 'MV55' ):
               weight = '700g'
           elif klasse in ('G18/19', 'MU20', 'MS', 'MU23', 'MV35', 'MV40', 'MV45' ):
               weight = '800g'
           elif klasse in ('MV70', 'MV75' ):
               weight = '500g'
           elif klasse >= 'MV80':
               weight = '400g'
    elif event == 'OT' :
        weight='150g'


    throws = {}
    throws['SP'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'KV35' : '4,0kg', 'KV40' : '4,0kg', 'KV45' : '4,0kg', 
                       'KV50' : '3,0kg', 'KV55' : '3,0kg', 'KV60' : '3,0kg', 'KV65' : '3,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg'
                       } 
    throws['DT'] = { 'J10' : '0,6kg', 'J11' : '0,6kg', 'J12' : '0,6kg', 'J13' : '0,6kg', 
                       'J14' : '0,75kg', 'J15' : '0,75kg', 'J16' : '0,75kg', 'J17' : '0,75kg',
                       'J18/19' : '1,0kg', 'KU20' : '1,0kg', 'KU23' : '1,0kg', 'KS' : '1,0kg', 
                       'G10' : '0,6kg', 'G11' : '0,6kg', 'G12' : '0,75kg', 'G13' : '0,75kg', 
                       'G14' : '1,0kg', 'G15' : '1,0kg', 'G16' : '1,5kg', 'G17' : '1,5kg',
                       'G18/19' : '1,75kg', 'MU20' : '1,75kg', 'MU23' : '2,0kg', 'MS' : '2,0kg', 
                       'MV35' : '2,0kg', 'MV40' : '2,0kg', 'MV45' : '2,0kg',
                       'MV50' : '1,5kg', 'MV55' : '1,5kg', 
                       'MV60' : '1,0kg', 'MV65' : '1,0kg', 'MV70' : '1,0kg', 'MV75' : '1,0kg' 
                       } 
    throws['HT'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg'} 
    throws['JT'] = { 'J10' : '400g', 'J11' : '400g', 'J12' : '400g', 'J13' : '400g', 
                       'J14' : '400g', 'J15' : '500g', 'J16' : '500g', 'J17' : '500g',
                       'J18/19' : '600g', 'KU20' : '600g', 'KU23' : '600g', 'KS' : '600g', 
                       'G10' : '400g', 'G11' : '400g', 'G12' : '400g', 'G13' : '400g', 
                       'G14' : '600g', 'G15' : '600g', 'G16' : '700g', 'G17' : '700g',
                       'G18/19' : '800g', 'MU20' : '800g', 'MU23' : '800g', 'MS' : '800g'} 
    throws['OT'] = { 'J10' : '150g', 'J11' : '150g', 'J12' : '150g', 'J13' : '150g', 'J14' : '150g', 
                             'G10' : '150g', 'G11' : '150g', 'G12' : '150g', 'G13' : '150g', 'G14' : '150g' }
    hurdles = {}
    hurdles['60H'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '76,2cm', 'J13' : '76,2cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'KV50' : '76,2cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '76,2cm', 'G13' : '76,2cm', 'G14' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm', 'G17' : '91,4cm',
                                 'G18/19' : '100cm','MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm' ,
                                 'MV35':'100cm', 'MV40':'91,4cm', 'MV45':'91,4cm', 'MV50':'91,4cm', 'MV55':'91,4cm', 
                                 'MV60':'84cm', 'MV65':'84cm', 'MV70':'76,2cm', 'MV75':'76,2cm', 
                                 }
    hurdles['80H'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'G14' : '84,0cm' } 
    hurdles['100H'] = { 'J16' : '76,2cm', 'J17' : '76,2cm', 'J18/19' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm'}
    hurdles['110H'] = { 'G17' : '91,4cm', 'G18/19' : '100cm','MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm' }
    hurdles['200H'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '68,0cm', 'J13' : '68,0cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '68,0cm', 'G13' : '68,0cm', 'G14' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '76,2cm', 'G17' : '76,2cm',
                                 'G18/19' : '76,2cm','MU20' : '76,2cm', 'MU23' : '76,2cm', 'MS' : '76,2cm' }
    hurdles['300H'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm',
                                 'default':''}
    hurdles['400H'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm' }


    if event in ('SP', 'DT', 'JT', 'HT', 'OT'):
       #e = event + ' ' + throws[event][klasse]
#       e = event_name(event) + ' ' + throws[event].get(klasse,'')
       e = event_name(event) + ' ' + weight
    elif event in ('60H', '80H', '100H', '110H', '200H', '300H', '400H'): 
#      e = event_name(event) + ' ' + hurdles[event][klasse]
       e = event_name(event) + ' ' + hurdles[event].get(klasse,'')
    else:
       e = event_name(event)

    return e

def club_name(club_code):
    if club_code == (u'AALEN'):
       club_name=u'ÅLEN IDRETTSLAG'
    elif club_code == (u'AASUN'):
       club_name=u'Ålesund Friidrettsklubb'
    elif club_code == (u'IKV'):
       club_name=u'Allianseidrettslaget Ik Våg'
    elif club_code == (u'ALM'):
       club_name=u'Almenning Il'
    elif club_code == (u'ALSV'):
       club_name=u'Alsvåg Idrettslag'
    elif club_code == (u'ALTA'):
       club_name=u'Alta Idrettsforening'
    elif club_code == (u'ANDAL'):
       club_name=u'Åndalsnes Idrettsforening'
    elif club_code == (u'ANDE'):
       club_name=u'Andebu Idrettslag'
    elif club_code == (u'ANDSK'):
       club_name=u'Andørja Sportsklubb'
    elif club_code == (u'ARE'):
       club_name=u'Aremark Idrettsforening'
    elif club_code == (u'ARNA'):
       club_name=u'Arna Turn & Idrettslag'
    elif club_code == (u'ASIL'):
       club_name=u'Ås Idrettslag'
    elif club_code == (u'AASEN'):
       club_name=u'Åsen Idrettslag'
    elif club_code == (u'AASER'):
       club_name=u'Åseral idrettslag'
    elif club_code == (u'ASK'):
       club_name=u'Ask Friidrett'
    elif club_code == (u'ASKFL'):
       club_name=u'Asker Fleridrettslag'
    elif club_code == (u'ASKSK'):
       club_name=u'Asker Skiklubb'
    elif club_code == (u'ASKIM'):
       club_name=u'Askim Idrettsforening'
    elif club_code == (u'ATNA'):
       club_name=u'Atna Idrettslag'
    elif club_code == (u'AURE'):
       club_name=u'Aure Idrettslag'
    elif club_code == (u'AURL'):
       club_name=u'Aurland Idrettslag'
    elif club_code == (u'AURS'):
       club_name=u'Aurskog-Høland Friidrettslag'
    elif club_code == (u'AUFJ'):
       club_name=u'Austefjord Idrettslag'
    elif club_code == (u'AUST'):
       club_name=u'Austevoll Idrettsklubb'
    elif club_code == (u'AUSTR'):
       club_name=u'Austrheim Idrettslag'
    elif club_code == (u'BAGN'):
       club_name=u'Bagn Idrettslag'
    elif club_code == (u'BAKKE'):
       club_name=u'Bakke IF'
    elif club_code == (u'BALE'):
       club_name=u'Balestrand IL'
    elif club_code == (u'BARD'):
       club_name=u'Bardu Idrettslag'
    elif club_code == (u'BIF'):
       club_name=u'Bøler Idrettslag'
    elif club_code == (u'BTSFJ'):
       club_name=u'Båtsfjord Sportsklubb'
    elif club_code == (u'BGND'):
       club_name=u'Begnadalen Idrettslag'
    elif club_code == (u'BEIT'):
       club_name=u'Beitstad Idrettslag'
    elif club_code == (u'BCK'):
       club_name=u'Bergen Cykleklubb'
    elif club_code == (u'BTC'):
       club_name=u'Bergen Triathlon Club'
    elif club_code == (u'BTU'):
       club_name=u'Bergens Turnforening'
    elif club_code == (u'BERGE'):
       club_name=u'Berger Idrettslag'
    elif club_code == (u'BFGL'):
       club_name=u'BFG Bergen Løpeklubb'
    elif club_code == (u'BJERR'):
       club_name=u'Bjerkreim Idrettslag'
    elif club_code == (u'BJERV'):
       club_name=u'Bjerkvik Idrettsforening'
    elif club_code == (u'BLA'):
       club_name=u'Blaker IL'
    elif club_code == (u'BLEFJ'):
       club_name=u'Blefjell Idrettslag'
    elif club_code == (u'BODO'):
       club_name=u'Bodø & Omegn IF Friidrett'
    elif club_code == (u'BODL'):
       club_name=u'Bodø Bauta Løpeklubb'
    elif club_code == (u'BODF'):
       club_name=u'Bodø Friidrettsklubb'
    elif club_code == (u'BOKN'):
       club_name=u'Bokn Idrettslag'
    elif club_code == (u'BOS'):
       club_name=u'Bossekop Ungdomslag'
    elif club_code == (u'BOTNA'):
       club_name=u'Botnan Idrettslag'
    elif club_code == (u'BOT'):
       club_name=u'Botne Skiklubb'
    elif club_code == (u'BRNDB'):
       club_name=u'Brandbu IF'
    elif club_code == (u'BRATS'):
       club_name=u'Bratsberg Idrettslag'
    elif club_code == (u'BRATTV'):
       club_name=u'Brattvåg Idrettslag'
    elif club_code == (u'BREIM'):
       club_name=u'Breimsbygda IL'
    elif club_code == (u'BREKK'):
       club_name=u'Brekke Idrettslag'
    elif club_code == (u'BREMA'):
       club_name=u'Bremanger Idrettslag'
    elif club_code == (u'BREM'):
       club_name=u'Bremnes Idrettslag'
    elif club_code == (u'BRE'):
       club_name=u'Brevik Idrettslag'
    elif club_code == (u'BROMM'):
       club_name=u'Bromma Idrettslag'
    elif club_code == (u'BRYF'):
       club_name=u'Bryne Friidrettsklubb'
    elif club_code == (u'BRYT'):
       club_name=u'BRYNE TRIATLONKLUBB'
    elif club_code == (u'BUD'):
       club_name=u'Bud Idrettslag'
    elif club_code == (u'BYS'):
       club_name=u'Byaasen Skiklub'
    elif club_code == (u'BYI'):
       club_name=u'Byåsen Idrettslag'
    elif club_code == (u'BYN'):
       club_name=u'Byneset IL Hovedlaget'
    elif club_code == (u'BSK'):
       club_name=u'Bækkelagets SK'
    elif club_code == (u'BRVHA'):
       club_name=u'Bærums Verk Hauger Idrettsforening'
    elif club_code == (u'BVRFJ'):
       club_name=u'Bæverfjord Idrettslag'
    elif club_code == (u'BIF'):
       club_name=u'Bøler Idrettsforening'
    elif club_code == (u'BMLO'):
       club_name=u'Bømlo Idrettslag'
    elif club_code == (u'BRSA'):
       club_name=u'Børsa Idrettslag'
    elif club_code == (u'DALE'):
       club_name=u'Dale Idrettslag'
    elif club_code == (u'DLN'):
       club_name=u'Dalen Idrettslag'
    elif club_code == (u'DIM'):
       club_name=u'Dimna IL'
    elif club_code == (u'DMB'):
       club_name=u'Dombås Idrettslag'
    elif club_code == (u'DRIV'):
       club_name=u'Driv Idrettslag'
    elif club_code == (u'DRIVA'):
       club_name=u'Driva IL'
    elif club_code == (u'DRFR'):
       club_name=u'Drøbak-Frogn Idrettslag'
    elif club_code == (u'DPVG'):
       club_name=u'Dypvåg Idrettsforening'
    elif club_code == (u'EGRSU'):
       club_name=u'Egersunds Idrettsklubb'
    elif club_code == (u'EIDIL'):
       club_name=u'Eid Idrettslag'
    elif club_code == (u'EIDA'):
       club_name=u'Eidanger Idrettslag'
    elif club_code == (u'EIDF'):
       club_name=u'Eidfjord Idrettslag'
    elif club_code == (u'EIDSB'):
       club_name=u'Eidsberg Idrettslag'
    elif club_code == (u'EIDS'):
       club_name=u'Eidsvåg Idrettslag'
    elif club_code == (u'EIDTU'):
       club_name=u'Eidsvold Turnforening Friidrett'
    elif club_code == (u'EIK'):
       club_name=u'Eikanger Idrettslag'
    elif club_code == (u'ESK'):
       club_name=u'Ekeberg Sports Klubb'
    elif club_code == (u'ESPA'):
       club_name=u'Espa Idrettslag'
    elif club_code == (u'ETNE'):
       club_name=u'Etne Idrettslag'
    elif club_code == (u'FAGIL'):
       club_name=u'Fagernes Idrettslag N'
    elif club_code == (u'FAG'):
       club_name=u'Fagernes Idrettslag O'
    elif club_code == (u'FALK'):
       club_name=u'Falkeid idrettslag'
    elif club_code == (u'FANA'):
       club_name=u'Fana IL'
    elif club_code == (u'FEIR'):
       club_name=u'Feiring Idrettslag'
    elif club_code == (u'FET'):
       club_name=u'Fet Friidrettsklubb'
    elif club_code == (u'FA77'):
       club_name=u'FIL AKS-77'
    elif club_code == (u'FINNY'):
       club_name=u'Finnøy Idrettslag'
    elif club_code == (u'FISIF'):
       club_name=u'Fiskå Idrettsforening'
    elif club_code == (u'FISIL'):
       club_name=u'Fiskå Idrettslag'
    elif club_code == (u'FITJ'):
       club_name=u'Fitjar Idrettslag'
    elif club_code == (u'FJVE'):
       club_name=u'Fjellhug/Vereide IL'
    elif club_code == (u'FLATS'):
       club_name=u'Flatås Idrettslag'
    elif club_code == (u'FLOR'):
       club_name=u'Florø T & IF'
    elif club_code == (u'FOLFO'):
       club_name=u'Follafoss Idrettslag'
    elif club_code == (u'FOL'):
       club_name=u'Folldal Idrettsforening'
    elif club_code == (u'FOLLO'):
       club_name=u'Follo Løpeklubb'
    elif club_code == (u'FORRA'):
       club_name=u'Forra Idrettslag'
    elif club_code == (u'FOSSU'):
       club_name=u'Fossum Idrettsforening'
    elif club_code == (u'FRED'):
       club_name=u'Fredrikstad Idrettsforening'
    elif club_code == (u'FREI'):
       club_name=u'Freidig Sportsklubben'
    elif club_code == (u'ORION'):
       club_name=u'Friidretsklubben Orion'
    elif club_code == (u'REN'):
       club_name=u'Friidrettsklubben Ren-Eng'
    elif club_code == (u'BAMSE'):
       club_name=u'Friidrettslaget Bamse'
    elif club_code == (u'BORG'):
       club_name=u'Friidrettslaget Borg'
    elif club_code == (u'FRISK'):
       club_name=u'Friidrettslaget Frisk'
    elif club_code == (u'FRO'):
       club_name=u'Frognerparken Idrettslag'
    elif club_code == (u'FROL'):
       club_name=u'Frol Idrettslag'
    elif club_code == (u'FROSTA'):
       club_name=u'Frosta Idrettslag'
    elif club_code == (u'FRLND'):
       club_name=u'Frøyland Idrettslag'
    elif club_code == (u'FUR'):
       club_name=u'Furuset Allidrett IF'
    elif club_code == (u'FYLL'):
       club_name=u'Fyllingen Idrettslag'
    elif club_code == (u'FRDE'):
       club_name=u'Førde Idrettslag'
    elif club_code == (u'GAULA'):
       club_name=u'Gaular IL'
    elif club_code == (u'GAU'):
       club_name=u'Gausdal Friidrettsklubb'
    elif club_code == (u'GEI'):
       club_name=u'Geilo Idrettslag'
    elif club_code == (u'GEIR'):
       club_name=u'Geiranger Idrettslag'
    elif club_code == (u'GJER'):
       club_name=u'Gjerpen Idrettsforening'
    elif club_code == (u'GJERS'):
       club_name=u'Gjerstad Idrettslag'
    elif club_code == (u'GJDAL'):
       club_name=u'Gjesdal Idrettslag'
    elif club_code == (u'GJFK'):
       club_name=u'Gjøvik Friidrettsklubb'
    elif club_code == (u'GJVIK'):
       club_name=u'Gjøvik Friidrettsklubb 2'
    elif club_code == (u'GLO'):
       club_name=u'Gloppen Friidrettslag'
    elif club_code == (u'GOL'):
       club_name=u'Gol Idrettslag'
    elif club_code == (u'GRON'):
       club_name=u'Grong Idrettslag'
    elif club_code == (u'GRO'):
       club_name=u'Groruddalen Friidrettsklubb'
    elif club_code == (u'GRUE'):
       club_name=u'Grue Idrettslag'
    elif club_code == (u'GTI'):
       club_name=u'GTI Friidrettsklubb'
    elif club_code == (u'GUI'):
       club_name=u'Gui Sportsklubb - Friidrett'
    elif club_code == (u'GUL'):
       club_name=u'Gulset Idrettsforening'
    elif club_code == (u'HAB'):
       club_name=u'HAB IL'
    elif club_code == (u'HADE'):
       club_name=u'Hadeland Friidrettsklubb'
    elif club_code == (u'HAGA'):
       club_name=u'Haga Idrettsforening '
    elif club_code == (u'HAL'):
       club_name=u'Halden Idrettslag'
    elif club_code == (u'HALMO'):
       club_name=u'Halmsås & Omegn Skilag'
    elif club_code == (u'HALSA'):
       club_name=u'Halsa Idrettslag'
    elif club_code == (u'HIL'):
       club_name=u'Hamar Idrettslag'
    elif club_code == (u'HANNEV'):
       club_name=u'Hannevikas Idrettslag'
    elif club_code == (u'HARDB'):
       club_name=u'Hardbagg Idrettslag'
    elif club_code == (u'HAREI'):
       club_name=u'Hareid Idrettslag'
    elif club_code == (u'HARE'):
       club_name=u'Harestua Idrettslag'
    elif club_code == (u'HATT'):
       club_name=u'Hattfjelldal Idrettslag'
    elif club_code == (u'HAUGN'):
       club_name=u'Haugen Idrettslag'
    elif club_code == (u'HAUGR'):
       club_name=u'Haugerud Idrettsforening'
    elif club_code == (u'HAUGF'):
       club_name=u'Haugesund Idrettslag Friidrett'
    elif club_code == (u'HAUGT'):
       club_name=u'Haugesund Triathlon Klubb'
    elif club_code == (u'HAV'):
       club_name=u'Havørn Allianseidrettslag'
    elif club_code == (u'HEGGF'):
       club_name=u'Heggedal Friidrettsklubb'
    elif club_code == (u'HEGGI'):
       club_name=u'Heggedal Idrettslag'
    elif club_code == (u'HELLU'):
       club_name=u'Hell Ultraløperklubb'
    elif club_code == (u'HEM'):
       club_name=u'Heming Idrettslaget'
    elif club_code == (u'HENN'):
       club_name=u'Henning I L'
    elif club_code == (u'HERA'):
       club_name=u'Herand Idrettslag'
    elif club_code == (u'HERK'):
       club_name=u'Herkules Friidrett'
    elif club_code == (u'HERY'):
       club_name=u'Herøy Idrettslag'
    elif club_code == (u'HIN'):
       club_name=u'Hinna Friidrett'
    elif club_code == (u'HITF'):
       club_name=u'Hitra Friidrettsklubb'
    elif club_code == (u'HITL'):
       club_name=u'Hitra Løpeklubb'
    elif club_code == (u'HOB'):
       club_name=u'Hobøl Idrettslag'
    elif club_code == (u'HOF'):
       club_name=u'Hof Idrettslag'
    elif club_code == (u'HOL'):
       club_name=u'Hol Idrettslag'
    elif club_code == (u'HOLMS'):
       club_name=u'Holmemstranda Idrettslag'
    elif club_code == (u'HOLUM'):
       club_name=u'Holum Idrettslag'
    elif club_code == (u'HMLV'):
       club_name=u'Hommelvik IL'
    elif club_code == (u'HOPE'):
       club_name=u'Hope Idrettslag'
    elif club_code == (u'HORNI'):
       club_name=u'Hornindal Idrettslag'
    elif club_code == (u'HORFR'):
       club_name=u'Horten Friidrettsklubb'
    elif club_code == (u'HUG'):
       club_name=u'Huglo Idrettslag'
    elif club_code == (u'HURD'):
       club_name=u'Hurdal Idrettslag'
    elif club_code == (u'HVAM'):
       club_name=u'Hvam Idrettslag'
    elif club_code == (u'HVFO'):
       club_name=u'Hvittingfoss Idrettslag'
    elif club_code == (u'HYEN'):
       club_name=u'Hyen Idrettslag'
    elif club_code == (u'HYLLS'):
       club_name=u'Hyllestad Idrettslag'
    elif club_code == (u'HSI'):
       club_name=u'Høybråten og Stovner IL'
    elif club_code == (u'HDMO'):
       club_name=u'Høydalsmo Idrottslag'
    elif club_code == (u'FJELLO'):
       club_name=u'I.l Fjellørnen'
    elif club_code == (u'FRAMS'):
       club_name=u'I.L. Framsteg'
    elif club_code == (u'NORSA'):
       club_name=u'Norna-Salhus IL'
    elif club_code == (u'NYBR'):
       club_name=u'I.L. Nybrott'
    elif club_code == (u'IDD'):
       club_name=u'Idd Sportsklubb'
    elif club_code == (u'BIRK'):
       club_name=u'Idrettsforeningen Birkebeineren'
    elif club_code == (u'FRAM'):
       club_name=u'Idrettsforeningen Fram'
    elif club_code == (u'HELLA'):
       club_name=u'IF Hellas'
    elif club_code == (u'NJAAL'):
       club_name=u'Idrettsforeningen Njaal'
    elif club_code == (u'STUR'):
       club_name=u'Sturla IF'
    elif club_code == (u'IFORN'):
       club_name=u'Idrettsforeningen Ørn'
    elif club_code == (u'BJARG'):
       club_name=u'Idrettslaget Bjarg'
    elif club_code == (u'ILBJ'):
       club_name=u'Idrettslaget Bjørn'
    elif club_code == (u'DLBR'):
       club_name=u'Idrettslaget Dalebrand'
    elif club_code == (u'DYREV'):
       club_name=u'Idrettslaget Dyre Vaa'
    elif club_code == (u'EXPR'):
       club_name=u'Idrettslaget Express'
    elif club_code == (u'FORSK'):
       club_name=u'Idrettslaget Forsøk'
    elif club_code == (u'FRI'):
       club_name=u'Idrettslaget Fri'
    elif club_code == (u'GNE'):
       club_name=u'Idrettslaget Gneist'
    elif club_code == (u'HOLE'):
       club_name=u'Idrettslaget Holeværingen'
    elif club_code == (u'BULT'):
       club_name=u'Idrettslaget I Bondeungdomslaget I Tromsø'
    elif club_code == (u'ILAR'):
       club_name=u'Idrettslaget Ilar'
    elif club_code == (u'IVRIG'):
       club_name=u'Idrettslaget Ivrig'
    elif club_code == (u'JARD'):
       club_name=u'Idrettslaget Jardar'
    elif club_code == (u'JUT'):
       club_name=u'Idrettslaget Jutul'
    elif club_code == (u'ILROS'):
       club_name=u'Idrettslaget Ros'
    elif club_code == (u'RUNAR'):
       club_name=u'IL Runar'
    elif club_code == (u'ILSAN'):
       club_name=u'Idrettslaget Sand'
    elif club_code == (u'SANDV'):
       club_name=u'IL Sandvin'
    elif club_code == (u'SKADE'):
       club_name=u'Idrettslaget Skade'
    elif club_code == (u'SKJA'):
       club_name=u'Idrettslaget Skjalg'
    elif club_code == (u'SYR'):
       club_name=u'Idrettslaget Syril'
    elif club_code == (u'TRY'):
       club_name=u'Idrettslaget Trysilgutten'
    elif club_code == (u'GULA'):
       club_name=u'IL Gular'
    elif club_code == (u'ILIBUL'):
       club_name=u'IDROTTSLAGET I BUL'
    elif club_code == (u'ILBUL'):
       club_name=u'IDROTTSLAGET I BUL'
    elif club_code == (u'JOT'):
       club_name=u'Idrottslaget Jotun'
    elif club_code == (u'IDUN'):
       club_name=u'Idun Idrettslag'
    elif club_code == (u'EIKKV'):
       club_name=u'If Eiker Kvikk'
    elif club_code == (u'KAVE'):
       club_name=u'IF Kamp/Vestheim'
    elif club_code == (u'KLYP'):
       club_name=u'If Klypetussen'
    elif club_code == (u'GRANE'):
       club_name=u'Ik Grane Arendal Friidrett'
    elif club_code == (u'HIND'):
       club_name=u'IK Hind'
    elif club_code == (u'IKORN'):
       club_name=u'Ikornnes Idrettslag'
    elif club_code == (u'AASG'):
       club_name=u'IL Aasguten'
    elif club_code == (u'ALVI'):
       club_name=u'IL Alvidra'
    elif club_code == (u'ILBEV'):
       club_name=u'IL Bever`n'
    elif club_code == (u'BRODD'):
       club_name=u'IL Brodd'
    elif club_code == (u'FLV'):
       club_name=u'IL Flåværingen'
    elif club_code == (u'GRY'):
       club_name=u'IL Gry'
    elif club_code == (u'ILNOR'):
       club_name=u'IL Norodd'
    elif club_code == (u'PIO'):
       club_name=u'IL Pioner Friidrett'
    elif club_code == (u'POL'):
       club_name=u'IL Polarstjernen'
    elif club_code == (u'SAMH'):
       club_name=u'IL Samhald'
    elif club_code == (u'SANT'):
       club_name=u'IL Santor'
    elif club_code == (u'STKAM'):
       club_name=u'IL Stålkameratene'
    elif club_code == (u'TRIUM'):
       club_name=u'IL Triumf'
    elif club_code == (u'VIND'):
       club_name=u'Il Vindbjart'
    elif club_code == (u'VING'):
       club_name=u'IL Vinger'
    elif club_code == (u'INDRY'):
       club_name=u'Inderøy Idrettslag'
    elif club_code == (u'INN'):
       club_name=u'Innstranda IL'
    elif club_code == (u'INSTA'):
       club_name=u'International School of Stavanger'
    elif club_code == (u'ISFJO'):
       club_name=u'Isfjorden Idrettslag'
    elif club_code == (u'JOND'):
       club_name=u'Jondalen Idrettslag'
    elif club_code == (u'JVTN'):
       club_name=u'Jægervatnet Idrettslag'
    elif club_code == (u'JIL'):
       club_name=u'Jøa Idrettslag'
    elif club_code == (u'JLSTE'):
       club_name=u'Jølster Idrettslag'
    elif club_code == (u'KAUP'):
       club_name=u'Kaupanger Idrettslag'
    elif club_code == (u'KFUM'):
       club_name=u'Kfum-kameratene Oslo'
    elif club_code == (u'KJ'):
       club_name=u'Kjelsås Idrettslag'
    elif club_code == (u'KLPP'):
       club_name=u'Klepp Idrettslag'
    elif club_code == (u'KLK'):
       club_name=u'Klæbu Løpeklubb'
    elif club_code == (u'KLIL'):
       club_name=u'Kløfta Idrettslag'
    elif club_code == (u'KLBK'):
       club_name=u'Kolbukameratene I L'
    elif club_code == (u'KOLL'):
       club_name=u'Koll Idrettslaget'
    elif club_code == (u'KLVIL'):
       club_name=u'Kolvereid Idrettslag'
    elif club_code == (u'KNGSB'):
       club_name=u'Kongsberg Idrettsforening'
    elif club_code == (u'KNGSV'):
       club_name=u'Kongsvinger IL Friidrett'
    elif club_code == (u'KONN'):
       club_name = u'Konnerud IL'
    elif club_code == (u'KOP'):
       club_name=u'Kopervik Idrettslag'
    elif club_code == (u'KORG'):
       club_name=u'Korgen Idrettslag'
    elif club_code == (u'KRAG'):
       club_name=u'Kragerø IF Friidrett'
    elif club_code == (u'KRAAK'):
       club_name=u'Kråkerøy Idrettslag'
    elif club_code == (u'KRSTD'):
       club_name=u'Kråkstad Idrettslag'
    elif club_code == (u'KRL'):
       club_name=u'Kristiansand Løpeklubb'
    elif club_code == (u'KIF'):
       club_name=u'Kristiansands IF'
    elif club_code == (u'KRHER'):
       club_name=u'Krødsherad Idrettslag'
    elif club_code == (u'KVINES'):
       club_name=u'Kvinesdal Idrettslag'
    elif club_code == (u'KVFJ'):
       club_name=u'Kvæfjord Idrettslag'
    elif club_code == (u'KYRK'):
       club_name=u'Kyrksæterøra Idrettslag Kil'
    elif club_code == (u'LAKS'):
       club_name=u'Laksevåg TIL'
    elif club_code == (u'LALM'):
       club_name=u'Lalm Idrettslag'
    elif club_code == (u'LAM'):
       club_name=u'Lambertseter IF'
    elif club_code == (u'LANGS'):
       club_name=u'Langesund Sykle- og triathlonklubb'
    elif club_code == (u'LNKEIL'):
       club_name=u'Lånke Idrettslag'
    elif club_code == (u'LRVK'):
       club_name=u'Larvik Turn & Idrettsforening'
    elif club_code == (u'LEINS'):
       club_name=u'Leinstrand Idrettslag'
    elif club_code == (u'LENA'):
       club_name=u'Lena Idrettsforening'
    elif club_code == (u'LIERN'):
       club_name=u'Lierne Idrettslag'
    elif club_code == (u'LIF'):
       club_name=u'Lillehammer IF'
    elif club_code == (u'LILLS'):
       club_name=u'Lillesand Idrettslag'
    elif club_code == (u'LISTA'):
       club_name=u'Lista Idrettslag'
    elif club_code == (u'LODD'):
       club_name=u'Loddefjord IL'
    elif club_code == (u'LFTR'):
       club_name=u'Lofoten Triatlonklubb'
    elif club_code == (u'LOM'):
       club_name=u'Lom Idrettslag'
    elif club_code == (u'LUND'):
       club_name=u'Lundamo Idrettslag'
    elif club_code == (u'LUNDH'):
       club_name=u'Lundehøgda IL'
    elif club_code == (u'LUST'):
       club_name=u'Luster Idrettslag'
    elif club_code == (u'LYE'):
       club_name=u'Lye Idrettslag'
    elif club_code == (u'LYN'):
       club_name=u'Lyn Ski'
    elif club_code == (u'LNGD'):
       club_name=u'Lyngdal Idrettslag'
    elif club_code == (u'LYKA'):
       club_name=u'Lyngen/ Karnes Il'
    elif club_code == (u'LYNGO'):
       club_name=u'Lyngstad og Omegn Idrettslag'
    elif club_code == (u'LRSKG'):
       club_name=u'Lørenskog Friidrettslag'
    elif club_code == (u'LFK'):
       club_name=u'Løten Friidrett'
    elif club_code == (u'LTN'):
       club_name=u'Løten Friidrett 2'
    elif club_code == (u'MALM'):
       club_name=u'Malm IL'
    elif club_code == (u'MLSEL'):
       club_name=u'Målselv Idrettslag'
    elif club_code == (u'MALV'):
       club_name=u'Malvik Idrettslag'
    elif club_code == (u'MAAL'):
       club_name=u'Måløy Idrettslag Hovedstyre'
    elif club_code == (u'MAHA'):
       club_name=u'Mandal & Halse I.l.'
    elif club_code == (u'MNDL'):
       club_name=u'Måndalen Idrettslag'
    elif club_code == (u'MABY'):
       club_name=u'Markabygda Idrettslag'
    elif club_code == (u'MARKA'):
       club_name=u'Markane IL'
    elif club_code == (u'MARNA'):
       club_name=u'Marnardal Idrettslag'
    elif club_code == (u'MEDKI'):
       club_name=u'Medkila Skilag'
    elif club_code == (u'MELD'):
       club_name=u'Meldal Idrettslag'
    elif club_code == (u'MELHU'):
       club_name=u'Melhus Idrettslag'
    elif club_code == (u'MDSND'):
       club_name=u'Midsund Idrettslag'
    elif club_code == (u'MJSD'):
       club_name=u'Mjøsdalen IL'
    elif club_code == (u'MOD'):
       club_name=u'Modum Friidrettsklubb'
    elif club_code == (u'MOELV'):
       club_name=u'Moelven IL'
    elif club_code == (u'MOI'):
       club_name=u'Moi Idrettslag'
    elif club_code == (u'MOLDE'):
       club_name=u'Molde og Omegn Idrettsforening'
    elif club_code == (u'OLYMP'):
       club_name=u'Molde Olymp'
    elif club_code == (u'MOITU'):
       club_name=u'Moltustranda Idrettslag'
    elif club_code == (u'MOSJ'):
       club_name=u'Mosjøen Friidrettsklubb'
    elif club_code == (u'MOSS'):
       club_name=u'Moss Idrettslag'
    elif club_code == (u'MOSV'):
       club_name=u'Mosvik Idrettslag'
    elif club_code == (u'MUIL'):
       club_name=u'MUIL - Mefjordvær Ungdoms- og Idrettslag'
    elif club_code == (u'NAML'):
       club_name=u'Namdal løpeklubb'
    elif club_code == (u'NAMDA'):
       club_name=u'Namdalseid Idrettslag'
    elif club_code == (u'NAMSE'):
       club_name=u'Namsen Fif'
    elif club_code == (u'NANN'):
       club_name=u'Nannestad Idrettslag'
    elif club_code == (u'NAR'):
       club_name=u'Narvik Idrettslag'
    elif club_code == (u'NESB'):
       club_name=u'Nesbyen Idrettslag'
    elif club_code == (u'NESO'):
       club_name=u'Nesodden IF'
    elif club_code == (u'NES'):
       club_name=u'Nesøya Idrettslag'
    elif club_code == (u'NID'):
       club_name=u'Nidelv Idrettslag'
    elif club_code == (u'NISS'):
       club_name=u'Nissedal Idrettslag'
    elif club_code == (u'NITT'):
       club_name=u'Nittedal Idrettslag'
    elif club_code == (u'NRDKJ'):
       club_name=u'Nordkjosbotn Idrettslag'
    elif club_code == (u'NEIDS'):
       club_name=u'Nordre Eidsvoll Idrettslag'
    elif club_code == (u'NFJEL'):
       club_name=u'Nordre Fjell Friidrett'
    elif club_code == (u'NLAND'):
       club_name=u'Nordre Land Idrettslag'
    elif club_code == (u'NTRY'):
       club_name=u'Nordre Trysil IL'
    elif club_code == (u'NORIL'):
       club_name=u'Nordøy Idrettslag'
    elif club_code == (u'NORR'):
       club_name=u'Norrøna IL'
    elif club_code == (u'NRUN'):
       club_name=u'Northern Runners'
    elif club_code == (u'NTNUI'):
       club_name=u'NTNUI - Norges Teknisk-Naturvitenskapelige Universitets Idrettsforening'
    elif club_code == (u'NYSK'):
       club_name=u'Nydalens Skiklub'
    elif club_code == (u'NYKIR'):
       club_name=u'Nykirke Idrettsforening'
    elif club_code == (u'NTTRY'):
       club_name=u'Nøtterøy Idrettsforening'
    elif club_code == (u'ODDA'):
       club_name=u'Odda Idrettslag'
    elif club_code == (u'OGND'):
       club_name=u'Ogndal Idrettslag Hovedlaget'
    elif club_code == (u'OLD'):
       club_name=u'Olden Idrettslag'
    elif club_code == (u'OLDA'):
       club_name=u'Olderdalen Idrettsklubb'
    elif club_code == (u'OPPD'):
       club_name=u'Oppdal IL Hovedlaget'
    elif club_code == (u'OPP'):
       club_name=u'Oppegård Idrettslag'
    elif club_code == (u'OPSL'):
       club_name=u'Oppsal Idrettsforening'
    elif club_code == (u'OPST'):
       club_name=u'Oppstad Idrettslag'
    elif club_code == (u'OPPST'):
       club_name=u'Oppstad Idrettslag 2'
    elif club_code == (u'OPSTR'):
       club_name=u'Oppstryn Idrettslag'
    elif club_code == (u'OPPT'):
       club_name=u'Opptur Motbakkeklubb'
    elif club_code == (u'ORKA'):
       club_name=u'Orkanger Idrettsforening'
    elif club_code == (u'ORKD'):
       club_name=u'Orkdal Idrettslag'
    elif club_code == (u'ORRE'):
       club_name=u'Orre Idrettslag'
    elif club_code == (u'OS'):
       club_name=u'Os Idrettslag'
    elif club_code == (u'OSTU'):
       club_name=u'Os Turnforening'
    elif club_code == (u'FRII'):
       club_name=u'OSI Friidrett'
    elif club_code == (u'POLIT'):
       club_name=u'Oslo Politis Idrettslag'
    elif club_code == (u'OSI'):
       club_name=u'Oslostudentenes Idrettsklubb'
    elif club_code == (u'OST'):
       club_name=u'Osterøy Idrottslag'
    elif club_code == (u'OTRA'):
       club_name=u'Otra IL'
    elif club_code == (u'OTTE'):
       club_name=u'Ottestad Idrettslag'
    elif club_code == (u'OTKS'):
       club_name=u'Ottestad Kast og Styrkeløft'
    elif club_code == (u'OVRH'):
       club_name=u'Overhalla Idrettslag'
    elif club_code == (u'PORS'):
       club_name=u'Porsanger Idrettslag'
    elif club_code == (u'RANA'):
       club_name=u'Rana Friidrettsklubb'
    elif club_code == (u'RAN'):
       club_name=u'Ranheim IL'
    elif club_code == (u'RAU'):
       club_name=u'Raufoss IL Friidrett'
    elif club_code == (u'RAUM'):
       club_name=u'Raumnes & Årnes Idrettslag'
    elif club_code == (u'RE'):
       club_name=u'Re Friidrettsklubb'
    elif club_code == (u'READY'):
       club_name=u'Ready Idrettsforeningen'
    elif club_code == (u'RENA'):
       club_name=u'Rena Idrettslag'
    elif club_code == (u'RENDA'):
       club_name=u'Rendalen Idrettslag'
    elif club_code == (u'RENB'):
       club_name=u'Rennebu Idrettslag'
    elif club_code == (u'RIND'):
       club_name=u'Rindal Idrettslag'
    elif club_code == (u'RING'):
       club_name=u'Ringerike Friidrettsklubb'
    elif club_code == (u'RIS'):
       club_name=u'Risør Idrettslag'
    elif club_code == (u'RJU'):
       club_name=u'Rjukan Idrettslag'
    elif club_code == (u'ROGNE'):
       club_name=u'Rogne Idrettslag'
    elif club_code == (u'ROMFR'):
       club_name=u'Romerike Friidrett'
    elif club_code == (u'ROMUL'):
       club_name=u'Romerike Ultraløperklubb'
    elif club_code == (u'ROMRA'):
       club_name=u'Romsdal Randoneklubb'
    elif club_code == (u'ROSEN'):
       club_name=u'Rosendal Turnlag'
    elif club_code == (u'ROYAL'):
       club_name=u'Royal Sport'
    elif club_code == (u'RUS'):
       club_name=u'Rustad Idrettslag'
    elif club_code == (u'RYGGE'):
       club_name=u'Rygge Idrettslag'
    elif club_code == (u'RIL'):
       club_name=u'Røa Allianseidrettslag'
    elif club_code == (u'RDIL'):
       club_name=u'Røldal Idrettslag'
    elif club_code == (u'ROSIL'):
       club_name=u'Røros Idrettslag'
    elif club_code == (u'RKEN'):
       club_name=u'Røyken UIL'
    elif club_code == (u'SALA'):
       club_name=u'Salangen IF Friidrett'
    elif club_code == (u'SAMN'):
       club_name=u'Samnanger Idrettslag'
    elif club_code == (u'SANTU'):
       club_name=u'Sandane Turn og Idrettslag'
    elif club_code == (u'STIF'):
       club_name=u'SANDEFJORD TURN & IDRETTSFORENING'
    elif club_code == (u'SAND'):
       club_name=u'Sandnes IL'
    elif club_code == (u'SNDI'):
       club_name=u'Sandnes Idrettslag 2'
    elif club_code == (u'SNDSJ'):
       club_name=u'Sandnessjøen Idrettslag'
    elif club_code == (u'SARP'):
       club_name=u'Sarpsborg IL'
    elif club_code == (u'SAUD'):
       club_name=u'Sauda Idrettslag'
    elif club_code == (u'SAUL'):
       club_name=u'Sauland Idrettslag'
    elif club_code == (u'SELB'):
       club_name=u'Selbu IL'
    elif club_code == (u'SELJE'):
       club_name=u'Selje Idrettslag'
    elif club_code == (u'SELJO'):
       club_name=u'Seljord Idrettslag'
    elif club_code == (u'SELS'):
       club_name=u'Selsbakk Idrettsforening'
    elif club_code == (u'SEM'):
       club_name=u'Sem Idrettsforening'
    elif club_code == (u'SIGFR'):
       club_name=u'Sigdal Friidrettsklubb'
    elif club_code == (u'SIGSK'):
       club_name=u'Sigdals Skiklub'
    elif club_code == (u'SILJ'):
       club_name=u'Siljan Idrettslag'
    elif club_code == (u'SIRMA'):
       club_name=u'Sirma Il'
    elif club_code == (u'SJET'):
       club_name=u'Sjetne Idrettslag'
    elif club_code == (u'VEDA'):
       club_name=u'Sk Vedavåg Karmøy'
    elif club_code == (u'VID'):
       club_name=u'SK Vidar'
    elif club_code == (u'SKAGE'):
       club_name=u'Skagerrak Sportsklubb'
    elif club_code == (u'SKLA'):
       club_name=u'Skåla Idrettslag'
    elif club_code == (u'SKRPH'):
       club_name=u'Skarphedin IL'
    elif club_code == (u'SKAU'):
       club_name=u'Skaubygda Il'
    elif club_code == (u'SKAUN'):
       club_name=u'Skaun Idrettslag'
    elif club_code == (u'SKI'):
       club_name=u'Ski IL Friidrett'
    elif club_code == (u'SKJK'):
       club_name=u'Skjåk IL'
    elif club_code == (u'SKJO'):
       club_name=u'Skjoldar Il'
    elif club_code == (u'SKOGN'):
       club_name=u'Skogn Idrettslag'
    elif club_code == (u'SKO'):
       club_name=u'Skotterud Idrettslag'
    elif club_code == (u'SKREIA'):
       club_name=u'Skreia Idrettslag'
    elif club_code == (u'SNSA'):
       club_name=u'Snåsa Idrettslag'
    elif club_code == (u'SNGG'):
       club_name=u'Snøgg Friidrett'
    elif club_code == (u'SIL'):
       club_name=u'Sogndal Idrettslag'
    elif club_code == (u'SOKND'):
       club_name=u'Sokndal Friidrettsklubb'
    elif club_code == (u'SOLA'):
       club_name=u'Sola Friidrett'
    elif club_code == (u'SOLN'):
       club_name=u'Solnut IL'
    elif club_code == (u'SORTL'):
       club_name=u'Sortland Friidrettsklubb'
    elif club_code == (u'SOT'):
       club_name=u'Sotra Sportsklubb'
    elif club_code == (u'SPILL'):
       club_name=u'Spillum Idrettslag'
    elif club_code == (u'SPRD'):
       club_name=u'Spiridon Langløperlag'
    elif club_code == (u'SPJVK'):
       club_name=u'Spjelkavik og Omegn Friidrettsklubb'
    elif club_code == (u'KRAFT'):
       club_name=u'Sportsklubben Kraft'
    elif club_code == (u'nan'):
       club_name=u'Sportsklubben Rye'
    elif club_code == (u'RYE'):
       club_name=u'Sportsklubben Rye 2'
    elif club_code == (u'VIDAR'):
       club_name=u'Sportsklubben Vidar'
    elif club_code == (u'SPYDE'):
       club_name=u'Spydeberg IL'
    elif club_code == (u'STJIL'):
       club_name=u'Staal Jørpeland IL'
    elif club_code == (u'STAD'):
       club_name=u'Stadsbygd IL'
    elif club_code == (u'STRHE'):
       club_name=u'Stårheim IL'
    elif club_code == (u'STDIF'):
       club_name=u'Stavanger Døve-Idrettsforening'
    elif club_code == (u'STAVA'):
       club_name=u'Stavanger Friidrettsklubb'
    elif club_code == (u'STAVIF'):
       club_name=u'Stavanger Idrettsforening Allianseidrettslag - Friidrett'
    elif club_code == (u'STEGA'):
       club_name=u'Stegaberg Idrettslag'
    elif club_code == (u'STEIN'):
       club_name=u'Stein Friidrettsklubb'
    elif club_code == (u'STEKJ'):
       club_name=u'Steinkjer Friidrettsklubb'
    elif club_code == (u'STSK'):
       club_name=u'Stettevik Sportsklubb'
    elif club_code == (u'STJF'):
       club_name=u'Stjørdal Fridrettsklubb'
    elif club_code == (u'STJP'):
       club_name=u'Stjørdal Paraidrettslag'
    elif club_code == (u'STJB'):
       club_name=u'Stjørdals-Blink IL'
    elif club_code == (u'STOKK'):
       club_name=u'Stokke Idrettslag'
    elif club_code == (u'STOKM'):
       club_name=u'Stokmarknes Idrettslag'
    elif club_code == (u'STO'):
       club_name=u'Stord Idrettslag'
    elif club_code == (u'STOFJ'):
       club_name=u'Storfjord idrettslag'
    elif club_code == (u'STRIL'):
       club_name=u'Stranda Idrottslag'
    elif club_code == (u'STRAB'):
       club_name=u'Strandebarm Idrettslag'
    elif club_code == (u'STRA'):
       club_name=u'Stranden Idrettslag'
    elif club_code == (u'STRAU'):
       club_name=u'Straumsnes Idrettslag'
    elif club_code == (u'STRI'):
       club_name=u'Strindheim Idrettslag'
    elif club_code == (u'STRY'):
       club_name=u'Stryn Turn og Idrettslag'
    elif club_code == (u'STREN'):
       club_name=u'Støren Sportsklubb'
    elif club_code == (u'SUNND'):
       club_name=u'Sunndal IL Friidrett'
    elif club_code == (u'SURN'):
       club_name=u'Surnadal Idrettslag'
    elif club_code == (u'SVTU'):
       club_name=u'Svalbard Turn Idrettslag'
    elif club_code == (u'SVARS'):
       club_name=u'Svarstad Idrettslag'
    elif club_code == (u'SVEIO'):
       club_name=u'Sveio Idrettslag'
    elif club_code == (u'SVEL'):
       club_name=u'Svelgen Turn og Idrettsforening'
    elif club_code == (u'SVINT'):
       club_name=u'Svint IL'
    elif club_code == (u'SVORK'):
       club_name=u'SVORKMO N.O.I.'
    elif club_code == (u'SYKK'):
       club_name=u'Sykkylven Idrottslag'
    elif club_code == (u'SYLL'):
       club_name=u'Sylling Idrettsforening'
    elif club_code == (u'SDAL'):
       club_name=u'Sædalen Idrettslag'
    elif club_code == (u'GRAA'):
       club_name=u'Sætre Idrætsforening Graabein'
    elif club_code == (u'STIL'):
       club_name=u'Søfteland Turn & Idrettslag'
    elif club_code == (u'SGNE'):
       club_name=u'Søgne Idrettslag'
    elif club_code == (u'SMNA'):
       club_name=u'Sømna Idrettslag'
    elif club_code == (u'SNDLA'):
       club_name=u'Søndre Land IL Friidrett'
    elif club_code == (u'SAAL'):
       club_name=u'Søre Ål Idrettslag'
    elif club_code == (u'SRILD'):
       club_name=u'Sørild Fridrettsklubb'
    elif club_code == (u'SRKDL'):
       club_name=u'Sørkedalens Idrettsforening'
    elif club_code == (u'HOVD'):
       club_name=u'T I L Hovding'
    elif club_code == (u'TAMBA'):
       club_name=u'IL Tambarskjelvar'
    elif club_code == (u'TAMSAN'):
       club_name=u'Tamil Sangam IL'
    elif club_code == (u'TING'):
       club_name=u'Tingvoll Friidrettsklubb'
    elif club_code == (u'TIST'):
        club_name = u'Tistedalen FL'
    elif club_code == (u'TJALV'):
       club_name=u'IK Tjalve'
    elif club_code == (u'TJØLL'):
       club_name=u'Tjølling Idrettsforening'
    elif club_code == (u'TJI'):
       club_name=u'Tjøme Idrettslag'
    elif club_code == (u'TJL'):
       club_name=u'Tjøme Løpeklubb'
    elif club_code == (u'TOL'):
       club_name=u'Tolga Idrettslag'
    elif club_code == (u'TOMR'):
       club_name=u'Tomrefjord Idrettslag'
    elif club_code == (u'TORO'):
       club_name=u'Torodd IF'
    elif club_code == (u'TORVI'):
       club_name=u'Torvikbukt Idrettslag'
    elif club_code == (u'TREU'):
       club_name=u'Treungen Idrettslag'
    elif club_code == (u'TRIO'):
       club_name=u'Trio idrettslag'
    elif club_code == (u'TRF'):
       club_name=u'Tromsø Friidrettsklubb'
    elif club_code == (u'TRL'):
       club_name=u'Tromsø Løpeklubb'
    elif club_code == (u'TRS'):
       club_name=u'Tromsø Svømmeklubb'
    elif club_code == (u'TROO'):
       club_name=u'Trondheim & Omegn Sportsklubb'
    elif club_code == (u'TROF'):
       club_name=u'Trondheim Friidrett'
    elif club_code == (u'TSK'):
       club_name=u'Trøgstad Skiklubb'
    elif club_code == (u'TUIL'):
       club_name=u'TUIL Tromsdalen Friidrett'
    elif club_code == (u'TVEDE'):
       club_name=u'Tvedestrand Turn & Idrettsforening'
    elif club_code == (u'TYR'):
       club_name=u'Tyrving IL'
    elif club_code == (u'TNSBF'):
       club_name=u'Tønsberg Friidrettsklubb'
    elif club_code == (u'TRBIL'):
       club_name=u'Tørvikbygd Idrettslag'
    elif club_code == (u'TYEN'):
       club_name=u'Tøyen Sportsklubb'
    elif club_code == (u'ULLK'):
       club_name=u'Ullensaker/Kisa IL Friidrett'
    elif club_code == (u'ULKI'):
       club_name=u'Ullensaker/Kisa IL Friidrett 2'
    elif club_code == (u'UND'):
       club_name=u'Undheim Idrettslag'
    elif club_code == (u'URFRI'):
       club_name=u'Urædd Friidrett'
    elif club_code == (u'UTL'):
       club_name=u'Utleira Idrettslag'
    elif club_code == (u'VAAL'):
       club_name=u'Vaaler Idrettsforening'
    elif club_code == (u'VA'):
       club_name=u'Vadsø Atletklubb'
    elif club_code == (u'VTF'):
       club_name=u'Vadsø Turnforening (Vtf)'
    elif club_code == (u'VGAA'):
       club_name=u'Vågå Idrettslag'
    elif club_code == (u'VIL'):
       club_name=u'Vågstranda Idrettslag'
    elif club_code == (u'VALK'):
       club_name=u'Valkyrien Idrettslag'
    elif club_code == (u'VALL'):
       club_name=u'Valldal Idrettslag'
    elif club_code == (u'VAL'):
       club_name=u'Vallset IL'
    elif club_code == (u'VAR'):
       club_name=u'Varegg Fleridrett'
    elif club_code == (u'VARH'):
       club_name=u'Varhaug Idrettslag'
    elif club_code == (u'VART'):
       club_name=u'Varteig Idrettslag'
    elif club_code == (u'VEG'):
       club_name=u'Vegårshei Idrettslag'
    elif club_code == (u'VELD'):
       club_name=u'Veldre Friidrett'
    elif club_code == (u'VELL'):
       club_name=u'Velledalen Idrettslag'
    elif club_code == (u'VERD'):
       club_name=u'Verdal Friidrettsklubb'
    elif club_code == (u'VESTB'):
       club_name=u'Vestby Idrettslag'
    elif club_code == (u'VESTB'):
       club_name=u'Vestby Idrettslag'
    elif club_code == (u'VESTF'):
       club_name=u'Vestfossen Idrettsforening'
    elif club_code == (u'VSPON'):
       club_name=u'Vestre Spone IF'
    elif club_code == (u'VIKIL'):
       club_name=u'Vik Idrettslag'
    elif club_code == (u'VIKAN'):
       club_name=u'Vikane IL'
    elif club_code == (u'VIK'):
       club_name=u'TIF Viking'
    elif club_code == (u'VIKSD'):
       club_name=u'Viksdalen Idrettslag'
    elif club_code == (u'VILJ'):
       club_name=u'Viljar IL'
    elif club_code == (u'VNDIL'):
       club_name=u'Vind Idrettslag'
    elif club_code == (u'VINDA'):
       club_name=u'Vindafjord Idrettslag'
    elif club_code == (u'VINJE'):
       club_name=u'Vinje Idrottslag'
    elif club_code == (u'VOLL'):
       club_name=u'Vollan Idrettsklubb'
    elif club_code == (u'VOSS'):
       club_name=u'Voss Idrottslag'
    elif club_code == (u'YTTER'):
       club_name=u'Ytterøy Idrettslag'
    elif club_code == (u'ORJIL'):
       club_name=u'Ørje Idrettslag'
    elif club_code == (u'ORSTA'):
       club_name=u'Ørsta Idrettslag'
    elif club_code == (u'OMARSJ'):
       club_name=u'Østmarka Marsjklubb'
    elif club_code == (u'OTRET'):
       club_name=u'Øyer/Tretten Idrettsforening'
    elif club_code == (u'OSLID'):
       club_name=u'Øystre Slidre Idrettslag'
    else:
       club_name=club_code

    return club_name
def get_organiser_name(key):
    organisers = { "1376e260-82f7-4bf6-9da6-064fd76c6d87" : "IL Koll", 
                   "575153d6-7f1b-4795-9276-5f8d57414944" : 'IK Tjalve'
            }
    return organisers.get(key, key)
#---------------------------------------
if len(sys.argv) < 2:
   sys.exit("Usage: %s <url>" % sys.argv[0])
   
url = sys.argv[1]
print(url)

r=requests.get(url+'json')
j = json.loads(r.text)
#with open('downloads.json', 'r') as f: 
#    j = json.load(f)

#print(type(j))
#print(j.keys())
#print(j['date'])
d  = j['date']
d2 = j['finishDate']
isodateformat = "%Y-%m-%d"
date0 = datetime.datetime.strptime(d, isodateformat)
#print(d, date)
date1 = datetime.datetime.strptime(d2, isodateformat)
#bdate = datetime.datetime.strptime('2005-06-24', isodateformat)
#print(get_category(bdate,date,'F'))
dates = []
d = date0
while d <= date1:
    dates.append(d)
    d += datetime.timedelta(days=1)
#print(dates)


if 'nameLocal' in j.keys():
    meetname = j['nameLocal']
meetname = j['fullName']

slug = j['slug']
outdoors = 'J'
if j.get('venue') == None: 
    venue = ''
else:
    venue = j['venue']['formalName']
    if j['venue']['indoor'] == 'true':
        outdoors = 'N'
if j['type']=="INDOOR":
    outdoors='N'


#print(meetname, venue)
organiser_name =  j['organiser']['name']
#organiser_name = get_organiser_name(organiser_key)


ignore_bibs = []
competitors = {}
#print(j['competitors'][0])
for c in j['competitors']:
#   print(c.keys())
    fn = ''; ln = ''; dob= ''; t=''
    bib  = c['competitorId']
    if 'firstName' in c.keys():
        fn   = c['firstName']
    if 'lastName' in c.keys():
        ln   = c['lastName']
    if 'dateOfBirth' in c.keys():
        d    = c['dateOfBirth']
        dob  = datetime.datetime.strptime(d, isodateformat)
    if 'teamId' in c.keys():
        t    = c['teamId']
#   if 'teamName' in c.keys():
#       t    = c['teamName']
#       print (t)

    if 'gender' in c.keys():
        g = c['gender']
    if fn=='':
        ignore_bibs.append(bib)
    else:
        competitors[bib] = [fn, ln, dob, g, t]
#       print(bib, competitors[bib])



#print(type(j['events']))
#print(type(j['events'][0]))
#print(j['events'][0].keys())
#print(j['events'][0]['units'][0])
outdoor = 'J'
if j['type'] == "INDOOR":
    outdoor = 'N'
"""
results = {}
for e in j['events']:
    eventcode = e['eventCode']
    if eventcode not in results:
        results[eventcode] = {}
    #print(eventcode)
    for u in e['units']:
        #print(u.keys())
        for r in u['results']:
            bib = r['bib']
            if bib not in ignore_bibs:
                #print(r)
                d = competitors[bib][2]
                g = competitors[bib][3]
                cat = get_category(d,date,g)
                #print(d, date.strftime('%d.%m.%Y'), cat)
                if cat not in results[eventcode].keys():
                    results[eventcode][cat] = []
                # might want to add a dict of performance attributes to this tuple
                # e.g. jump series, sort keys ...
                rfiesults[eventcode][cat].append( (bib, r['performance']) )    
 
    print( bib, (fn, ln, dob.strftime('%d.%m.%Y'), t) )
    competitors[bib] = (fn, ln, dob, t)
#print(competitors)
"""
#print(competitors)

poolnr = 0
results ={}
series = {}
for e in j["events"]:
#   pp = pprint.PrettyPrinter(indent=4)
#   pp.pprint(e)
    day = e["day"]
    event_code = e["eventCode"]
    category = e["category"]
    event_key = (category, event_code)
#   print(event_code, event_key)
    if 'x' in event_key[1]:
        continue
    series[event_key] = {}
    #if day not in e.keys():
    if day not in results.keys():
        results[day] = {}
    if event_key not in e.keys():
        results[day][event_key] = {}
#       for u in e["units"]:
        trials = {}
        if event_code not in ( 'BI', 'TRI', 'QUAD', 'PEN', 'HEX', 'HEP', 'OCT', 'ENN', 'DEC', 'HEN', 'DOD', 'ICO'):
           for pool, u in enumerate(e["units"]):
#       for pool, u in zip(range(len(e["units"])),e["units"]):
               #results[event_code] ={}
               if "windAssistance" in u.keys():
                   wind = u["windAssistance"]
               else:
                   wind = None
               #print(wind)

               for r in u["results"]:
#               print(r)
                   if "bib" in r.keys():
                       bib = r["bib"]
                   
                   if bib not in ignore_bibs:
                        bdate = competitors[bib][2]
                        g = competitors[bib][3]
                        cat = get_category(bdate,date0,g)
                        if results[day][event_key].get(cat) == None:
                            results[day][event_key][cat] = {}
                        if results[day][event_key][cat].get(pool) == None:
                            #results[event_code][cat][pool] = []
                            results[day][event_key][cat][pool] = {'marks' : []}
                        if not wind == None:
                            results[day][event_key][cat][pool]['wind'] = wind
#                    x
#                       print(r.keys())
                        if 'performance' in r.keys():
                            res = r['performance']
                        else:
                            res = ''

                        if "place" in r.keys():
                            pl = r["place"]
                        else:
                            pl = noplace
                      
#                    if "order" in r.keys():
#                        pl = r["order"]
                       
#                    print (event_code, bib, res, pl, pool)
                        #results[event_code][cat][pool].append((bib, res, pl))
                        results[day][event_key][cat][pool]['marks'].append((bib, res, pl))
                        #print (bib, res, pl, pool)
                        
#                       t = r['teamId']
#                       competitors[bib][4] = t
#           poolnr = poolnr + 1
#           print (type(u['trials']))
#           print (u['trials'])
               if event_code in ('HJ', 'SHJ', 'PV'):
                   for t in u['trials']:
                       bib = t['bib']
                       if trials.get(bib)==None:
                           trials[bib] = {}
                       height = t['height']
                       if trials[bib].get(height)==None:
                           trials[bib][height] = []
                       trials[bib][height].append(t['result'])
                   for bib in trials.keys():
                       s = ''
                       for height in sorted(trials[bib].keys() ):
                           s += height + '(' + ''.join(trials[bib][height]) + ') ' 
                       s = s.replace('.',',')
#                    print(s)
                       i0 = i1 = len(s)
                       if 'x' in s:
                           i0 = s.index('x')
                       if 'o' in s:
                           i1 = s.index('o')
                       ij = min(i0,i1)
                       if ij < len(s):
                           series[event_key][bib] = s[ij-5:]
                       else:
                           series[event_key][bib] = ''
               elif event_code in ('LJ', 'TJ', 'SP', 'DT', 'HT', 'JT', 'OT', 'BT'):
                   for t in u['trials']:
                       bib = t['bib']
                       if trials.get(bib)==None:
                           trials[bib] = {}
#                   print(event_code, t)
                       rond = t['round']
                       if trials[bib].get(rond)==None:
                           trials[bib][rond] = {}
                       trials[bib][rond]['result'] = t['result']
                       if 'wind' in t.keys():
                           trials[bib][rond]['wind'] = t['wind']

                   for bib in trials.keys():
                       s = ''
                       for rond in sorted(trials[bib].keys() ):
                           s += trials[bib][rond]['result'] 
                           if 'wind' in trials[bib][rond].keys():
                               s += "(%3.1f)" % (trials[bib][rond]['wind'])
                           s += '/'    
                       s = s.replace('.',',')
                       series[event_key][bib] = s[:-1]
        else:
            if ( 'BI', 'TRI', 'QUAD', 'PEN', 'HEX', 'HEP', 'OCT', 'ENN', 'DEC', 'HEN', 'DOD', 'ICO').index(event_code) > 4:
                del results[day][event_key] 
                day +=1
                results[day][event_key] = {}
            print(event_key, day)
#           print(event_key)
#           print( results[day][event_key] )
            for r in e['results']:
#               print(r)
                if "bib" in r.keys():
                    bib = r["bib"]
                
                if bib not in ignore_bibs:
                     bdate = competitors[bib][2]
                     g = competitors[bib][3]
                     cat = get_category(bdate,date0,g)
#                    cat = r['category']
                     pool = 0
                     if results[day][event_key].get(cat) == None:
                         results[day][event_key][cat] = { pool : { 'marks' : [] } }
#                    if results[day][event_key][cat].get(pool) == None:
#                    results[event_code][cat][pool] = []
#                        results[day][event_key][cat][pool] = {'marks' : []}
#                    if not wind == None:
#                        results[day][event_key][cat][pool]['wind'] = wind
                     if 'total' in r.keys():
                         res = f'{r["total"]}'
                     if "place" in r.keys():
                         pl = r["place"]
                         if pl is None:
                             pl = noplace

#                        t = r['teamName']
#                        competitors[bib][4] = t
                     else:
                         pl = noplace
#                    print(bib, cat, res, pl)

#                    pool = 0
                     results[day][event_key][cat][pool]['marks'].append((bib, res, pl))


#       #   elif event_code in ( 'BI', 'TRI', 'QUAD', 'PEN', 'HEX', 'HEP', 'OCT', 'ENN', 'DEC', 'HEN', 'DOD', 'ICO'):
#               print(event_code)

#pp = pprint.PrettyPrinter(indent=4)
#pp.pprint(results)

#... write template for Results to xlsx workbook
wb = Workbook()

ws = wb.active
    
greenfont = Font(name='Calibri', color="0000FF00")
#greenfont = Font(name='Calibri', color=xlcolors.GREEN)
boldfont = Font(name='Calibri', bold=True, underline="single")
    
ws.title = "Resultatliste"
    
ws['a1'] = 'Stevne:';         ws['b1'] = meetname
ws['a2'] = 'Stevnested:';     ws['b2'] = venue
ws['a3'] = 'Stevnedato:';     ws['b3'] = date0.strftime('%d.%m.%Y'); ws['c3'] = date1.strftime('%d.%m.%Y')
ws['a4'] = 'Arrangør:';       ws['b4'] = organiser_name; #b4=ws['b4']; b4.font=greenfont
ws['a5'] = 'Kontaktperson:';  ws['b5'] = '<navn>'    ; b5=ws['b5']; b5.font=greenfont
ws['a6'] = 'Erklæring*: ';    ws['b6'] = 'J'     #; b6=ws['b6']; b6.font=greenfont
ws['a7'] = 'Telefon:';        ws['b7'] = '<tlf>'     ; b7=ws['b7']; b7.font=greenfont
ws['a8'] = 'Epost:';          ws['b8'] = j['contactDetails']  ; #b8=ws['b8']; b8.font=greenfont
ws['a9'] = 'Utendørs:';       ws['b9'] = outdoors #   ; b9=ws['b9']; b9.font=greenfont
ws['a10'] = 'Kommentar:';     ws['b10'] = url
ws['a11'] = 'Kommentar:';     


row_counter = 13 

day = 1
#for day,date in zip(range(1,len(dates)+1), dates):
for day,date in enumerate(dates):
    day +=1
    ws[f'A{row_counter}'] = 'Resultater';     ws[f'B{row_counter}'] = date.strftime('%d.%m.%Y')
    row_counter +=2
    for event_key in sorted(results[day].keys()):
#       print(event_key)
        event = event_key[1]
        for cat in sorted(results[day][event_key].keys() ):
            ws["A%(row_counter)d"%vars()] = cat; arc = ws["A%(row_counter)d"%vars()]; arc.font=boldfont
            ws["B%(row_counter)d"%vars()] = event_spec(event,cat) ; brc = ws["B%(row_counter)d"%vars()]; brc.font=boldfont
            row_counter +=1
            heats = sorted(results[day][event_key][cat].keys() )
            for h, heat in zip(range(len(heats)), heats):
                ws["A%(row_counter)d"%vars()] = "Heat:";  ws["B%(row_counter)d"%vars()] = h+1;  
                if 'wind' in results[day][event_key][cat][heat].keys():
                    ws["C%(row_counter)d"%vars()] = "Vind:";  ws["D%(row_counter)d"%vars()] = results[day][event_key][cat][heat]['wind']
                row_counter +=1
#               print( results[day][event_key][cat][heat]['marks'] )
                sorted_results = sorted(results[day][event_key][cat][heat]['marks'], key=lambda tup: tup[2])
                pat = "[GJ](\d?\d)"
                match = re.search(pat,event_key[0])
                if match: 
                    age = int(match.group(1))
                    if age < 11:
                        sorted_results = results[day][event_key][cat][heat]['marks']
                        random.shuffle(sorted_results)
                for i,r in zip(range(len(sorted_results)),sorted_results):
                    bib = r[0]
                    perf = r[1].replace('.',',')
                    place = r[2]
    
                    fn  = competitors[bib][0]
                    ln  = competitors[bib][1]
                    dob = competitors[bib][2]
                    club = competitors[bib][4]
    
                    if place == noplace:
                        pl = ''
                    else:
                        pl = i+1
                    ws["A%(row_counter)d"%vars()] = pl
                    #ws["B%(row_counter)d"%vars()] = bib
                    ws["C%(row_counter)d"%vars()] = ' '.join((fn,ln))
                    ws["D%(row_counter)d"%vars()] = dob.strftime('%Y')
                    ws["E%(row_counter)d"%vars()] = club_name(club)
#                   ws["E%(row_counter)d"%vars()] = club
                    ws["F%(row_counter)d"%vars()] = perf
    
    #--- extract wind for best performance from series
                    s = series[event_key].get(bib, 'no_series')
                    if event in ('LJ', 'TJ') and not s == 'no_series':
                        pat = r'/?%(perf)s\(([+-]?\d,\d)\)/?' % vars()
                        match = re.search(pat,s)
                        if match:
                            ws["G%(row_counter)d"%vars()] = match.group(1)
    
                    if not s=='no_series':
                        row_counter +=1
                        ws["A%(row_counter)d"%vars()] = s
                    row_counter +=1
            row_counter +=1
        
print("done")

"""
class_keys = athlete_by_event_by_class.keys()
class_keys.sort()
for klasse in class_keys:
   event_keys = athlete_by_event_by_class[klasse].keys()
   event_keys.sort()
   for event in event_keys:
           
       e = event_spec(event,klasse)
       ws["A%(row_counter)d"%vars()] = klasse; arc = ws["A%(row_counter)d"%vars()]; arc.font=boldfont
       ws["B%(row_counter)d"%vars()] = e     ; brc = ws["B%(row_counter)d"%vars()]; brc.font=boldfont
       ws["C%(row_counter)d"%vars()] = "<spesiell konkurransestatus>";  crc = ws["C%(row_counter)d"%vars()]; crc.font=greenfont
       row_counter +=1

       if istrack(event):
           ws["A%(row_counter)d"%vars()] = "<Heat | Finale:>";  arc = ws["A%(row_counter)d"%vars()]; arc.font=greenfont
           ws["C%(row_counter)d"%vars()] = "Vind:";  crc = ws["C%(row_counter)d"%vars()]; crc.font=greenfont
           row_counter +=1
          
       for athlete in athlete_by_event_by_class[klasse][event]:
          ws["C%(row_counter)d"%vars()] = athlete['name']
          ws["D%(row_counter)d"%vars()] = athlete['dob'][-4:]
          ws["E%(row_counter)d"%vars()] = athlete['club']
          ws["F%(row_counter)d"%vars()] = "<resultat>"
          if ishjump(event):
             ws["G%(row_counter)d"%vars()] = "<vind>";  grc = ws["G%(row_counter)d"%vars()]; grc.font=greenfont
             ws["H%(row_counter)d"%vars()] = "<resultat>";  hrc = ws["H%(row_counter)d"%vars()]; hrc.font=greenfont

             ws["I%(row_counter)d"%vars()] = "<vind>";  irc = ws["I%(row_counter)d"%vars()]; irc.font=greenfont
          if isfield(event):
             row_counter +=1 # add blank line for series
             ws["A%(row_counter)d"%vars()] = "<hopp-/kastserie>";  arc = ws["A%(row_counter)d"%vars()]; arc.font=greenfont
    
          row_counter +=1
       row_counter +=1
           
    
fname = output_file_name(tree)
xlname = fname+'.xlsx'
"""
xlname = slug + '-' + date0.strftime(isodateformat) + '.xlsx'
wb.save(xlname)
