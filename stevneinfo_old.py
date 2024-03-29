﻿#!/usr/bin/python
# -*- coding: utf-8 -*-
import sys
from xml.etree import ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import colors as xlcolors
from openpyxl.styles import Font, Color
#from tabulate import tabulate
from reportlab.lib import colors as rlcolors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from unidecode import unidecode
from collections import defaultdict
import requests
from bs4 import BeautifulSoup
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import datetime
import re
#import requests
#from bs4 import BeautifulSoup


def read_xml_into_tree(infile):
   #with open(infile, 'rt') as f:
   with open(infile, 'rb') as f:
      tree = ET.parse(f)
   return tree

def extract_competition_data(tree):
   s = tree.find('.//Competition')
   mn = s.attrib['name']
   md = s.attrib['startDate']

   v = s.find('CompetitionVenue')
   mv = v.attrib['startingvenue']
   return{'meet_name' : mn, 'meet_date' : md, 'venue' : mv}

def output_file_name(tree):
   cdata = extract_competition_data(tree)
   mn = cdata['meet_name']
   md = cdata['meet_date']
   mv = cdata['venue']

   sn =  mn + ' ' + mv + ' ' + md
   sn = sn.replace(' ','_')
   sn = sn.replace('/','-')
   fname = unidecode( sn.replace(' ','_') )
   return fname

def save_xml_copy(tree):
   fname = output_file_name(tree)
   # save a copy of the element tree
   tree.write(fname+'.xml', encoding="utf-8")

def istrack(event):
    return 'meter' in event

def ishurdles(event):
    return istrack(event) and 'hekk' in event

def issteeple(event):
    return istrack(event) and 'hinder' in event

def isfield(event):
    return isvjump(event) or ishjump(event) or isthrow(event)

def isvjump(event):
    return event in [u'Høyde', u'Stav', u'Høyde uten tilløp']

def ishjump(event):
    return event in [u'Lengde', u'Lengde satssone', u'Tresteg', u'Lengde uten tilløp', u'Tresteg uten tilløp']

def isthrow(event):
    return event in [u'Kule', u'Diskos', u'Slegge', u'Spyd', u'Vektkast', u'Liten ball']

def ismulti(event):
    return 'kamp' in event

def sort_athletes_by_class_by_event(tree):
    athlete_by_class_by_event = {}
    for c in tree.findall('.//Competitor'):
       p = c.find('Person')
       club = p.attrib['clubName']
       cs = club.split(',')
       if len(cs) > 1:
          club = cs[1].strip() + ' ' + cs[0].strip()
       n = p.find('./Name')
       name = "Name"
       fn = n.find('Given')
       en = n.find('Family')
       bd = p.find('BirthDate')
       name = fn.text + ' ' + en.text
       dd = int(bd.attrib['day'])
       mm = int(bd.attrib['month'])
       yyyy = int(bd.attrib['year'])
       dob = "%(dd)02d.%(mm)02d.%(yyyy)04d" % vars()
       ec = c.find('./Entry/EntryClass')
       klasse = ec.attrib['classCode']
       ec = c.find('./Entry/Exercise')
       event = ec.attrib['name'] 
       athlete_key = name + dob + club
      
       if event not in athlete_by_class_by_event.keys():
          athlete_by_class_by_event[event]={}
       if klasse not in athlete_by_class_by_event[event].keys():
          athlete_by_class_by_event[event][klasse]=[]
       athlete_by_class_by_event[event][klasse].append({'name': name, 'dob': dob, 'club' : club }) 

    return athlete_by_class_by_event


def list_entries(tree):
    entries_list = []
    for c in tree.findall('.//Competitor'):
       p = c.find('Person')
       club = p.attrib['clubName']
       cs = club.split(',')
       if len(cs) > 1:
          club = cs[1].strip() + ' ' + cs[0].strip()
       n = p.find('./Name')
       name = "Name"
       fn = n.find('Given')
       en = n.find('Family')
       bd = p.find('BirthDate')
       dd = int(bd.attrib['day'])
       mm = int(bd.attrib['month'])
       yyyy = int(bd.attrib['year'])
       dob = "%(dd)02d.%(mm)02d.%(yyyy)04d" % vars()
       ec = c.find('./Entry/EntryClass')
       klasse = ec.attrib['classCode']
       ec = c.find('./Entry/Exercise')
       event = ec.attrib['name'] 
       entry = {'first_name': fn.text, 'last_name' : en.text, 'birth_date' : dob, 'club' : club, 'event' : event}

       entries_list.append( entry )
    
    return entries_list

def sort_athletes_by_event_by_class(tree):
    athlete_by_event_by_class = {}
    for c in tree.findall('.//Competitor'):
       p = c.find('Person')
       club = p.attrib['clubName']
       cs = club.split(',')
       if len(cs) > 1:
          club = cs[1].strip() + ' ' + cs[0].strip()
       n = p.find('./Name')
       name = "Name"
       fn = n.find('Given')
       en = n.find('Family')
       bd = p.find('BirthDate')
       name = fn.text + ' ' + en.text
       dd = int(bd.attrib['day'])
       mm = int(bd.attrib['month'])
       yyyy = int(bd.attrib['year'])
       dob = "%(dd)02d.%(mm)02d.%(yyyy)04d" % vars()
       ec = c.find('./Entry/EntryClass')
       klasse = ec.attrib['classCode']
       ec = c.find('./Entry/Exercise')
       event = ec.attrib['name'] 
       athlete_key = name + dob + club

       if klasse not in athlete_by_event_by_class.keys():
           athlete_by_event_by_class[klasse] = {}
       if event not in athlete_by_event_by_class[klasse].keys():
           athlete_by_event_by_class[klasse][event] = [] 
       athlete_by_event_by_class[klasse][event].append({'name': name, 'dob': dob, 'club' : club }) 


    return athlete_by_event_by_class

def sort_events_by_athlete(tree):
    events_by_athlete= {}
    for c in tree.findall('.//Competitor'):
       p = c.find('Person')
       club = p.attrib['clubName']
       g = p.attrib['sex']
       cs = club.split(',')
       if len(cs) > 1:
          club = cs[1].strip() + ' ' + cs[0].strip()
       n = p.find('./Name')
       name = "Name"
       fn = n.find('Given')
       en = n.find('Family')
       bd = p.find('BirthDate')
       idd = p.find('Identity')
       idt = idd.attrib['value']
       name = fn.text + ' ' + en.text
       dd = int(bd.attrib['day'])
       mm = int(bd.attrib['month'])
       yyyy = int(bd.attrib['year'])
       dob = "%(dd)02d.%(mm)02d.%(yyyy)04d" % vars()
       ec = c.find('./Entry/EntryClass')
       klasse = ec.attrib['classCode']
       if klasse == '':
          klasse = ec.attrib['shortName']
       ec = c.find('./Entry/Exercise')
       #event = klasse + ' ' +ec.attrib['name'] 
       event = ( klasse, ec.attrib['name'] )
       athlete_key = '+'.join((fn.text, en.text, dob, g, club,idt))

       if athlete_key not in events_by_athlete.keys():
           events_by_athlete[athlete_key] = []
       if event not in events_by_athlete[athlete_key]:
           events_by_athlete[athlete_key].append(event)

    return events_by_athlete

def write_start_lists_as_html(tree):
    competition_data = extract_competition_data(tree)
    #mn = competition_data['meet_name']
    #md = competition_data['meet_date']
    #mv = competition_data['venue']
    mn = competition_data['meet_name'].encode('utf-8')
    md = competition_data['meet_date'].encode('utf-8')
    mv = competition_data['venue'].encode('utf-8')

    athlete_by_class_by_event = sort_athletes_by_class_by_event(tree)

    fname = output_file_name(tree)
    of = open (fname+"_deltagere_pr_ovelse.html".encode('utf-8'), 'w')
    of.write(""" <!DOCTYPE html>
    <meta charset="UTF-8">
    <html>
    <body>
    <title> %(mn)s </title>
    <h1> %(mn)s </h1>
    """ % vars() )
    #of.write("%s, %s" % ( md, mv.encode('utf-8') ) )
    of.write("%s, %s" % ( md, mv ) )
    event_keys = athlete_by_class_by_event.keys() 
    event_keys.sort()
    for event_key in event_keys:
        class_keys = athlete_by_class_by_event[event_key].keys()
        class_keys.sort()
        for class_key in  class_keys:
           of.write( "<h2>%s %s </h2>\n<ul style=\"list-style-type:none\">\n"% (event_key.encode('utf-8'), class_key) )
           for athlete in athlete_by_class_by_event[event_key][class_key]:
               of.write("<li>" +athlete['name'].encode('utf-8') + ' (' + athlete['dob'][-4:] +'), ' + athlete['club'].encode('utf-8') +  "</li>\n" )
           of.write("</ul>\n")
    
    of.write("""</body>
    </html>""")
    of.close()


def write_opentrack_import(tree):
    competition_data = extract_competition_data(tree)
    mn = competition_data['meet_name']
    md = competition_data['meet_date']
    mv = competition_data['venue']

    events = list_events(tree)
    events_by_athlete = sort_events_by_athlete(tree)

    isodateformat = "%Y-%m-%d"
    #... write template for Results to xlsx workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Competitors'
    ws1 = wb.create_sheet('Events')
    
    row_counter = 1
    ws["A1"] = 'Competitor Id'
    ws["B1"] = 'National Id'
    ws["C1"] = 'First name'
    ws["D1"] = 'Last name'
    ws["E1"] = 'Gender'
    ws["F1"] = 'Date of birth'
    ws["G1"] = 'Team ID'
    ws["H1"] = 'Nationality'
    ws["I1"] = 'Event'
    ws["J1"] = 'Pb'
    ws["K1"] = 'Sb'
    ws1["A1"] = 'Event selection'
    row_counter = 2

    jf = 0
    jt = 0
    jm = 0
    full_events = {}
    print(events)
    for e in events:
        if isfield(e[1]):
            jf +=1
            event_ref = "F%02d"%jf
        elif ismulti(e[1]):
            jm +=1
            event_ref = "M%02d"%jm
        else:
            jt +=1
            event_ref = "T%02d"%jt

        #print(e)
        full_events[ ( class_code(e[0]) , e[1] ) ]  = event_ref + ' - ' + ' '.join(( class_code(e[0]), event_spec(e[1], class_code(e[0])) ))
        ws1["A%d"%row_counter] = event_ref + ' - '  + ' '.join([e[0], event_spec(e[1], class_code(e[0]))])
        ws1["B%d"%row_counter] = event_ref
        ws1["C%d"%row_counter] = event_code(e[1])
        ws1["D%d"%row_counter] = age_group(class_code(e[0]))
        ws1["E%d"%row_counter] = gender(class_code(e[0]))
        ws1["F%d"%row_counter] = class_code(e[0])
#       ws1["G%d"%row_counter] = age_group(class_code(e[0]))

        ws1["H%d"%row_counter] = ' '.join(( class_code(e[0]), event_spec(e[1], class_code(e[0])) ))
        ws1["I%d"%row_counter] = '1'
        ws1["J%d"%row_counter] = '1'
        ws1["K%d"%row_counter] = '12:00'
        
        row_counter +=1
#   ws.insert_cols(13)

#   print (full_events)
    row_counter = 2    
    bib = 0
    for key in events_by_athlete.keys():
        bib +=1
        k = key.split('+')
        fn = k[0]
        en = k[1]
        dob = '-'.join(( k[2][6:10], k[2][3:5], k[2][0:2] ))
        dob = datetime.datetime.strptime(dob,isodateformat)
        g = k[3]
        club = k[4]
        ident = k[5]

        for e in events_by_athlete[key]:
            ws["A%d"%row_counter] = bib
            ws["B%d"%row_counter] = ident
            ws["C%d"%row_counter] = fn
            ws["D%d"%row_counter] = en
            ws["E%d"%row_counter] = gender(g)
            ws["F%d"%row_counter] = dob.strftime(isodateformat)
            ws["G%d"%row_counter] = club_code(club)
            ws["I%d"%row_counter] = full_events[e]

            #print(e, full_events[e])
            event = e[1]
            print(event)
            if not isfield(e[1]):
                if event == "60 meter": # for Bassen sprint
                    event = "100 meter"
                res1 = get_seed_marks(' '.join((fn, en)), dob, event, e[0], '2021' )
                #res2 = get_seed_marks(' '.join((fn, en)), dob, e[1], e[0], '2020' )
                #print(res1,res2)
                #res = min(res1,res2)
                res = res1
                if res=='nm':
                    res=''
                if e[1] == '10 000 meter':
                    res3 = get_seed_marks(' '.join((fn, en)), dob, '5000 meter', e[0], '2021' )
                    res4 = get_seed_marks(' '.join((fn, en)), dob, '5000 meter', e[0], '2020' )
                    res3 = min(res3,res4)
                    if res3 == 'nm':
                        res3 = ''
                    ws["M%d"%row_counter] = res3
            else:
                res = ''
            ws["J%d"%row_counter] = res

            row_counter +=1

    fname = output_file_name(tree)
    xlname = fname+'_opentrack.xlsx'
    wb.save(xlname)

def write_xlsx_results_template(tree):
    competition_data = extract_competition_data(tree)
    mn = competition_data['meet_name']
    md = competition_data['meet_date']
    mv = competition_data['venue']

    athlete_by_event_by_class = sort_athletes_by_event_by_class(tree)

    #... write template for Results to xlsx workbook
    wb = Workbook()
    ws = wb.active
    
    greenfont = Font(name='Calibri', color=xlcolors.GREEN)
    boldfont = Font(name='Calibri', bold=True, underline="single")
    
    ws.title = "Resultatliste"
    
    ws['a1'] = 'Stevne:';         ws['b1'] = mn
    ws['a2'] = 'Stevnested:';     ws['b2'] = mv
    ws['a3'] = 'Stevnedato:';     ws['b3'] = md          ; ws['c3'] = '<til dato>'; c3=ws['c3']; c3.font=greenfont
    ws['a4'] = 'Arrangør:';       ws['b4'] = '<arrangør>'; b4=ws['b4']; b4.font=greenfont
    ws['a5'] = 'Kontaktperson:';  ws['b5'] = '<navn>'    ; b5=ws['b5']; b5.font=greenfont
    ws['a6'] = 'Erklæring*: ';    ws['b6'] = '<J/N>'     ; b6=ws['b6']; b6.font=greenfont
    ws['a7'] = 'Telefon:';        ws['b7'] = '<tlf>'     ; b7=ws['b7']; b7.font=greenfont
    ws['a8'] = 'Epost:';          ws['b8'] = '<e-post>'  ; b8=ws['b8']; b8.font=greenfont
    ws['a9'] = 'Utendørs:';       ws['b9'] = '<J/N>'     ; b9=ws['b9']; b9.font=greenfont
    ws['a10'] = 'Kommentar:'
    
    ws['a12'] = 'Resultater';     ws['b12'] = md
    
    row_counter = 14
    class_keys = athlete_by_event_by_class.keys()
    class_keys.sort()
    for klasse in class_keys:
       event_keys = athlete_by_event_by_class[klasse].keys()
       event_keys.sort()
       for event in event_keys:
           
           e = event_spec(event,klasse)
           ws["A%(row_counter)d"%vars()] = klasse; arc = ws["A%(row_counter)d"%vars()]; arc.font=boldfont
           ws["B%(row_counter)d"%vars()] = e     ; brc = ws["B%(row_counter)d"%vars()]; brc.font=boldfont
           ws["C%(row_counter)d"%vars()] = "<spesiell konkurransestatus>";  crc = ws["C%(row_counter)d"%vars()]; crc.font=greenfont
           row_counter +=1

           if istrack(event):
               ws["A%(row_counter)d"%vars()] = "<Heat | Finale:>";  arc = ws["A%(row_counter)d"%vars()]; arc.font=greenfont
               ws["C%(row_counter)d"%vars()] = "Vind:";  crc = ws["C%(row_counter)d"%vars()]; crc.font=greenfont
               row_counter +=1
          
           for athlete in athlete_by_event_by_class[klasse][event]:
              ws["C%(row_counter)d"%vars()] = athlete['name']
              ws["D%(row_counter)d"%vars()] = athlete['dob'][-4:]
              ws["E%(row_counter)d"%vars()] = athlete['club']
              ws["F%(row_counter)d"%vars()] = "<resultat>"
              if ishjump(event):
                 ws["G%(row_counter)d"%vars()] = "<vind>";  grc = ws["G%(row_counter)d"%vars()]; grc.font=greenfont
                 ws["H%(row_counter)d"%vars()] = "<resultat>";  hrc = ws["H%(row_counter)d"%vars()]; hrc.font=greenfont
                 ws["I%(row_counter)d"%vars()] = "<vind>";  irc = ws["I%(row_counter)d"%vars()]; irc.font=greenfont
              if isfield(event):
                 row_counter +=1 # add blank line for series
                 ws["A%(row_counter)d"%vars()] = "<hopp-/kastserie>";  arc = ws["A%(row_counter)d"%vars()]; arc.font=greenfont
    
              row_counter +=1
           row_counter +=1
           
    
    fname = output_file_name(tree)
    xlname = fname+'.xlsx'
    wb.save(xlname)

def make_horizontal_protocol(tree, event, classes):
    """make the event protocol sheet for a horizontal jump or throw event
    writes the protocol sheet to a pdf
    Input:
    	tree: ElementTree
    	event: the event (fails if event is not a hjump or throw)
 	classes: list of classes """
    if not ishjump(event) and not isthrow(event):
       sys.exit('make_horizontal_protocol: event is not hjump or throw')
    athlete_by_event_by_class = sort_athletes_by_event_by_class(tree)
    athlete_by_class_by_event = sort_athletes_by_class_by_event(tree)
 
    eventclass = event + ' ' + '+'.join(classes)
 
    fname = output_file_name(tree) + '_' + event + '-' + '+'.join(classes) +'.pdf'
    fname = fname.replace(' ', '_')
    fname = fname.replace('/', '-')
    doc = SimpleDocTemplate(fname, pagesize=A4)
    doc.pagesize = landscape(A4)
 
    rows_on_page = 12
    # container for the 'Flowable' objects
    elements = []
    
    styles = getSampleStyleSheet()
 
    data= [ [event + ': ' + ', '.join(classes)],
            ['Klasse', 'Navn', 'F.år', 'Klubb', 'Forsøk 1', 'Forsøk 2', 'Forsøk 3', 'Forsøk 4', 'Forsøk 5', 'Forsøk 6', 'Resultat'] ]
    if ishjump(event):
       data[1].append('Vind')
 
    rows = 0
    for c in classes:
       if c in athlete_by_class_by_event[event].keys():
          for athlete in athlete_by_class_by_event[event][c]:
             data.append( [ c, athlete['name'], athlete['dob'][-4:], athlete['club'] ] )
             rows +=1
    pages = int(rows/(rows_on_page-1)) + 1
 
    if rows%rows_on_page > 5:
       pages +=1
    rows_in_table = pages*rows_on_page 
 
    if rows < rows_in_table:
       for i in range(rows_in_table-rows-2):
          data.append([ ' ' ])
 
 
 
    t=Table(data, [1.9*cm, 6.0*cm, 2.1*cm, 2.6*cm , 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm], rows_in_table*[1.4*cm], repeatRows=2)
 
    t.setStyle(TableStyle([
                           ('SPAN',(0,0),(-1,0)),
                           ('ALIGN',(0,1),(0,-1),'CENTER'),
                           ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                           ('ALIGN',(1,1),(1,-1),'LEFT'),
                           ('INNERGRID', (0,0), (-1,-1), 0.25, rlcolors.black),
                           ('BOX', (0,0), (-1,-1), 0.25, rlcolors.black),
                           ]))
    elements.append(t)
    # write the document to disk
    doc.build(elements)

def make_vertical_protocol(tree, event, classes):
    """make the event protocol sheet for a horizontal jump or throw event
    writes the protocol sheet to a pdf
    Input:
    	tree: ElementTree
    	event: the event (fails if event is not a vjump )
 	classes: list of classes """
    if not isvjump(event):
       sys.exit('make_vertical_protocol: event is not vjump')
    athlete_by_event_by_class = sort_athletes_by_event_by_class(tree)
    athlete_by_class_by_event = sort_athletes_by_class_by_event(tree)
 
    eventclass = event + ' ' + '+'.join(classes)
 
    fname = output_file_name(tree) + '_' + event + '-' + '+'.join(classes) +'.pdf'
    fname = fname.replace(' ', '_')
    fname = fname.replace('/', '-')
    print(fname)
    doc = SimpleDocTemplate(fname, pagesize=A4)
    doc.pagesize = landscape(A4)
 
    rows_on_page = 22
    # container for the 'Flowable' objects
    elements = []
    
    styles = getSampleStyleSheet()
 
    data= [ [event + ': ' + ', '.join(classes) ],
            ['Klasse', 'Navn', 'F.år', 'Klubb', '', '', '', '', '', '','','','','','','', 'Res', 'Pl' ] ]
 
    rows = 0
    for c in classes:
       if c in athlete_by_class_by_event[event].keys():
          for athlete in athlete_by_class_by_event[event][c]:
             data.append( [ c, athlete['name'], athlete['dob'][-4:], athlete['club'][0:11] ] )
             rows +=1
    pages = int(rows/(rows_on_page-1)) + 1
    #print(pages)
 
    if rows%rows_on_page > 15:
       pages +=1
    rows_in_table = pages*rows_on_page 
 
    if rows < rows_in_table:
       for i in range(rows_in_table-rows-2):
          data.append([ ' ' ])
 
 
 
    t=Table(data, [1.9*cm, 6.0*cm, 2.1*cm, 2.6*cm , 0.8*cm, 0.8*cm, 0.8*cm, 0.8*cm, 0.8*cm, 0.8*cm, 0.8*cm,0.8*cm,0.8*cm,0.8*cm,0.8*cm, 0.8*cm, 1.6*cm, 0.8*cm], rows_in_table*[0.7*cm], repeatRows=2)
 
    t.setStyle(TableStyle([
                           ('SPAN',(0,0),(-1,0)),
                           ('ALIGN',(0,1),(0,-1),'CENTER'),
                           ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                           ('ALIGN',(1,1),(1,-1),'LEFT'),
                           ('INNERGRID', (0,0), (-1,-1), 0.25, rlcolors.black),
                           ('BOX', (0,0), (-1,-1), 0.25, rlcolors.black),
                           ]))
    elements.append(t)
    # write the document to disk
    doc.build(elements)

def list_events(tree):
    event_list = []
    for c in tree.findall('.//Competitor'):
       ec = c.find('./Entry/EntryClass')
       kl= ec.attrib['shortName']
       ec = c.find('./Entry/Exercise')
       ev= ec.attrib['name'] 
       event = (kl,ev)
       if event not in event_list:
           event_list.append(event)
    event_list.sort()
    return event_list

def make_crosstable(tree):
    events_by_athlete = sort_events_by_athlete(tree)
    crosstable = {}

    for a in events_by_athlete.keys():
        for e1 in events_by_athlete[a]:
            if e1 not in crosstable.keys():
                crosstable[e1] = {}
            for e2 in events_by_athlete[a]:
                if e2 not in crosstable[e1].keys():
                    crosstable[e1][e2] = 0
                crosstable[e1][e2] +=1
    return crosstable

def event_code(event):
    event_codes = {
            u'60 meter'          : '60', 
            u'80 meter'          : '80', 
            u'100 meter'         : '100', 
            u'150 meter'         : '150', 
            u'200 meter'         : '200', 
            u'300 meter'         : '300', 
            u'400 meter'         : '400', 
            u'600 meter'         : '600', 
            u'800 meter'         : '800', 
            u'1000 meter'        : '1000', 
            u'1500 meter'        : '1500', 
            u'3000 meter'        : '3000', 
            u'5000 meter'        : '5000', 
            u'10 000 meter'       : '10000', 
            u'60 meter hekk'     : '60H', 
            u'80 meter hekk'     : '80H', 
            u'100 meter hekk'    : '100H', 
            u'110 meter hekk'    : '110H', 
            u'200 meter hekk'    : '200H', 
            u'300 meter hekk'    : '300H', 
            u'400 meter hekk'    : '400H', 
            u'2000 meter hinder' : '2000SC', 
            u'3000 meter hinder' : '3000SC', 
            u'Kappgang 1000 meter' : '1000W', 
            u'Kappgang 3000 meter' : '3000W', 
            u'Kappgang 5000 meter' : '5000W', 
            u'Kappgang 5 km'     : '5000W', 
            u'Kappgang'          : '1500W', 
            u'Høyde'             : 'HJ', 
            u'Høyde uten tilløp' : 'SHJ', 
            u'Stav'              : 'PV', 
            u'Lengde'            : 'LJ', 
            u'Lengde satssone'   : 'LJ', 
            u'Lengde uten tilløp': 'SLJ', 
            u'Tresteg'           : 'TJ', 
            u'Tresteg satssone'  : 'TJ', 
            u'Tresteg uten tilløp'  : 'STJ', 
            u'Kule'              : 'SP', 
            u'Diskos'            : 'DT', 
            u'Slegge'            : 'HT', 
            u'Spyd'              : 'JT', 
            u'Liten ball'        : 'OT', 
            u'Tikamp'            : 'DEC', 
            u'Sjukamp'           : 'HEP' ,
            u'7-kamp'           : 'HEP' ,
            u'5-kamp'           : 'PEN' ,
            u'4x200 meter stafett' : '4x200' 
            }
    return event_codes[event]
 
def event_name(code):
    event_names = {
            '60'     : '60 meter'          , 
            '80'     : '80 meter'          , 
            '100'    : '100 meter'         , 
            '150'    : '150 meter'         , 
            '200'    : '200 meter'         , 
            '300'    : '300 meter'         , 
            '400'    : '400 meter'         , 
            '600'    : '600 meter'         , 
            '800'    : '800 meter'         , 
            '1000'   : '1000 meter'        , 
            '1500'   : '1500 meter'        , 
            '3000'   : '3000 meter'        , 
            '5000'   : '5000 meter'        , 
            '10000'  : '10000 meter'       , 
            '60H'    : '60 meter hekk'     , 
            '80H'    : '80 meter hekk'     , 
            '100H'   : '100 meter hekk'    , 
            '110H'   : '110 meter hekk'    , 
            '200H'   : '200 meter hekk'    ,
            '300H'   : '300 meter hekk'    , 
            '400H'   : '400 meter hekk'    , 
            '2000SC' : '2000 meter hinder' , 
            '3000SC' : '3000 meter hinder' , 
            'HJ'     : 'Høyde'             , 
            'PV'     : 'Stav'              , 
            'LJ'     : 'Lengde'            , 
            'TJ'     : 'Tresteg'           , 
            'SP'     : 'Kule'              , 
            'DT'     : 'Diskos'            , 
            'HT'     : 'Slegge'            , 
            'JT'     : 'Spyd'              , 
            'DEC'    : 'Tikamp'            , 
            'HEP'    : 'Sjukamp'           
            }
    return event_names[code]

def class_code(name):
    class_codes = {
            u'6 år Fellesklasse' : u'F6' ,
            u'7 år Fellesklasse' : u'F7' ,
            u'Gutter 6 - 7'     : u'G 6-7'          , 
            u'Gutter 8'     : u'G 8'          , 
            u'Gutter 9'     : u'G 9'          , 
            u'Gutter 10'    : u'G10'          , 
            u'Gutter 11'    : u'G11'          , 
            u'Gutter 12'    : u'G12'          , 
            u'Gutter 13'    : u'G13'          , 
            u'Gutter 14'    : u'G14'          , 
            u'Gutter 15'    : u'G15'          , 
            u'Gutter 16'    : u'G16'          , 
            u'Gutter 17'    : u'G17'          , 
            u'Gutter 18/19' : u'G18/19'       , 
            u'Gutter alle klasser' : u'GALLE'       , 
            u'Menn junior'  : u'MJ'           , 
            u'Menn U20'     : u'MU20'         , 
            u'Menn U23'     : u'MU23'         , 
            u'Menn senior'  : u'MS'           , 
            u'Menn alle klasser'  : u'MALLE'           , 
            u'Menn veteraner' : u'MV'         , 
            u'Jenter 6 - 7'     : u'J 6-7'          , 
            u'Jenter 8'     : u'J 8'          , 
            u'Jenter 9'     : u'J 9'          , 
            u'Jenter 10'    : u'J10'          , 
            u'Jenter 11'    : u'J11'          , 
            u'Jenter 12'    : u'J12'          , 
            u'Jenter 13'    : u'J13'          , 
            u'Jenter 14'    : u'J14'          , 
            u'Jenter 15'    : u'J15'          , 
            u'Jenter 16'    : u'J16'          , 
            u'Jenter 17'    : u'J17'          , 
            u'Jenter 18/19' : u'J18/19'       , 
            u'Jenter alle klasser' : u'JALLE'       , 
            u'Kvinner junior'  : u'KJ'        , 
            u'Kvinner U20'     : u'KU20'      , 
            u'Kvinner U23'     : u'KU23'      , 
            u'Kvinner senior'  : u'KS'        , 
            u'Kvinner alle klasser'  : u'KALLE'           , 
            u'Kvinner veteraner' : u'KV'      ,
            u'Funksjonshemmede' : u'FH'      ,
            u'Ikke valgt klasse' : u'IVK'
            }
    #print(name, class_codes[name.strip()])
    #return class_codes[name.strip()]
    return class_codes.get(name.strip(), name.strip())

def age_group(class_code):
    age_groups = {
            'F6'    : '6',
            'F7'    : '7',
            'G 6-7'    : '6-7',
            'G 8'    : '8',
            'G 9'    : '9',
            'G10'    : '10',
            'G11'    : '11',
            'G12'    : '12',
            'G13'    : '13',
            'G14'    : '14',
            'G15'    : '15',
            'G16'    : '16',
            'G17'    : '17',
            'G18/19' : '18-19',
            'GALLE' : 'ALL',
            'MJ'     : 'U20' ,
            'MS'     : 'SEN' ,
            'MALLE'     : 'ALL' ,
            'MV'     : 'V35' ,
            'MV35'   : 'V35' ,
            'J 6-7'    : '6-7',
            'J 8'    : '8',
            'J 9'    : '9',
            'J10'    : '10',
            'J11'    : '11',
            'J12'    : '12',
            'J13'    : '13',
            'J14'    : '14',
            'J15'    : '15',
            'J16'    : '16',
            'J17'    : '17',
            'J18/19' : '18/19',
            'JALLE' : 'ALL',
            'KJ'     : 'U20',
            'KS'     : 'SEN' ,
            'KV'     : 'V35' ,
            'KALLE'     : 'ALL' ,
            'FH'   : 'ALL' ,
            'IVK'    : 'ALL'  
            }

    #return age_groups[class_code]
    return age_groups.get(class_code, class_code)

def gender(class_code):
    if class_code[0] in ('G', 'M'):
        g = 'M'
    elif class_code[0] in ('J', 'K'):
        g = 'F'
    else:
        g = 'MF'

    return g

def event_spec(event, klasse):
    throws = {}
    throws['Kule'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg', 
                       'default' : ''} 
    throws['Diskos'] = { 'J10' : '0,6kg', 'J11' : '0,6kg', 'J12' : '0,6kg', 'J13' : '0,6kg', 
                       'J14' : '0,75kg', 'J15' : '0,75kg', 'J16' : '0,75kg', 'J17' : '0,75kg',
                       'J18/19' : '1,0kg', 'KU20' : '1,0kg', 'KU23' : '1,0kg', 'KS' : '1,0kg', 
                       'G10' : '0,6kg', 'G11' : '0,6kg', 'G12' : '0,75kg', 'G13' : '0,75kg', 
                       'G14' : '1,0kg', 'G15' : '1,0kg', 'G16' : '1,5kg', 'G17' : '1,5kg',
                       'G18/19' : '1,75kg', 'MU20' : '1,75kg', 'MU23' : '2,0kg', 'MS' : '2,0kg',
                       'default': ''} 
    throws['Slegge'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg',
                       'default': ''} 
    throws['Spyd'] = { 'J10' : '400g', 'J11' : '400g', 'J12' : '400g', 'J13' : '400g', 
                       'J14' : '400g', 'J15' : '500g', 'J16' : '500g', 'J17' : '500g',
                       'J18/19' : '600g', 'KU20' : '600g', 'KU23' : '600g', 'KS' : '600g', 
                       'G10' : '400g', 'G11' : '400g', 'G12' : '400g', 'G13' : '400g', 
                       'G14' : '600g', 'G15' : '600g', 'G16' : '700g', 'G17' : '700g',
                       'G18/19' : '800g', 'MU20' : '800g', 'MU23' : '800g', 'MS' : '800g',
                       'default': ''} 
#    throws['Liten ball'] = { 'J10' : '150g', 'J11' : '150g', 'J12' : '150g', 'J13' : '150g', 'J14' : '150g', 
#                             'G10' : '150g', 'G11' : '150g', 'G12' : '150g', 'G13' : '150g', 'G14' : '150g' 
#                             }
    throws['Liten ball'] = defaultdict(lambda : '150g') 
    hurdles = {}
    hurdles['60 meter hekk'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '76,2cm', 'J13' : '76,2cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '84,0cm','KJ' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '76,2cm', 'G13' : '76,2cm', 'G14' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm', 'G17' : '91,4cm',
                                 'G18/19' : '100cm','MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm', 'default':'' }
    hurdles['80 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'G14' : '84,0cm' } 
    hurdles['100 meter hekk'] = { 'J16' : '76,2cm', 'J17' : '76,2cm', 'J18/19' : '84,0cm','KJ' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm'}

    hurdles['110 meter hekk'] = { 'G17' : '91,4cm', 'G18/19' : '100cm','MJ' : '100cm', 'MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm' }
    hurdles['200 meter hekk'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '68,0cm', 'J13' : '68,0cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KJ' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '68,0cm', 'G13' : '68,0cm', 'G14' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '76,2cm', 'G17' : '76,2cm',
                                 'G18/19' : '76,2cm','MJ' : '76,2cm', 'MU20' : '76,2cm', 'MU23' : '76,2cm', 'MS' : '76,2cm' }
    hurdles['300 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KJ' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MJ' : '91,4cm', 'MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm' }
    hurdles['400 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KJ' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MJ' : '91,4cm','MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm' }

    if isthrow(event):
       #e = event + ' ' + throws[event][klasse]
       t = throws[event].get(klasse,None)
       if t == None:
           t = throws[event]['default']
       e = event + ' ' + t
    elif ishurdles(event):
       h = hurdles[event].get(klasse,None)
       if h == None:
           h = hurdles[event]['default']
       e = event + ' ' + h
    else:
       e = event

    return e


def club_code(club_name):
    if club_name in (u'ÅLEN IDRETTSLAG'):
       club_code=u'AALEN'
    elif club_name in (u'Ålesund Friidrettsklubb', u'Ålesund FIK'):
       club_code=u'AASUN'
    elif club_name in (u'Allianseidrettslaget Ik Våg'):
       club_code=u'IKV'
    elif club_name in (u'Almenning Il'):
       club_code=u'ALM'
    elif club_name in (u'Alsvåg Idrettslag'):
       club_code=u'ALSV'
    elif club_name in (u'Alta Idrettsforening'):
       club_code=u'ALTA'
    elif club_name in (u'Åndalsnes Idrettsforening'):
       club_code=u'ANDAL'
    elif club_name in (u'Andebu Idrettslag'):
       club_code=u'ANDE'
    elif club_name in (u'Andørja Sportsklubb'):
       club_code=u'ANDSK'
    elif club_name in (u'Aremark Idrettsforening', u'Aremark IF'):
       club_code=u'ARE'
    elif club_name in (u'Arna Turn & Idrettslag'):
       club_code=u'ARNA'
    elif club_name in (u'Ås Idrettslag', u'Ås IL'):
       club_code=u'ASIL'
    elif club_name in (u'Åsen Idrettslag'):
       club_code=u'AASEN'
    elif club_name in (u'Åseral idrettslag'):
       club_code=u'AASER'
    elif club_name in (u'Ask Friidrett'):
       club_code=u'ASK'
    elif club_name in (u'Asker Fleridrettslag'):
       club_code=u'ASKFL'
    elif club_name in (u'Asker Skiklubb', u'Asker Sk. Friidrett'):
       club_code=u'ASKSK'
    elif club_name in (u'Askim Idrettsforening', u'Askim IF'):
       club_code=u'ASKIM'
    elif club_name in (u'Atna Idrettslag'):
       club_code=u'ATNA'
    elif club_name in (u'Aure Idrettslag'):
       club_code=u'AURE'
    elif club_name in (u'Aurland Idrettslag'):
       club_code=u'AURL'
    elif club_name in (u'Aurskog-Høland Friidrettslag', u'Aurskog - Høland Friidrettslag - Friidrett'):
       club_code=u'AURS'
    elif club_name in (u'Austefjord Idrettslag'):
       club_code=u'AUFJ'
    elif club_name in (u'Austevoll Idrettsklubb'):
       club_code=u'AUST'
    elif club_name in (u'Austrheim Idrettslag'):
       club_code=u'AUSTR'
    elif club_name in (u'Bagn Idrettslag'):
       club_code=u'BAGN'
    elif club_name in (u'Bakke IF'):
       club_code=u'BAKKE'
    elif club_name in (u'Balestrand Idrettslag', u'Balestrand IL'):
       club_code=u'BALE'
    elif club_name in (u'Bardu Idrettslag'):
       club_code=u'BARD'
    elif club_name in (u'Båtsfjord Sportsklubb'):
       club_code=u'BTSFJ'
    elif club_name in (u'Begnadalen Idrettslag'):
       club_code=u'BGND'
    elif club_name in (u'Beitstad Idrettslag'):
       club_code=u'BEIT'
    elif club_name in (u'Bergen Cykleklubb'):
       club_code=u'BCK'
    elif club_name in (u'Bergen Triathlon Club'):
       club_code=u'BTC'
    elif club_name in (u'Bergens Turnforening', u'Bergens TF'):
       club_code=u'BTU'
    elif club_name in (u'Berger Idrettslag'):
       club_code=u'BERGE'
    elif club_name in (u'BFG Bergen Løpeklubb'):
       club_code=u'BFGL'
    elif club_name in (u'Bjerkreim Idrettslag'):
       club_code=u'BJERR'
    elif club_name in (u'Bjerkvik Idrettsforening'):
       club_code=u'BJERV'
    elif club_name in (u'Blaker IL'):
       club_code=u'BLA'
    elif club_name in (u'Blefjell Idrettslag'):
       club_code=u'BLEFJ'
    elif club_name in (u'Bodø & Omegn IF Friidrett'):
       club_code=u'BODO'
    elif club_name in (u'Bodø Bauta Løpeklubb'):
       club_code=u'BODL'
    elif club_name in (u'Bodø Friidrettsklubb'):
       club_code=u'BODF'
    elif club_name in (u'Bokn Idrettslag'):
       club_code=u'BOKN'
    elif club_name in (u'Bossekop Ungdomslag'):
       club_code=u'BOS'
    elif club_name in (u'Botnan Idrettslag'):
       club_code=u'BOTNA'
    elif club_name in (u'Botne Skiklubb'):
       club_code=u'BOT'
    elif club_name in (u'Brandbu Idretsforening', u'Brandbu IF'):
       club_code=u'BRNDB'
    elif club_name in (u'Bratsberg Idrettslag'):
       club_code=u'BRATS'
    elif club_name in (u'Brattvåg Idrettslag'):
       club_code=u'BRATTV'
    elif club_name in (u'Breimsbygda IL'):
       club_code=u'BREIM'
    elif club_name in (u'Brekke Idrettslag'):
       club_code=u'BREKK'
    elif club_name in (u'Bremanger Idrettslag'):
       club_code=u'BREMA'
    elif club_name in (u'Bremnes Idrettslag'):
       club_code=u'BREM'
    elif club_name in (u'Brevik Idrettslag'):
       club_code=u'BRE'
    elif club_name in (u'Bromma Idrettslag'):
       club_code=u'BROMM'
    elif club_name in (u'Bryne Friidrettsklubb'):
       club_code=u'BRYF'
    elif club_name in (u'BRYNE TRIATLONKLUBB'):
       club_code=u'BRYT'
    elif club_name in (u'Bud Idrettslag'):
       club_code=u'BUD'
    elif club_name in (u'Byaasen Skiklub'):
       club_code=u'BYS'
    elif club_name in (u'Byåsen Idrettslag'):
       club_code=u'BYI'
    elif club_name in (u'Byneset IL Hovedlaget'):
       club_code=u'BYN'
    elif club_name in (u'Bøler IF'):
       club_code=u'BIF'
    elif club_name in (u'Bækkelagets SK'):
       club_code=u'BSK'
    elif club_name in (u'Bærums Verk Hauger Idrettsforening'):
       club_code=u'BRVHA'
    elif club_name in (u'Bæverfjord Idrettslag'):
       club_code=u'BVRFJ'
    elif club_name in (u'Bøler Idrettsforening'):
       club_code=u'BIF'
    elif club_name in (u'Bømlo Idrettslag'):
       club_code=u'BMLO'
    elif club_name in (u'Børsa Idrettslag', u'Børsa IL'):
       club_code=u'BRSA'
    elif club_name in (u'Dale Idrettslag'):
       club_code=u'DALE'
    elif club_name in (u'Dalen Idrettslag'):
       club_code=u'DLN'
    elif club_name in (u'Dimna IL'):
       club_code=u'DIM'
    elif club_name in (u'Dombås Idrettslag'):
       club_code=u'DMB'
    elif club_name in (u'Driv Idrettslag', u'Driv Idrettslag - Friidrett ' ):
       club_code=u'DRIV'
    elif club_name in (u'Driva IL'):
       club_code=u'DRIVA'
    elif club_name in (u'Drøbak-Frogn Idrettslag', u'Drøbak-Frogn IL'):
       club_code=u'DRFR'
    elif club_name in (u'Dypvåg Idrettsforening'):
       club_code=u'DPVG'
    elif club_name in (u'Egersunds Idrettsklubb'):
       club_code=u'EGRSU'
    elif club_name in (u'Eid Idrettslag'):
       club_code=u'EIDIL'
    elif club_name in (u'Eidanger Idrettslag'):
       club_code=u'EIDA'
    elif club_name in (u'Eidfjord Idrettslag'):
       club_code=u'EIDF'
    elif club_name in (u'Eidsberg Idrettslag'):
       club_code=u'EIDSB'
    elif club_name in (u'Eidsvåg Idrettslag'):
       club_code=u'EIDS'
    elif club_name in (u'Eidsvold Turnforening Friidrett'):
       club_code=u'EIDTU'
    elif club_name in (u'Eikanger Idrettslag'):
       club_code=u'EIK'
    elif club_name in (u'Ekeberg Sports Klubb'):
       club_code=u'ESK'
    elif club_name in (u'Espa Idrettslag'):
       club_code=u'ESPA'
    elif club_name in (u'Etne Idrettslag'):
       club_code=u'ETNE'
    elif club_name in (u'Fagernes Idrettslag N'):
       club_code=u'FAGIL'
    elif club_name in (u'Fagernes Idrettslag O', u'Fagernes IL'):
       club_code=u'FAG'
    elif club_name in (u'Falkeid idrettslag'):
       club_code=u'FALK'
    elif club_name in (u'Fana Idrettslag', u'Fana IL'):
       club_code=u'FANA'
    elif club_name in (u'Feiring Idrettslag'):
       club_code=u'FEIR'
    elif club_name in (u'Fet Friidrettsklubb'):
       club_code=u'FET'
    elif club_name in (u'FIL AKS-77'):
       club_code=u'FA77'
    elif club_name in (u'Finnøy Idrettslag'):
       club_code=u'FINNY'
    elif club_name in (u'Fiskå Idrettsforening'):
       club_code=u'FISIF'
    elif club_name in (u'Fiskå Idrettslag', u'Fiskå IL'):
       club_code=u'FISIL'
    elif club_name in (u'Fitjar Idrettslag'):
       club_code=u'FITJ'
    elif club_name in (u'Fjellhug/Vereide IL'):
       club_code=u'FJVE'
    elif club_name in (u'Flatås Idrettslag'):
       club_code=u'FLATS'
    elif club_name in (u'Florø Turn og Idrettsforening', u'Florø T og IF', u'Florø T&IF'):
       club_code=u'FLOR'
    elif club_name in (u'Follafoss Idrettslag'):
       club_code=u'FOLFO'
    elif club_name in (u'Folldal Idrettsforening'):
       club_code=u'FOL'
    elif club_name in (u'Follo Løpeklubb'):
       club_code=u'FOLLO'
    elif club_name in (u'Forra Idrettslag'):
       club_code=u'FORRA'
    elif club_name in (u'Fossum Idrettsforening', u'Fossum IF'):
       club_code=u'FOSSU'
    elif club_name in (u'Fredrikstad Idrettsforening', u'Fredrikstad IF'):
       club_code=u'FRED'
    elif club_name in (u'Freidig Sportsklubben'):
       club_code=u'FREI'
    elif club_name in (u'Friidretsklubben Orion', u'FIK Orion'):
       club_code=u'ORION'
    elif club_name in (u'Friidrettsklubben Ren-Eng'):
       club_code=u'REN'
    elif club_name in (u'Friidrettslaget Bamse'):
       club_code=u'BAMSE'
    elif club_name in (u'Friidrettslaget Borg'):
       club_code=u'BORG'
    elif club_name in (u'Friidrettslaget Frisk'):
       club_code=u'FRISK'
    elif club_name in (u'Frognerparken Idrettslag'):
       club_code=u'FRO'
    elif club_name in (u'Frol Idrettslag'):
       club_code=u'FROL'
    elif club_name in (u'Frosta Idrettslag'):
       club_code=u'FROSTA'
    elif club_name in (u'Frøyland Idrettslag'):
       club_code=u'FRLND'
    elif club_name in (u'Furuset Allidrett IF'):
       club_code=u'FUR'
    elif club_name in (u'Fyllingen Idrettslag'):
       club_code=u'FYLL'
    elif club_name in (u'Førde Idrettslag', u'Førde IL'):
       club_code=u'FRDE'
    elif club_name in (u'Gaular IL'):
       club_code=u'GAULA'
    elif club_name in (u'Gausdal Friidrettsklubb'):
       club_code=u'GAU'
    elif club_name in (u'Geilo Idrettslag'):
       club_code=u'GEI'
    elif club_name in (u'Geiranger Idrettslag'):
       club_code=u'GEIR'
    elif club_name in (u'Gjerpen Idrettsforening'):
       club_code=u'GJER'
    elif club_name in (u'Gjerstad Idrettslag'):
       club_code=u'GJERS'
    elif club_name in (u'Gjesdal Idrettslag' , 'Gjesdal IL'):
       club_code=u'GJDAL'
    elif club_name in (u'Gjøvik Friidrettsklubb'):
       club_code=u'GJFK'
    elif club_name in (u'Gjøvik Friidrettsklubb 2'):
       club_code=u'GJVIK'
    elif club_name in (u'Gloppen Friidrettslag'):
       club_code=u'GLO'
    elif club_name in (u'Gol Idrettslag'):
       club_code=u'GOL'
    elif club_name in (u'Grong Idrettslag'):
       club_code=u'GRON'
    elif club_name in (u'Groruddalen Friidrettsklubb'):
       club_code=u'GRO'
    elif club_name in (u'Grue Idrettslag'):
       club_code=u'GRUE'
    elif club_name in (u'GTI Friidrettsklubb', u'GTI Friidrettsklubb - gr.  '):
       club_code=u'GTI'
    elif club_name in (u'Gui Sportsklubb - Friidrett'):
       club_code=u'GUI'
    elif club_name in (u'Gulset Idrettsforening'):
       club_code=u'GUL'
    elif club_name in (u'HAB IL'):
       club_code=u'HAB'
    elif club_name in (u'Hadeland Friidrettsklubb'):
       club_code=u'HADE'
    elif club_name in (u'Haga Idrettsforening '):
       club_code=u'HAGA'
    elif club_name in (u'Halden Idrettslag', u'Halden IL'):
       club_code=u'HAL'
    elif club_name in (u'Halmsås & Omegn Skilag'):
       club_code=u'HALMO'
    elif club_name in (u'Halsa Idrettslag'):
       club_code=u'HALSA'
    elif club_name in (u'Hamar Idrettslag Hovedlaget', u'Hamar IL'):
       club_code=u'HIL'
    elif club_name in (u'Hannevikas Idrettslag'):
       club_code=u'HANNEV'
    elif club_name in (u'Hardbagg Idrettslag'):
       club_code=u'HARDB'
    elif club_name in (u'Hareid Idrettslag'):
       club_code=u'HAREI'
    elif club_name in (u'Harestua Idrettslag'):
       club_code=u'HARE'
    elif club_name in (u'Hattfjelldal Idrettslag', u'Hattfjelldal IL'):
       club_code=u'HATT'
    elif club_name in (u'Haugen Idrettslag'):
       club_code=u'HAUGN'
    elif club_name in (u'Haugerud Idrettsforening'):
       club_code=u'HAUGR'
    elif club_name in (u'Haugesund Idrettslag Friidrett', u'Haugesund Idrettslag Friidrett - gr'):
       club_code=u'HAUGF'
    elif club_name in (u'Haugesund Triathlon Klubb'):
       club_code=u'HAUGT'
    elif club_name in (u'Havørn Allianseidrettslag'):
       club_code=u'HAV'
    elif club_name in (u'Heggedal Friidrettsklubb'):
       club_code=u'HEGGF'
    elif club_name in (u'Heggedal Idrettslag'):
       club_code=u'HEGGI'
    elif club_name in (u'Hell Ultraløperklubb'):
       club_code=u'HELLU'
    elif club_name in (u'Heming Idrettslaget', u'IL Heming'):
       club_code=u'HEM'
    elif club_name in (u'Henning I L'):
       club_code=u'HENN'
    elif club_name in (u'Herand Idrettslag'):
       club_code=u'HERA'
    elif club_name in (u'Herkules Friidrett'):
       club_code=u'HERK'
    elif club_name in (u'Herøy Idrettslag'):
       club_code=u'HERY'
    elif club_name in (u'Hinna Friidrett'):
       club_code=u'HIN'
    elif club_name in (u'Hitra Friidrettsklubb'):
       club_code=u'HITF'
    elif club_name in (u'Hitra Løpeklubb'):
       club_code=u'HITL'
    elif club_name in (u'Hobøl Idrettslag'):
       club_code=u'HOB'
    elif club_name in (u'Hof Idrettslag'):
       club_code=u'HOF'
    elif club_name in (u'Hol Idrettslag'):
       club_code=u'HOL'
    elif club_name in (u'Holmemstranda Idrettslag'):
       club_code=u'HOLMS'
    elif club_name in (u'Holum Idrettslag'):
       club_code=u'HOLUM'
    elif club_name in (u'Hommelvik Idrettslag', u'Hommelvik IL'):
       club_code=u'HMLV'
    elif club_name in (u'Hope Idrettslag'):
       club_code=u'HOPE'
    elif club_name in (u'Hornindal Idrettslag'):
       club_code=u'HORNI'
    elif club_name in (u'Horten Friidrettsklubb'):
       club_code=u'HORFR'
    elif club_name in (u'Huglo Idrettslag'):
       club_code=u'HUG'
    elif club_name in (u'Hurdal Idrettslag', u'Hurdal IL'):
       club_code=u'HURD'
    elif club_name in (u'Hvam Idrettslag', u'Hvam IL'):
       club_code=u'HVAM'
    elif club_name in (u'Hvittingfoss Idrettslag'):
       club_code=u'HVFO'
    elif club_name in (u'Hyen Idrettslag'):
       club_code=u'HYEN'
    elif club_name in (u'Hyllestad Idrettslag'):
       club_code=u'HYLLS'
    elif club_name in (u'Høybråten og Stovner IL'):
       club_code=u'HSI'
    elif club_name in (u'Høydalsmo Idrottslag'):
       club_code=u'HDMO'
    elif club_name in (u'I.l Fjellørnen'):
       club_code=u'FJELLO'
    elif club_name in (u'I.L. Framsteg'):
       club_code=u'FRAMS'
    elif club_name in (u'I.L. Norna Salhus', u'Norna-Salhus IL'):
       club_code=u'NORSA'
    elif club_name in (u'I.L. Nybrott'):
       club_code=u'NYBR'
    elif club_name in (u'Idd Sportsklubb'):
       club_code=u'IDD'
    elif club_name in (u'Idrettsforeningen Birkebeineren'):
       club_code=u'BIRK'
    elif club_name in (u'Idrettsforeningen Fram'):
       club_code=u'FRAM'
    elif club_name in (u'Idrettsforeningen Hellas', u'IF Hellas'):
       club_code=u'HELLA'
    elif club_name in (u'Idrettsforeningen Njaal'):
       club_code=u'NJAAL'
    elif club_name in (u'Idrettsforeningen Sturla', u'Sturla IF'):
       club_code=u'STUR'
    elif club_name in (u'Idrettsforeningen Ørn'):
       club_code=u'IFORN'
    elif club_name in (u'Idrettslaget Bjarg'):
       club_code=u'BJARG'
    elif club_name in (u'Idrettslaget Bjørn'):
       club_code=u'ILBJ'
    elif club_name in (u'Idrettslaget Dalebrand'):
       club_code=u'DLBR'
    elif club_name in (u'Idrettslaget Dyre Vaa'):
       club_code=u'DYREV'
    elif club_name in (u'Idrettslaget Express'):
       club_code=u'EXPR'
    elif club_name in (u'Idrettslaget Forsøk'):
       club_code=u'FORSK'
    elif club_name in (u'Idrettslaget Fri'):
       club_code=u'FRI'
    elif club_name in (u'Idrettslaget Gneist' u'IL Gneist'):
       club_code=u'GNE'
    elif club_name in (u'Idrettslaget Holeværingen'):
       club_code=u'HOLE'
    elif club_name in (u'Idrettslaget I Bondeungdomslaget I Tromsø'):
       club_code=u'BULT'
    elif club_name in (u'Idrettslaget Ilar'):
       club_code=u'ILAR'
    elif club_name in (u'Idrettslaget Ivrig'):
       club_code=u'IVRIG'
    elif club_name in (u'Idrettslaget Jardar', u'IL Jardar'):
       club_code=u'JARD'
    elif club_name in (u'IL Jutul', u'Idrettslaget Jutul'):
       club_code=u'JUT'
    elif club_name in (u'Idrettslaget Ros'):
       club_code=u'ILROS'
    elif club_name in (u'Idrettslaget Runar', u'IL Runar'):
       club_code=u'RUNAR'
    elif club_name in (u'Idrettslaget Sand'):
       club_code=u'ILSAN'
    elif club_name in (u'Idrettslaget Sandvin', u'IL Sandvin'):
       club_code=u'SANDV'
    elif club_name in (u'Idrettslaget Skade'):
       club_code=u'SKADE'
    elif club_name in (u'Idrettslaget Skjalg'):
       club_code=u'SKJA'
    elif club_name in (u'Idrettslaget Syril'):
       club_code=u'SYR'
    elif club_name in (u'Idrettslaget Trysilgutten'):
       club_code=u'TRY'
    elif club_name in (u'Idrottslaget Gular Bygdeungdomen I Bergen', u'IL Gular'):
       club_code=u'GULA'
    elif club_name in (u'IDROTTSLAGET I BUL', u'IL i BUL'):
       club_code=u'ILBUL'
#   elif club_name in (u'IDROTTSLAGET I BUL 2'):
#      club_code=u'ILBUL'
    elif club_name in (u'Idrottslaget Jotun', u'Jotun IL'):
       club_code=u'JOT'
    elif club_name in (u'Idun Idrettslag'):
       club_code=u'IDUN'
    elif club_name in (u'IF Eiker-Kvikk', u'If Eiker Kvikk'):
       club_code=u'EIKKV'
    elif club_name in (u'IF Kamp/Vestheim', u'Kamp/Vestheim IF'):
       club_code=u'KAVE'
    elif club_name in (u'If Klypetussen'):
       club_code=u'KLYP'
    elif club_name in (u'Ik Grane Arendal Friidrett'):
       club_code=u'GRANE'
    elif club_name in (u'IK Hind', u'IK Hind '):
       club_code=u'HIND'
    elif club_name in (u'Ikornnes Idrettslag'):
       club_code=u'IKORN'
    elif club_name in (u'IL Aasguten'):
       club_code=u'AASG'
    elif club_name in (u'IL Alvidra'):
       club_code=u'ALVI'
    elif club_name in (u'IL Bever`n'):
       club_code=u'ILBEV'
    elif club_name in (u'IL Brodd'):
       club_code=u'BRODD'
    elif club_name in (u'IL Flåværingen ', u'Flåværingen IL '):
       club_code=u'FLV'
    elif club_name in (u'IL Gry'):
       club_code=u'GRY'
    elif club_name in (u'IL Norodd'):
       club_code=u'ILNOR'
    elif club_name in (u'IL Pioner Friidrett'):
       club_code=u'PIO'
    elif club_name in (u'IL Polarstjernen'):
       club_code=u'POL'
    elif club_name in (u'IL Samhald'):
       club_code=u'SAMH'
    elif club_name in (u'IL Santor'):
       club_code=u'SANT'
    elif club_name in (u'IL Stålkameratene'):
       club_code=u'STKAM'
    elif club_name in (u'IL Tambarskjelvar'):
       club_code=u'TAMBA'
    elif club_name in (u'IL Triumf'):
       club_code=u'TRIUM'
    elif club_name in (u'Vestby IL'):
       club_code=u'VESTB'
    elif club_name in (u'Il Vindbjart'):
       club_code=u'VIND'
    elif club_name in (u'IL Vinger'):
       club_code=u'VING'
    elif club_name in (u'Inderøy Idrettslag'):
       club_code=u'INDRY'
    elif club_name in (u'Innstranda IL'):
       club_code=u'INN'
    elif club_name in (u'International School of Stavanger'):
       club_code=u'INSTA'
    elif club_name in (u'Isfjorden Idrettslag'):
       club_code=u'ISFJO'
    elif club_name in (u'Jondalen Idrettslag'):
       club_code=u'JOND'
    elif club_name in (u'Jægervatnet Idrettslag'):
       club_code=u'JVTN'
    elif club_name in (u'Jøa Idrettslag'):
       club_code=u'JIL'
    elif club_name in (u'Jølster Idrettslag', u'Jølster IL'):
       club_code=u'JLSTE'
    elif club_name in (u'Kaupanger Idrettslag'):
       club_code=u'KAUP'
    elif club_name in (u'Kfum-kameratene Oslo'):
       club_code=u'KFUM'
    elif club_name in (u'Kjelsås Idrettslag', u'Kjelsås IL'):
       club_code=u'KJ'
    elif club_name in (u'Klepp Idrettslag'):
       club_code=u'KLPP'
    elif club_name in (u'Klæbu Løpeklubb'):
       club_code=u'KLK'
    elif club_name in (u'Kløfta Idrettslag'):
       club_code=u'KLIL'
    elif club_name in (u'Kolbukameratene I L'):
       club_code=u'KLBK'
    elif club_name in (u'Koll Idrettslaget', 'IL Koll'):
       club_code=u'KOLL'
    elif club_name in (u'Kolvereid Idrettslag'):
       club_code=u'KLVIL'
    elif club_name in (u'Kongsberg Idrettsforening', u'Kongsberg IF'):
       club_code=u'KNGSB'
    elif club_name in (u'Kongsvinger IL Friidrett'):
       club_code=u'KNGSV'
    elif club_name in (u'Konnerud IL Friidrett'):
       club_code=u'KONN'
    elif club_name in (u'Kopervik Idrettslag'):
       club_code=u'KOP'
    elif club_name in (u'Korgen Idrettslag'):
       club_code=u'KORG'
    elif club_name in (u'Kragerø IF Friidrett'):
       club_code=u'KRAG'
    elif club_name in (u'Kråkerøy Idrettslag'):
       club_code=u'KRAAK'
    elif club_name in (u'Kråkstad Idrettslag'):
       club_code=u'KRSTD'
    elif club_name in (u'Kristiansand Løpeklubb'):
       club_code=u'KRL'
    elif club_name in (u'Kristiansands Idrettsforening Friidrett', u'Kristiansands IF'):
       club_code=u'KIF'
    elif club_name in (u'Krødsherad Idrettslag'):
       club_code=u'KRHER'
    elif club_name in (u'Kvinesdal Idrettslag'):
       club_code=u'KVINES'
    elif club_name in (u'Kvæfjord Idrettslag'):
       club_code=u'KVFJ'
    elif club_name in (u'Kyrksæterøra Idrettslag Kil'):
       club_code=u'KYRK'
    elif club_name in (u'Laksevåg Turn og Idrettslag', u'Laksevåg TIL'):
       club_code=u'LAKS'
    elif club_name in (u'Lalm Idrettslag'):
       club_code=u'LALM'
    elif club_name in (u'Lambertseter IF'):
       club_code=u'LAM'
    elif club_name in (u'Langesund Sykle- og triathlonklubb'):
       club_code=u'LANGS'
    elif club_name in (u'Lånke Idrettslag'):
       club_code=u'LNKEIL'
    elif club_name in (u'Larvik Turn & Idrettsforening', u'Larvik Turn & IF'):
       club_code=u'LRVK'
    elif club_name in (u'Leinstrand Idrettslag'):
       club_code=u'LEINS'
    elif club_name in (u'Lena Idrettsforening'):
       club_code=u'LENA'
    elif club_name in (u'Lierne Idrettslag'):
       club_code=u'LIERN'
    elif club_name in (u'Lillehammer Idrettsforening', u'Lillehammer IF'):
       club_code=u'LIF'
    elif club_name in (u'Lillesand Idrettslag'):
       club_code=u'LILLS'
    elif club_name in (u'Lista Idrettslag'):
       club_code=u'LISTA'
    elif club_name in (u'Loddefjord IL'):
       club_code=u'LODD'
    elif club_name in (u'Lofoten Triatlonklubb'):
       club_code=u'LFTR'
    elif club_name in (u'Lom Idrettslag'):
       club_code=u'LOM'
    elif club_name in (u'Lundamo Idrettslag'):
       club_code=u'LUND'
    elif club_name in (u'Lundehøgda IL'):
       club_code=u'LUNDH'
    elif club_name in (u'Luster Idrettslag'):
       club_code=u'LUST'
    elif club_name in (u'Lye Idrettslag'):
       club_code=u'LYE'
    elif club_name in (u'Lyn Ski'):
       club_code=u'LYN'
    elif club_name in (u'Lyngdal Idrettslag', u'Lyngdal IL'):
       club_code=u'LNGD'
    elif club_name in (u'Lyngen/ Karnes Il'):
       club_code=u'LYKA'
    elif club_name in (u'Lyngstad og Omegn Idrettslag'):
       club_code=u'LYNGO'
    elif club_name in (u'Lørenskog Friidrettslag'):
       club_code=u'LRSKG'
    elif club_name in (u'Løten Friidrett'):
       club_code=u'LFK'
    elif club_name in (u'Løten Friidrett 2'):
       club_code=u'LTN'
    elif club_name in (u'Malm IL'):
       club_code=u'MALM'
    elif club_name in (u'Målselv Idrettslag'):
       club_code=u'MLSEL'
    elif club_name in (u'Malvik Idrettslag'):
       club_code=u'MALV'
    elif club_name in (u'Måløy Idrettslag Hovedstyre'):
       club_code=u'MAAL'
    elif club_name in (u'Mandal & Halse I.l.'):
       club_code=u'MAHA'
    elif club_name in (u'Måndalen Idrettslag'):
       club_code=u'MNDL'
    elif club_name in (u'Markabygda Idrettslag'):
       club_code=u'MABY'
    elif club_name in (u'Markane Idrettslag', u'Markane IL'):
       club_code=u'MARKA'
    elif club_name in (u'Marnardal Idrettslag'):
       club_code=u'MARNA'
    elif club_name in (u'Medkila Skilag'):
       club_code=u'MEDKI'
    elif club_name in (u'Meldal Idrettslag'):
       club_code=u'MELD'
    elif club_name in (u'Melhus Idrettslag'):
       club_code=u'MELHU'
    elif club_name in (u'Midsund Idrettslag'):
       club_code=u'MDSND'
    elif club_name in (u'Mjøsdalen IL'):
       club_code=u'MJSD'
    elif club_name in (u'Modum Friidrettsklubb'):
       club_code=u'MOD'
    elif club_name in (u'Moelven Idrettslag', u'Moelven IL'):
       club_code=u'MOELV'
    elif club_name in (u'Moi Idrettslag'):
       club_code=u'MOI'
    elif club_name in (u'Molde og Omegn Idrettsforening'):
       club_code=u'MOLDE'
    elif club_name in (u'Molde Olymp'):
       club_code=u'OLYMP'
    elif club_name in (u'Moltustranda Idrettslag'):
       club_code=u'MOITU'
    elif club_name in (u'Mosjøen Friidrettsklubb'):
       club_code=u'MOSJ'
    elif club_name in (u'Moss Idrettslag', u'Moss IL'):
       club_code=u'MOSS'
    elif club_name in (u'Mosvik Idrettslag', 'Mosvik IL - Friidrett'):
       club_code=u'MOSV'
    elif club_name in (u'MUIL - Mefjordvær Ungdoms- og Idrettslag'):
       club_code=u'MUIL'
    elif club_name in (u'Namdal løpeklubb'):
       club_code=u'NAML'
    elif club_name in (u'Namdalseid Idrettslag'):
       club_code=u'NAMDA'
    elif club_name in (u'Namsen Fif'):
       club_code=u'NAMSE'
    elif club_name in (u'Nannestad Idrettslag'):
       club_code=u'NANN'
    elif club_name in (u'Narvik Idrettslag'):
       club_code=u'NAR'
    elif club_name in (u'Nesbyen Idrettslag', u'Nesbyen IL Friidrett'):
       club_code=u'NESB'
    elif club_name in (u'Nesodden IF'):
       club_code=u'NESO'
    elif club_name in (u'Nesøya Idrettslag', u'Nesøya IL'):
       club_code=u'NES'
    elif club_name in (u'Nidelv Idrettslag'):
       club_code=u'NID'
    elif club_name in (u'Nissedal Idrettslag'):
       club_code=u'NISS'
    elif club_name in (u'Nittedal Idrettslag', u'Nittedal IL'):
       club_code=u'NITT'
    elif club_name in (u'Nordkjosbotn Idrettslag'):
       club_code=u'NRDKJ'
    elif club_name in (u'Nordre Eidsvoll Idrettslag'):
       club_code=u'NEIDS'
    elif club_name in (u'Nordre Fjell Friidrett'):
       club_code=u'NFJEL'
    elif club_name in (u'Nordre Land Idrettslag', u'Nordre Land IL'):
       club_code=u'NLAND'
    elif club_name in (u'Nordre Trysil IL'):
       club_code=u'NTRY'
    elif club_name in (u'Nordøy Idrettslag'):
       club_code=u'NORIL'
    elif club_name in (u'Norrøna IL'):
       club_code=u'NORR'
    elif club_name in (u'Northern Runners'):
       club_code=u'NRUN'
    elif club_name in (u'NTNUI - Norges Teknisk-Naturvitenskapelige Universitets Idrettsforening'):
       club_code=u'NTNUI'
    elif club_name in (u'Nydalens Skiklub'):
       club_code=u'NYSK'
    elif club_name in (u'Nykirke Idrettsforening'):
       club_code=u'NYKIR'
    elif club_name in (u'Nøtterøy Idrettsforening'):
       club_code=u'NTTRY'
    elif club_name in (u'Odda Idrettslag'):
       club_code=u'ODDA'
    elif club_name in (u'Ogndal Idrettslag Hovedlaget'):
       club_code=u'OGND'
    elif club_name in (u'Olden Idrettslag'):
       club_code=u'OLD'
    elif club_name in (u'Olderdalen Idrettsklubb'):
       club_code=u'OLDA'
    elif club_name in (u'Oppdal IL Hovedlaget'):
       club_code=u'OPPD'
    elif club_name in (u'Oppegård Idrettslag', u'Oppegård IL'):
       club_code=u'OPP'
    elif club_name in (u'Oppsal Idrettsforening', u'Oppsal IF'):
       club_code=u'OPSL'
    elif club_name in (u'Oppstad Idrettslag'):
       club_code=u'OPST'
    elif club_name in (u'Oppstad Idrettslag 2'):
       club_code=u'OPPST'
    elif club_name in (u'Oppstryn Idrettslag'):
       club_code=u'OPSTR'
    elif club_name in (u'Opptur Motbakkeklubb'):
       club_code=u'OPPT'
    elif club_name in (u'Orkanger Idrettsforening'):
       club_code=u'ORKA'
    elif club_name in (u'Orkdal Idrettslag'):
       club_code=u'ORKD'
    elif club_name in (u'Orre Idrettslag'):
       club_code=u'ORRE'
    elif club_name in (u'Os Idrettslag'):
       club_code=u'OS'
    elif club_name in (u'Os Turnforening'):
       club_code=u'OSTU'
    elif club_name in (u'OSI Friidrett'):
       club_code=u'FRII'
    elif club_name in (u'Oslo Politis Idrettslag', u'Oslo Politis IL'):
       club_code=u'POLIT'
    elif club_name in (u'Oslostudentenes Idrettsklubb', 'Oslostudentenes IK'):
       club_code=u'OSI'
    elif club_name in (u'Osterøy Idrottslag', u'Osterøy IL'):
       club_code=u'OST'
    elif club_name in (u'Otra IL'):
       club_code=u'OTRA'
    elif club_name in (u'Ottestad Idrettslag'):
       club_code=u'OTTE'
    elif club_name in (u'Ottestad Kast og Styrkeløft'):
       club_code=u'OTKS'
    elif club_name in (u'Overhalla Idrettslag', u'Overhalla IL'):
       club_code=u'OVRH'
    elif club_name in (u'Porsanger Idrettslag'):
       club_code=u'PORS'
    elif club_name in (u'Rana Friidrettsklubb'):
       club_code=u'RANA'
    elif club_name in (u'Ranheim Idrettslag', u'Ranheim IL'):
       club_code=u'RAN'
    elif club_name in (u'Raufoss IL Friidrett', 'Raufoss Friidrett'):
       club_code=u'RAU'
    elif club_name in (u'Raumnes & Årnes Idrettslag', u'Raumnes & Årnes IL'):
       club_code=u'RAUM'
    elif club_name in (u'Re Friidrettsklubb'):
       club_code=u'RE'
    elif club_name in (u'Ready Idrettsforeningen'):
       club_code=u'READY'
    elif club_name in (u'Rena Idrettslag'):
       club_code=u'RENA'
    elif club_name in (u'Rendalen Idrettslag'):
       club_code=u'RENDA'
    elif club_name in (u'Rennebu Idrettslag'):
       club_code=u'RENB'
    elif club_name in (u'Rindal Idrettslag', u'Rindal IL'):
       club_code=u'RIND'
    elif club_name in (u'Ringerike Friidrettsklubb'):
       club_code=u'RING'
    elif club_name in (u'Risør Idrettslag'):
       club_code=u'RIS'
    elif club_name in (u'Rjukan Idrettslag'):
       club_code=u'RJU'
    elif club_name in (u'Rogne Idrettslag'):
       club_code=u'ROGNE'
    elif club_name in (u'Romerike Friidrett'):
       club_code=u'ROMFR'
    elif club_name in (u'Romerike Ultraløperklubb'):
       club_code=u'ROMUL'
    elif club_name in (u'Romsdal Randoneklubb'):
       club_code=u'ROMRA'
    elif club_name in (u'Rosendal Turnlag', u'Rosendal TL'):
       club_code=u'ROSEN'
    elif club_name in (u'Royal Sport'):
       club_code=u'ROYAL'
    elif club_name in (u'Rustad Idrettslag', u'Rustad IL'):
       club_code=u'RUS'
    elif club_name in (u'Rygge Idrettslag', u'Rygge IL'):
       club_code=u'RYGGE'
    elif club_name in (u'Røa Allianseidrettslag'):
       club_code=u'RIL'
    elif club_name in (u'Røldal Idrettslag'):
       club_code=u'RDIL'
    elif club_name in (u'Røros Idrettslag'):
       club_code=u'ROSIL'
    elif club_name in (u'Røyken UIL'):
       club_code=u'RKEN'
    elif club_name in (u'Salangen IF Friidrett', u'Salangen IF - Friidrett'):
       club_code=u'SALA'
    elif club_name in (u'Samnanger Idrettslag'):
       club_code=u'SAMN'
    elif club_name in (u'Sandane Turn og Idrettslag'):
       club_code=u'SANTU'
    elif club_name in (u'Sandefjord Turn & Idrettsforening', u'SANDEFJORD TURN & IDRETTSFORENING'):
       club_code=u'STIF'
    elif club_name in (u'Sandnes Idrettslag', u'Sandnes IL'):
       club_code=u'SAND'
    elif club_name in (u'Sandnes Idrettslag 2'):
       club_code=u'SNDI'
    elif club_name in (u'Sandnessjøen Idrettslag', u'Sandnessjøen IL'):
       club_code=u'SNDSJ'
    elif club_name in (u'Sarpsborg Allianseidrettslag', u'Sarpsborg IL'):
       club_code=u'SARP'
    elif club_name in (u'Sauda Idrettslag'):
       club_code=u'SAUD'
    elif club_name in (u'Sauland Idrettslag'):
       club_code=u'SAUL'
    elif club_name in (u'Selbu IL'):
       club_code=u'SELB'
    elif club_name in (u'Selje Idrettslag'):
       club_code=u'SELJE'
    elif club_name in (u'Seljord Idrettslag'):
       club_code=u'SELJO'
    elif club_name in (u'Selsbakk Idrettsforening'):
       club_code=u'SELS'
    elif club_name in (u'Sem Idrettsforening', u'Sem IF'):
       club_code=u'SEM'
    elif club_name in (u'Sigdal Friidrettsklubb'):
       club_code=u'SIGFR'
    elif club_name in (u'Sigdals Skiklub'):
       club_code=u'SIGSK'
    elif club_name in (u'Siljan Idrettslag'):
       club_code=u'SILJ'
    elif club_name in (u'Sirma Il'):
       club_code=u'SIRMA'
    elif club_name in (u'Sjetne Idrettslag'):
       club_code=u'SJET'
    elif club_name in (u'Sk Vedavåg Karmøy'):
       club_code=u'VEDA'
    elif club_name in (u'SK Vidar'):
       club_code=u'VID'
    elif club_name in (u'Skagerrak Sportsklubb'):
       club_code=u'SKAGE'
    elif club_name in (u'Skåla Idrettslag'):
       club_code=u'SKLA'
    elif club_name in (u'Skarphedin IL', u'Idrettslaget Skarphedin'):
       club_code=u'SKRPH'
    elif club_name in (u'Skaubygda Il'):
       club_code=u'SKAU'
    elif club_name in (u'Skaun Idrettslag'):
       club_code=u'SKAUN'
    elif club_name in (u'Ski IL Friidrett'):
       club_code=u'SKI'
    elif club_name in (u'Skjåk IL'):
       club_code=u'SKJK'
    elif club_name in (u'Skjoldar Il'):
       club_code=u'SKJO'
    elif club_name in (u'Skogn Idrettslag'):
       club_code=u'SKOGN'
    elif club_name in (u'Skotterud Idrettslag'):
       club_code=u'SKO'
    elif club_name in (u'Skreia Idrettslag'):
       club_code=u'SKREIA'
    elif club_name in (u'Snåsa Idrettslag'):
       club_code=u'SNSA'
    elif club_name in (u'Snøgg Friidrett'):
       club_code=u'SNGG'
    elif club_name in (u'Sogndal Idrettslag'):
       club_code=u'SIL'
    elif club_name in (u'Sokndal Friidrettsklubb'):
       club_code=u'SOKND'
    elif club_name in (u'Sola Friidrett'):
       club_code=u'SOLA'
    elif club_name in (u'Solnut IL'):
       club_code=u'SOLN'
    elif club_name in (u'Sortland Friidrettsklubb'):
       club_code=u'SORTL'
    elif club_name in (u'Sotra Sportsklubb'):
       club_code=u'SOT'
    elif club_name in (u'Spillum Idrettslag'):
       club_code=u'SPILL'
    elif club_name in (u'Spiridon Langløperlag'):
       club_code=u'SPRD'
    elif club_name in (u'Spjelkavik og Omegn Friidrettsklubb'):
       club_code=u'SPJVK'
    elif club_name in (u'Sportsklubben Kraft'):
       club_code=u'KRAFT'
    elif club_name in (u'Sportsklubben Rye', u'Rye Sp.Kl.'):
       club_code=u'RYE'
    elif club_name in (u'Sportsklubben Vidar'):
       club_code=u'VIDAR'
    elif club_name in (u'Spydeberg IL'):
       club_code=u'SPYDE'
    elif club_name in (u'Staal Jørpeland IL'):
       club_code=u'STJIL'
    elif club_name in (u'Stadsbygd IL'):
       club_code=u'STAD'
    elif club_name in (u'Stårheim Idrettslag', u'Stårheim IL'):
       club_code=u'STRHE'
    elif club_name in (u'Stavanger Døve-Idrettsforening'):
       club_code=u'STDIF'
    elif club_name in (u'Stavanger Friidrettsklubb'):
       club_code=u'STAVA'
    elif club_name in (u'Stavanger Idrettsforening Allianseidrettslag - Friidrett'):
       club_code=u'STAVIF'
    elif club_name in (u'Stegaberg Idrettslag'):
       club_code=u'STEGA'
    elif club_name in (u'Stein Friidrettsklubb'):
       club_code=u'STEIN'
    elif club_name in (u'Steinkjer Friidrettsklubb'):
       club_code=u'STEKJ'
    elif club_name in (u'Stettevik Sportsklubb'):
       club_code=u'STSK'
    elif club_name in (u'Stjørdal Fridrettsklubb', u'Stjørdal Friidrettsklubb'):
       club_code=u'STJF'
    elif club_name in (u'Stjørdal Paraidrettslag'):
       club_code=u'STJP'
    elif club_name in (u'Stjørdals-Blink IL'):
       club_code=u'STJB'
    elif club_name in (u'Stokke Idrettslag'):
       club_code=u'STOKK'
    elif club_name in (u'Stokmarknes Idrettslag'):
       club_code=u'STOKM'
    elif club_name in (u'Stord Idrettslag', u'Stord IL (Allianseidrettslag)'):
       club_code=u'STO'
    elif club_name in (u'Storfjord idrettslag'):
       club_code=u'STOFJ'
    elif club_name in (u'Stranda Idrottslag'):
       club_code=u'STRIL'
    elif club_name in (u'Strandebarm Idrettslag'):
       club_code=u'STRAB'
    elif club_name in (u'Stranden Idrettslag'):
       club_code=u'STRA'
    elif club_name in (u'Straumsnes Idrettslag'):
       club_code=u'STRAU'
    elif club_name in (u'Strindheim Idrettslag', u'Strindheim il'):
       club_code=u'STRI'
    elif club_name in (u'Stryn Turn og Idrettslag'):
       club_code=u'STRY'
    elif club_name in (u'Støren Sportsklubb'):
       club_code=u'STREN'
    elif club_name in (u'Sunndal IL Friidrett'):
       club_code=u'SUNND'
    elif club_name in (u'Surnadal Idrettslag'):
       club_code=u'SURN'
    elif club_name in (u'Svalbard Turn Idrettslag'):
       club_code=u'SVTU'
    elif club_name in (u'Svarstad Idrettslag', u'Svarstad IL'):
       club_code=u'SVARS'
    elif club_name in (u'Sveio Idrettslag'):
       club_code=u'SVEIO'
    elif club_name in (u'Svelgen Turn og Idrettsforening'):
       club_code=u'SVEL'
    elif club_name in (u'Svint IL'):
       club_code=u'SVINT'
    elif club_name in (u'SVORKMO N.O.I.', u'Svorkmo N.O.I.'):
       club_code=u'SVORK'
    elif club_name in (u'Sykkylven Idrottslag'):
       club_code=u'SYKK'
    elif club_name in (u'Sylling Idrettsforening'):
       club_code=u'SYLL'
    elif club_name in (u'Sædalen Idrettslag'):
       club_code=u'SDAL'
    elif club_name in (u'Sætre Idrætsforening Graabein'):
       club_code=u'GRAA'
    elif club_name in (u'Søfteland Turn & Idrettslag'):
       club_code=u'STIL'
    elif club_name in (u'Søgne Idrettslag'):
       club_code=u'SGNE'
    elif club_name in (u'Sømna Idrettslag', u'Sømna IL'):
       club_code=u'SMNA'
    elif club_name in (u'Søndre Land IL Friidrett', u'Friidrett Søndre Land IL'):
       club_code=u'SNDLA'
    elif club_name in (u'Søre Ål Idrettslag'):
       club_code=u'SAAL'
    elif club_name in (u'Sørild Fridrettsklubb', u'Sørild FIK'):
       club_code=u'SRILD'
    elif club_name in (u'Sørkedalens Idrettsforening'):
       club_code=u'SRKDL'
    elif club_name in (u'T I L Hovding'):
       club_code=u'HOVD'
    elif club_name in (u'Tamil Sangam IL'):
       club_code=u'TAMSAN'
    elif club_name in (u'Tistedalen FL'):
       club_code=u'TIST'
    elif club_name in (u'Tingvoll Friidrettsklubb', u'Tingvoll Friidrettskl.'):
       club_code=u'TING'
    elif club_name in ( 'IK Tjalve', 'Idrettsklubben Tjalve', 'Tjalve, IK', 'Tjalve, Idrettsklubben', 'Tjalve Idrettsklubben' ):
       club_code=u'TJAL'
    elif club_name in (u'Tjølling Idrettsforening'):
       club_code=u'TJØLL'
    elif club_name in (u'Tjøme Idrettslag'):
       club_code=u'TJI'
    elif club_name in (u'Tjøme Løpeklubb'):
       club_code=u'TJL'
    elif club_name in (u'Tolga Idrettslag'):
       club_code=u'TOL'
    elif club_name in (u'Tomrefjord Idrettslag'):
       club_code=u'TOMR'
    elif club_name in (u'Torodd IF'):
       club_code=u'TORO'
    elif club_name in (u'Torvikbukt Idrettslag'):
       club_code=u'TORVI'
    elif club_name in (u'Treungen Idrettslag'):
       club_code=u'TREU'
    elif club_name in (u'Trio idrettslag'):
       club_code=u'TRIO'
    elif club_name in (u'Tromsø Friidrettsklubb'):
       club_code=u'TRF'
    elif club_name in (u'Tromsø Løpeklubb'):
       club_code=u'TRL'
    elif club_name in (u'Tromsø Svømmeklubb'):
       club_code=u'TRS'
    elif club_name in (u'Trondheim & Omegn Sportsklubb'):
       club_code=u'TROO'
    elif club_name in (u'Trondheim Friidrett', u'Trondheim Friidrett - Friidrett'):
       club_code=u'TROF'
    elif club_name in (u'Trøgstad Skiklubb'):
       club_code=u'TSK'
    elif club_name in (u'TUIL Tromsdalen Friidrett'):
       club_code=u'TUIL'
    elif club_name in (u'Tvedestrand Turn & Idrettsforening'):
       club_code=u'TVEDE'
    elif club_name in (u'Tyrving Idrettslag', u'Tyrving IL'):
       club_code=u'TYR'
    elif club_name in (u'Tønsberg Friidrettsklubb'):
       club_code=u'TNSBF'
    elif club_name in (u'Tørvikbygd Idrettslag'):
       club_code=u'TRBIL'
    elif club_name in (u'Tøyen Sportsklubb'):
       club_code=u'TYEN'
    elif club_name in (u'Ullensaker/Kisa IL Friidrett'):
       club_code=u'ULLK'
    elif club_name in (u'Ullensaker/Kisa IL Friidrett 2'):
       club_code=u'ULKI'
    elif club_name in (u'Undheim Idrettslag'):
       club_code=u'UND'
    elif club_name in (u'Urædd Friidrett'):
       club_code=u'URFRI'
    elif club_name in (u'Utleira Idrettslag'):
       club_code=u'UTL'
    elif club_name in (u'Vaaler Idrettsforening'):
       club_code=u'VAAL'
    elif club_name in (u'Vadsø Atletklubb'):
       club_code=u'VA'
    elif club_name in (u'Vadsø Turnforening (Vtf)'):
       club_code=u'VTF'
    elif club_name in (u'Vågå Idrettslag'):
       club_code=u'VGAA'
    elif club_name in (u'Vågstranda Idrettslag'):
       club_code=u'VIL'
    elif club_name in (u'Valkyrien Idrettslag'):
       club_code=u'VALK'
    elif club_name in (u'Valldal Idrettslag'):
       club_code=u'VALL'
    elif club_name in (u'Vallset IL'):
       club_code=u'VAL'
    elif club_name in (u'Varegg Fleridrett'):
       club_code=u'VAR'
    elif club_name in (u'Varhaug Idrettslag'):
       club_code=u'VARH'
    elif club_name in (u'Varteig Idrettslag'):
       club_code=u'VART'
    elif club_name in (u'Vegårshei Idrettslag'):
       club_code=u'VEG'
    elif club_name in (u'Veldre Friidrett'):
       club_code=u'VELD'
    elif club_name in (u'Velledalen Idrettslag'):
       club_code=u'VELL'
    elif club_name in (u'Verdal Friidrettsklubb'):
       club_code=u'VERD'
    elif club_name in (u'Vestby Idrettslag'):
       club_code=u'VESTB'
    elif club_name in (u'Vestfossen Idrettsforening', u'Vestfossen IF'):
       club_code=u'VESTF'
    elif club_name in (u'Vestre Spone IF'):
       club_code=u'VSPON'
    elif club_name in (u'Vik Idrettslag', u'Vik IL'):
       club_code=u'VIKIL'
    elif club_name in (u'Vikane IL'):
       club_code=u'VIKAN'
    elif club_name in (u'Viking Turn og Idrettsforening', u'TIF Viking'):
       club_code=u'VIK'
    elif club_name in (u'Viksdalen Idrettslag'):
       club_code=u'VIKSD'
    elif club_name in (u'Viljar IL'):
       club_code=u'VILJ'
    elif club_name in (u'Vind Idrettslag'):
       club_code=u'VNDIL'
    elif club_name in (u'Vindafjord Idrettslag'):
       club_code=u'VINDA'
    elif club_name in (u'Vinje Idrottslag'):
       club_code=u'VINJE'
    elif club_name in (u'Vollan Idrettsklubb', u'Vollan I.K.'):
       club_code=u'VOLL'
    elif club_name in (u'Voss Idrottslag'):
       club_code=u'VOSS'
    elif club_name in (u'Ytterøy Idrettslag'):
       club_code=u'YTTER'
    elif club_name in (u'Ørje Idrettslag'):
       club_code=u'ORJIL'
    elif club_name in (u'Ørsta Idrettslag', u'Ørsta IL'):
       club_code=u'ORSTA'
    elif club_name in (u'Østmarka Marsjklubb'):
       club_code=u'OMARSJ'
    elif club_name in (u'Øyer/Tretten Idrettsforening'):
       club_code=u'OTRET'
    elif club_name in (u'Øystre Slidre Idrettslag'):
       club_code=u'OSLID'
    else:
       club_code = club_name
 
    return club_code
"""
    if club_name in ('IL Koll', 'Idrettslaget Koll', 'Koll, IL', 'Koll, Idrettslaget'):
        club_code = 'KOLL'
    elif club_name in ('IL i BUL', 'Idrottslaget i BUL', 'BUL, IL i', 'BUL, Idrottslaget i'):
        club_code = 'ILBUL'
    elif club_name in ( 'IK Tjalve', 'Idrettsklubben Tjalve', 'Tjalve, IK', 'Tjalve, Idrettsklubben', 'Tjalve Idrettsklubben' ):
        club_code = 'TJAL'
    elif club_name in ( 'Tyrving IL', 'Tyrving Idrettslag' ):
        club_code = 'TYR'
    elif club_name in ( 'Romerike Friidrett' ):
        club_code = 'ROMFR'
    elif club_name in ( u'Bækkelagets SK' ): 
        club_code = 'BSK'
    elif club_name in ( 'Nesodden IF' ): 
        club_code = 'NESO'
    elif club_name in ( 'Groruddalen Friidrettsklubb' ): 
        club_code = 'GRO'
    elif club_name in ( 'Idrettslaget Sandvin', 'IL Sandvin' ): 
        club_code = 'SANDV'
    elif club_name in ( 'Eidanger Idrettslag' ): 
        club_code = 'EIDA'
    elif club_name in ( 'Ski IL Friidrett' ): 
        club_code = 'SKI'
    elif club_name in ( 'Gui Sportsklubb' ): 
        club_code = 'GUI'
    elif club_name in ( 'Sturla IF', 'Idrettsforeningen Sturla'): 
        club_code = 'STUR'
    elif club_name in ( 'Modum Friidrettsklubb'):
        club_code = 'MOD'
    elif club_name in ( u'Krødsherad Idrettslag', u'Krødsherad IL'):
        club_code = 'KRHER'
    elif club_name in ( u'Moss IL' ):
        club_code = 'MOSS'
    elif club_name in ( u'Norna-Salhus IL' ):
        club_code = 'NORSA'
    elif club_name in ( u'Sandnes IL' ):
        club_code = 'SAND'
    elif club_name in ( u'Askim IF' ):
        club_code = 'ASKIM'
    else:
        club_code = club_name

    return club_code
"""
def club_name(club_code):
    """
    if club_code == 'KOLL':
        club_name = 'IL Koll'
    elif club_code in ( 'ILBUL', 'ILIBUL'):
        club_name = 'IL i BUL'
    elif club_code ==  'TJAL':
        club_name = 'IK Tjalve'
    elif club_code ==  'TYR':
        club_name = 'Tyrving IL'
    elif club_code ==  'ROMFR':
        club_name = 'Romerike Friidrett'
    elif club_code ==  'BSK':
        club_name = u'Bækkelagets SK'
    elif club_code ==  'NESO':
        club_name = 'Nesodden IF'
    elif club_code ==  'GRO':
        club_name = 'Groruddalen FIK'
    elif club_code ==  'SANDV':
        club_name = 'IL Sandvin'
    elif club_code ==  'EIDA':
        club_name = 'Eidanger Idrettslag'
    elif club_code ==  'SKI':
        club_name = 'Ski IL Friidrett'
    elif club_code ==  'GUI':
        club_name = 'Gui Sportsklubb'
    elif club_code ==  'STUR':
        club_name = 'IF Sturla'
    elif club_code in ( 'MOD' ):
        club_name = 'Modum Friidrettsklubb'
    elif club_code in ( 'KRHER' ):
        club_name = u'Krødsherad Idrettslag'
    elif club_code in ( 'MOSS' ):
        club_name = 'Moss IL'
    elif club_code in ( 'NORSA'  ):
        club_name = 'Norna-Salhus IL'
    elif club_code in ( 'SAND'  ):
        club_name = 'Sandnes IL'
    elif club_code in ( 'ASKIM' ):
        club_name = 'Askim IF
    """
    if club_code == (u'AALEN'):
       club_name=u'ÅLEN IDRETTSLAG'
    elif club_code == (u'AASUN'):
       club_name=u'Ålesund Friidrettsklubb'
    elif club_code == (u'IKV'):
       club_name=u'Allianseidrettslaget Ik Våg'
    elif club_code == (u'ALM'):
       club_name=u'Almenning Il'
    elif club_code == (u'ALSV'):
       club_name=u'Alsvåg Idrettslag'
    elif club_code == (u'ALTA'):
       club_name=u'Alta Idrettsforening'
    elif club_code == (u'ANDAL'):
       club_name=u'Åndalsnes Idrettsforening'
    elif club_code == (u'ANDE'):
       club_name=u'Andebu Idrettslag'
    elif club_code == (u'ANDSK'):
       club_name=u'Andørja Sportsklubb'
    elif club_code == (u'ARE'):
       club_name=u'Aremark Idrettsforening'
    elif club_code == (u'ARNA'):
       club_name=u'Arna Turn & Idrettslag'
    elif club_code == (u'ASIL'):
       club_name=u'Ås Idrettslag'
    elif club_code == (u'AASEN'):
       club_name=u'Åsen Idrettslag'
    elif club_code == (u'AASER'):
       club_name=u'Åseral idrettslag'
    elif club_code == (u'ASK'):
       club_name=u'Ask Friidrett'
    elif club_code == (u'ASKFL'):
       club_name=u'Asker Fleridrettslag'
    elif club_code == (u'ASKSK'):
       club_name=u'Asker Skiklubb'
    elif club_code == (u'ASKIM'):
       club_name=u'Askim Idrettsforening'
    elif club_code == (u'ATNA'):
       club_name=u'Atna Idrettslag'
    elif club_code == (u'AURE'):
       club_name=u'Aure Idrettslag'
    elif club_code == (u'AURL'):
       club_name=u'Aurland Idrettslag'
    elif club_code == (u'AURS'):
       club_name=u'Aurskog-Høland Friidrettslag'
    elif club_code == (u'AUFJ'):
       club_name=u'Austefjord Idrettslag'
    elif club_code == (u'AUST'):
       club_name=u'Austevoll Idrettsklubb'
    elif club_code == (u'AUSTR'):
       club_name=u'Austrheim Idrettslag'
    elif club_code == (u'BAGN'):
       club_name=u'Bagn Idrettslag'
    elif club_code == (u'BAKKE'):
       club_name=u'Bakke IF'
    elif club_code == (u'BALE'):
       club_name=u'Balestrand IL'
    elif club_code == (u'BARD'):
       club_name=u'Bardu Idrettslag'
    elif club_code == (u'BIF'):
       club_name=u'Bøler Idrettslag'
    elif club_code == (u'BTSFJ'):
       club_name=u'Båtsfjord Sportsklubb'
    elif club_code == (u'BGND'):
       club_name=u'Begnadalen Idrettslag'
    elif club_code == (u'BEIT'):
       club_name=u'Beitstad Idrettslag'
    elif club_code == (u'BCK'):
       club_name=u'Bergen Cykleklubb'
    elif club_code == (u'BTC'):
       club_name=u'Bergen Triathlon Club'
    elif club_code == (u'BTU'):
       club_name=u'Bergens Turnforening'
    elif club_code == (u'BERGE'):
       club_name=u'Berger Idrettslag'
    elif club_code == (u'BFGL'):
       club_name=u'BFG Bergen Løpeklubb'
    elif club_code == (u'BJERR'):
       club_name=u'Bjerkreim Idrettslag'
    elif club_code == (u'BJERV'):
       club_name=u'Bjerkvik Idrettsforening'
    elif club_code == (u'BLA'):
       club_name=u'Blaker IL'
    elif club_code == (u'BLEFJ'):
       club_name=u'Blefjell Idrettslag'
    elif club_code == (u'BODO'):
       club_name=u'Bodø & Omegn IF Friidrett'
    elif club_code == (u'BODL'):
       club_name=u'Bodø Bauta Løpeklubb'
    elif club_code == (u'BODF'):
       club_name=u'Bodø Friidrettsklubb'
    elif club_code == (u'BOKN'):
       club_name=u'Bokn Idrettslag'
    elif club_code == (u'BOS'):
       club_name=u'Bossekop Ungdomslag'
    elif club_code == (u'BOTNA'):
       club_name=u'Botnan Idrettslag'
    elif club_code == (u'BOT'):
       club_name=u'Botne Skiklubb'
    elif club_code == (u'BRNDB'):
       club_name=u'Brandbu IF'
    elif club_code == (u'BRATS'):
       club_name=u'Bratsberg Idrettslag'
    elif club_code == (u'BRATTV'):
       club_name=u'Brattvåg Idrettslag'
    elif club_code == (u'BREIM'):
       club_name=u'Breimsbygda IL'
    elif club_code == (u'BREKK'):
       club_name=u'Brekke Idrettslag'
    elif club_code == (u'BREMA'):
       club_name=u'Bremanger Idrettslag'
    elif club_code == (u'BREM'):
       club_name=u'Bremnes Idrettslag'
    elif club_code == (u'BRE'):
       club_name=u'Brevik Idrettslag'
    elif club_code == (u'BROMM'):
       club_name=u'Bromma Idrettslag'
    elif club_code == (u'BRYF'):
       club_name=u'Bryne Friidrettsklubb'
    elif club_code == (u'BRYT'):
       club_name=u'BRYNE TRIATLONKLUBB'
    elif club_code == (u'BUD'):
       club_name=u'Bud Idrettslag'
    elif club_code == (u'BYS'):
       club_name=u'Byaasen Skiklub'
    elif club_code == (u'BYI'):
       club_name=u'Byåsen Idrettslag'
    elif club_code == (u'BYN'):
       club_name=u'Byneset IL Hovedlaget'
    elif club_code == (u'BSK'):
       club_name=u'Bækkelagets SK'
    elif club_code == (u'BRVHA'):
       club_name=u'Bærums Verk Hauger Idrettsforening'
    elif club_code == (u'BVRFJ'):
       club_name=u'Bæverfjord Idrettslag'
    elif club_code == (u'BIF'):
       club_name=u'Bøler Idrettsforening'
    elif club_code == (u'BMLO'):
       club_name=u'Bømlo Idrettslag'
    elif club_code == (u'BRSA'):
       club_name=u'Børsa Idrettslag'
    elif club_code == (u'DALE'):
       club_name=u'Dale Idrettslag'
    elif club_code == (u'DLN'):
       club_name=u'Dalen Idrettslag'
    elif club_code == (u'DIM'):
       club_name=u'Dimna IL'
    elif club_code == (u'DMB'):
       club_name=u'Dombås Idrettslag'
    elif club_code == (u'DRIV'):
       club_name=u'Driv Idrettslag'
    elif club_code == (u'DRIVA'):
       club_name=u'Driva IL'
    elif club_code == (u'DRFR'):
       club_name=u'Drøbak-Frogn Idrettslag'
    elif club_code == (u'DPVG'):
       club_name=u'Dypvåg Idrettsforening'
    elif club_code == (u'EGRSU'):
       club_name=u'Egersunds Idrettsklubb'
    elif club_code == (u'EIDIL'):
       club_name=u'Eid Idrettslag'
    elif club_code == (u'EIDA'):
       club_name=u'Eidanger Idrettslag'
    elif club_code == (u'EIDF'):
       club_name=u'Eidfjord Idrettslag'
    elif club_code == (u'EIDSB'):
       club_name=u'Eidsberg Idrettslag'
    elif club_code == (u'EIDS'):
       club_name=u'Eidsvåg Idrettslag'
    elif club_code == (u'EIDTU'):
       club_name=u'Eidsvold Turnforening Friidrett'
    elif club_code == (u'EIK'):
       club_name=u'Eikanger Idrettslag'
    elif club_code == (u'ESK'):
       club_name=u'Ekeberg Sports Klubb'
    elif club_code == (u'ESPA'):
       club_name=u'Espa Idrettslag'
    elif club_code == (u'ETNE'):
       club_name=u'Etne Idrettslag'
    elif club_code == (u'FAGIL'):
       club_name=u'Fagernes Idrettslag N'
    elif club_code == (u'FAG'):
       club_name=u'Fagernes Idrettslag O'
    elif club_code == (u'FALK'):
       club_name=u'Falkeid idrettslag'
    elif club_code == (u'FANA'):
       club_name=u'Fana IL'
    elif club_code == (u'FEIR'):
       club_name=u'Feiring Idrettslag'
    elif club_code == (u'FET'):
       club_name=u'Fet Friidrettsklubb'
    elif club_code == (u'FA77'):
       club_name=u'FIL AKS-77'
    elif club_code == (u'FINNY'):
       club_name=u'Finnøy Idrettslag'
    elif club_code == (u'FISIF'):
       club_name=u'Fiskå Idrettsforening'
    elif club_code == (u'FISIL'):
       club_name=u'Fiskå Idrettslag'
    elif club_code == (u'FITJ'):
       club_name=u'Fitjar Idrettslag'
    elif club_code == (u'FJVE'):
       club_name=u'Fjellhug/Vereide IL'
    elif club_code == (u'FLATS'):
       club_name=u'Flatås Idrettslag'
    elif club_code == (u'FLOR'):
       club_name=u'Florø T & IF'
    elif club_code == (u'FOLFO'):
       club_name=u'Follafoss Idrettslag'
    elif club_code == (u'FOL'):
       club_name=u'Folldal Idrettsforening'
    elif club_code == (u'FOLLO'):
       club_name=u'Follo Løpeklubb'
    elif club_code == (u'FORRA'):
       club_name=u'Forra Idrettslag'
    elif club_code == (u'FOSSU'):
       club_name=u'Fossum Idrettsforening'
    elif club_code == (u'FRED'):
       club_name=u'Fredrikstad Idrettsforening'
    elif club_code == (u'FREI'):
       club_name=u'Freidig Sportsklubben'
    elif club_code == (u'ORION'):
       club_name=u'Friidretsklubben Orion'
    elif club_code == (u'REN'):
       club_name=u'Friidrettsklubben Ren-Eng'
    elif club_code == (u'BAMSE'):
       club_name=u'Friidrettslaget Bamse'
    elif club_code == (u'BORG'):
       club_name=u'Friidrettslaget Borg'
    elif club_code == (u'FRISK'):
       club_name=u'Friidrettslaget Frisk'
    elif club_code == (u'FRO'):
       club_name=u'Frognerparken Idrettslag'
    elif club_code == (u'FROL'):
       club_name=u'Frol Idrettslag'
    elif club_code == (u'FROSTA'):
       club_name=u'Frosta Idrettslag'
    elif club_code == (u'FRLND'):
       club_name=u'Frøyland Idrettslag'
    elif club_code == (u'FUR'):
       club_name=u'Furuset Allidrett IF'
    elif club_code == (u'FYLL'):
       club_name=u'Fyllingen Idrettslag'
    elif club_code == (u'FRDE'):
       club_name=u'Førde Idrettslag'
    elif club_code == (u'GAULA'):
       club_name=u'Gaular IL'
    elif club_code == (u'GAU'):
       club_name=u'Gausdal Friidrettsklubb'
    elif club_code == (u'GEI'):
       club_name=u'Geilo Idrettslag'
    elif club_code == (u'GEIR'):
       club_name=u'Geiranger Idrettslag'
    elif club_code == (u'GJER'):
       club_name=u'Gjerpen Idrettsforening'
    elif club_code == (u'GJERS'):
       club_name=u'Gjerstad Idrettslag'
    elif club_code == (u'GJDAL'):
       club_name=u'Gjesdal Idrettslag'
    elif club_code == (u'GJFK'):
       club_name=u'Gjøvik Friidrettsklubb'
    elif club_code == (u'GJVIK'):
       club_name=u'Gjøvik Friidrettsklubb 2'
    elif club_code == (u'GLO'):
       club_name=u'Gloppen Friidrettslag'
    elif club_code == (u'GOL'):
       club_name=u'Gol Idrettslag'
    elif club_code == (u'GRON'):
       club_name=u'Grong Idrettslag'
    elif club_code == (u'GRO'):
       club_name=u'Groruddalen Friidrettsklubb'
    elif club_code == (u'GRUE'):
       club_name=u'Grue Idrettslag'
    elif club_code == (u'GTI'):
       club_name=u'GTI Friidrettsklubb'
    elif club_code == (u'GUI'):
       club_name=u'Gui Sportsklubb - Friidrett'
    elif club_code == (u'GUL'):
       club_name=u'Gulset Idrettsforening'
    elif club_code == (u'HAB'):
       club_name=u'HAB IL'
    elif club_code == (u'HADE'):
       club_name=u'Hadeland Friidrettsklubb'
    elif club_code == (u'HAGA'):
       club_name=u'Haga Idrettsforening '
    elif club_code == (u'HAL'):
       club_name=u'Halden Idrettslag'
    elif club_code == (u'HALMO'):
       club_name=u'Halmsås & Omegn Skilag'
    elif club_code == (u'HALSA'):
       club_name=u'Halsa Idrettslag'
    elif club_code == (u'HIL'):
       club_name=u'Hamar Idrettslag'
    elif club_code == (u'HANNEV'):
       club_name=u'Hannevikas Idrettslag'
    elif club_code == (u'HARDB'):
       club_name=u'Hardbagg Idrettslag'
    elif club_code == (u'HAREI'):
       club_name=u'Hareid Idrettslag'
    elif club_code == (u'HARE'):
       club_name=u'Harestua Idrettslag'
    elif club_code == (u'HATT'):
       club_name=u'Hattfjelldal Idrettslag'
    elif club_code == (u'HAUGN'):
       club_name=u'Haugen Idrettslag'
    elif club_code == (u'HAUGR'):
       club_name=u'Haugerud Idrettsforening'
    elif club_code == (u'HAUGF'):
       club_name=u'Haugesund Idrettslag Friidrett'
    elif club_code == (u'HAUGT'):
       club_name=u'Haugesund Triathlon Klubb'
    elif club_code == (u'HAV'):
       club_name=u'Havørn Allianseidrettslag'
    elif club_code == (u'HEGGF'):
       club_name=u'Heggedal Friidrettsklubb'
    elif club_code == (u'HEGGI'):
       club_name=u'Heggedal Idrettslag'
    elif club_code == (u'HELLU'):
       club_name=u'Hell Ultraløperklubb'
    elif club_code == (u'HEM'):
       club_name=u'Heming Idrettslaget'
    elif club_code == (u'HENN'):
       club_name=u'Henning I L'
    elif club_code == (u'HERA'):
       club_name=u'Herand Idrettslag'
    elif club_code == (u'HERK'):
       club_name=u'Herkules Friidrett'
    elif club_code == (u'HERY'):
       club_name=u'Herøy Idrettslag'
    elif club_code == (u'HIN'):
       club_name=u'Hinna Friidrett'
    elif club_code == (u'HITF'):
       club_name=u'Hitra Friidrettsklubb'
    elif club_code == (u'HITL'):
       club_name=u'Hitra Løpeklubb'
    elif club_code == (u'HOB'):
       club_name=u'Hobøl Idrettslag'
    elif club_code == (u'HOF'):
       club_name=u'Hof Idrettslag'
    elif club_code == (u'HOL'):
       club_name=u'Hol Idrettslag'
    elif club_code == (u'HOLMS'):
       club_name=u'Holmemstranda Idrettslag'
    elif club_code == (u'HOLUM'):
       club_name=u'Holum Idrettslag'
    elif club_code == (u'HMLV'):
       club_name=u'Hommelvik IL'
    elif club_code == (u'HOPE'):
       club_name=u'Hope Idrettslag'
    elif club_code == (u'HORNI'):
       club_name=u'Hornindal Idrettslag'
    elif club_code == (u'HORFR'):
       club_name=u'Horten Friidrettsklubb'
    elif club_code == (u'HUG'):
       club_name=u'Huglo Idrettslag'
    elif club_code == (u'HURD'):
       club_name=u'Hurdal Idrettslag'
    elif club_code == (u'HVAM'):
       club_name=u'Hvam Idrettslag'
    elif club_code == (u'HVFO'):
       club_name=u'Hvittingfoss Idrettslag'
    elif club_code == (u'HYEN'):
       club_name=u'Hyen Idrettslag'
    elif club_code == (u'HYLLS'):
       club_name=u'Hyllestad Idrettslag'
    elif club_code == (u'HSI'):
       club_name=u'Høybråten og Stovner IL'
    elif club_code == (u'HDMO'):
       club_name=u'Høydalsmo Idrottslag'
    elif club_code == (u'FJELLO'):
       club_name=u'I.l Fjellørnen'
    elif club_code == (u'FRAMS'):
       club_name=u'I.L. Framsteg'
    elif club_code == (u'NORSA'):
       club_name=u'Norna-Salhus IL'
    elif club_code == (u'NYBR'):
       club_name=u'I.L. Nybrott'
    elif club_code == (u'IDD'):
       club_name=u'Idd Sportsklubb'
    elif club_code == (u'BIRK'):
       club_name=u'Idrettsforeningen Birkebeineren'
    elif club_code == (u'FRAM'):
       club_name=u'Idrettsforeningen Fram'
    elif club_code == (u'HELLA'):
       club_name=u'IF Hellas'
    elif club_code == (u'NJAAL'):
       club_name=u'Idrettsforeningen Njaal'
    elif club_code == (u'STUR'):
       club_name=u'Sturla IF'
    elif club_code == (u'IFORN'):
       club_name=u'Idrettsforeningen Ørn'
    elif club_code == (u'BJARG'):
       club_name=u'Idrettslaget Bjarg'
    elif club_code == (u'ILBJ'):
       club_name=u'Idrettslaget Bjørn'
    elif club_code == (u'DLBR'):
       club_name=u'Idrettslaget Dalebrand'
    elif club_code == (u'DYREV'):
       club_name=u'Idrettslaget Dyre Vaa'
    elif club_code == (u'EXPR'):
       club_name=u'Idrettslaget Express'
    elif club_code == (u'FORSK'):
       club_name=u'Idrettslaget Forsøk'
    elif club_code == (u'FRI'):
       club_name=u'Idrettslaget Fri'
    elif club_code == (u'GNE'):
       club_name=u'Idrettslaget Gneist'
    elif club_code == (u'HOLE'):
       club_name=u'Idrettslaget Holeværingen'
    elif club_code == (u'BULT'):
       club_name=u'Idrettslaget I Bondeungdomslaget I Tromsø'
    elif club_code == (u'ILAR'):
       club_name=u'Idrettslaget Ilar'
    elif club_code == (u'IVRIG'):
       club_name=u'Idrettslaget Ivrig'
    elif club_code == (u'JARD'):
       club_name=u'Idrettslaget Jardar'
    elif club_code == (u'JUT'):
       club_name=u'Idrettslaget Jutul'
    elif club_code == (u'ILROS'):
       club_name=u'Idrettslaget Ros'
    elif club_code == (u'RUNAR'):
       club_name=u'IL Runar'
    elif club_code == (u'ILSAN'):
       club_name=u'Idrettslaget Sand'
    elif club_code == (u'SANDV'):
       club_name=u'IL Sandvin'
    elif club_code == (u'SKADE'):
       club_name=u'Idrettslaget Skade'
    elif club_code == (u'SKJA'):
       club_name=u'Idrettslaget Skjalg'
    elif club_code == (u'SYR'):
       club_name=u'Idrettslaget Syril'
    elif club_code == (u'TRY'):
       club_name=u'Idrettslaget Trysilgutten'
    elif club_code == (u'GULA'):
       club_name=u'IL Gular'
    elif club_code == (u'ILIBUL'):
       club_name=u'IDROTTSLAGET I BUL'
    elif club_code == (u'ILBUL'):
       club_name=u'IDROTTSLAGET I BUL'
    elif club_code == (u'JOT'):
       club_name=u'Idrottslaget Jotun'
    elif club_code == (u'IDUN'):
       club_name=u'Idun Idrettslag'
    elif club_code == (u'EIKKV'):
       club_name=u'If Eiker Kvikk'
    elif club_code == (u'KAVE'):
       club_name=u'IF Kamp/Vestheim'
    elif club_code == (u'KLYP'):
       club_name=u'If Klypetussen'
    elif club_code == (u'GRANE'):
       club_name=u'Ik Grane Arendal Friidrett'
    elif club_code == (u'HIND'):
       club_name=u'IK Hind'
    elif club_code == (u'IKORN'):
       club_name=u'Ikornnes Idrettslag'
    elif club_code == (u'AASG'):
       club_name=u'IL Aasguten'
    elif club_code == (u'ALVI'):
       club_name=u'IL Alvidra'
    elif club_code == (u'ILBEV'):
       club_name=u'IL Bever`n'
    elif club_code == (u'BRODD'):
       club_name=u'IL Brodd'
    elif club_code == (u'FLV'):
       club_name=u'IL Flåværingen'
    elif club_code == (u'GRY'):
       club_name=u'IL Gry'
    elif club_code == (u'ILNOR'):
       club_name=u'IL Norodd'
    elif club_code == (u'PIO'):
       club_name=u'IL Pioner Friidrett'
    elif club_code == (u'POL'):
       club_name=u'IL Polarstjernen'
    elif club_code == (u'SAMH'):
       club_name=u'IL Samhald'
    elif club_code == (u'SANT'):
       club_name=u'IL Santor'
    elif club_code == (u'STKAM'):
       club_name=u'IL Stålkameratene'
    elif club_code == (u'TRIUM'):
       club_name=u'IL Triumf'
    elif club_code == (u'VIND'):
       club_name=u'Il Vindbjart'
    elif club_code == (u'VING'):
       club_name=u'IL Vinger'
    elif club_code == (u'INDRY'):
       club_name=u'Inderøy Idrettslag'
    elif club_code == (u'INN'):
       club_name=u'Innstranda IL'
    elif club_code == (u'INSTA'):
       club_name=u'International School of Stavanger'
    elif club_code == (u'ISFJO'):
       club_name=u'Isfjorden Idrettslag'
    elif club_code == (u'JOND'):
       club_name=u'Jondalen Idrettslag'
    elif club_code == (u'JVTN'):
       club_name=u'Jægervatnet Idrettslag'
    elif club_code == (u'JIL'):
       club_name=u'Jøa Idrettslag'
    elif club_code == (u'JLSTE'):
       club_name=u'Jølster Idrettslag'
    elif club_code == (u'KAUP'):
       club_name=u'Kaupanger Idrettslag'
    elif club_code == (u'KFUM'):
       club_name=u'Kfum-kameratene Oslo'
    elif club_code == (u'KJ'):
       club_name=u'Kjelsås Idrettslag'
    elif club_code == (u'KLPP'):
       club_name=u'Klepp Idrettslag'
    elif club_code == (u'KLK'):
       club_name=u'Klæbu Løpeklubb'
    elif club_code == (u'KLIL'):
       club_name=u'Kløfta Idrettslag'
    elif club_code == (u'KLBK'):
       club_name=u'Kolbukameratene I L'
    elif club_code == (u'KOLL'):
       club_name=u'Koll Idrettslaget'
    elif club_code == (u'KLVIL'):
       club_name=u'Kolvereid Idrettslag'
    elif club_code == (u'KNGSB'):
       club_name=u'Kongsberg Idrettsforening'
    elif club_code == (u'KNGSV'):
       club_name=u'Kongsvinger IL Friidrett'
    elif club_code == (u'KONN'):
       club_name = u'Konnerud IL'
    elif club_code == (u'KOP'):
       club_name=u'Kopervik Idrettslag'
    elif club_code == (u'KORG'):
       club_name=u'Korgen Idrettslag'
    elif club_code == (u'KRAG'):
       club_name=u'Kragerø IF Friidrett'
    elif club_code == (u'KRAAK'):
       club_name=u'Kråkerøy Idrettslag'
    elif club_code == (u'KRSTD'):
       club_name=u'Kråkstad Idrettslag'
    elif club_code == (u'KRL'):
       club_name=u'Kristiansand Løpeklubb'
    elif club_code == (u'KIF'):
       club_name=u'Kristiansands IF'
    elif club_code == (u'KRHER'):
       club_name=u'Krødsherad Idrettslag'
    elif club_code == (u'KVINES'):
       club_name=u'Kvinesdal Idrettslag'
    elif club_code == (u'KVFJ'):
       club_name=u'Kvæfjord Idrettslag'
    elif club_code == (u'KYRK'):
       club_name=u'Kyrksæterøra Idrettslag Kil'
    elif club_code == (u'LAKS'):
       club_name=u'Laksevåg TIL'
    elif club_code == (u'LALM'):
       club_name=u'Lalm Idrettslag'
    elif club_code == (u'LAM'):
       club_name=u'Lambertseter IF'
    elif club_code == (u'LANGS'):
       club_name=u'Langesund Sykle- og triathlonklubb'
    elif club_code == (u'LNKEIL'):
       club_name=u'Lånke Idrettslag'
    elif club_code == (u'LRVK'):
       club_name=u'Larvik Turn & Idrettsforening'
    elif club_code == (u'LEINS'):
       club_name=u'Leinstrand Idrettslag'
    elif club_code == (u'LENA'):
       club_name=u'Lena Idrettsforening'
    elif club_code == (u'LIERN'):
       club_name=u'Lierne Idrettslag'
    elif club_code == (u'LIF'):
       club_name=u'Lillehammer IF'
    elif club_code == (u'LILLS'):
       club_name=u'Lillesand Idrettslag'
    elif club_code == (u'LISTA'):
       club_name=u'Lista Idrettslag'
    elif club_code == (u'LODD'):
       club_name=u'Loddefjord IL'
    elif club_code == (u'LFTR'):
       club_name=u'Lofoten Triatlonklubb'
    elif club_code == (u'LOM'):
       club_name=u'Lom Idrettslag'
    elif club_code == (u'LUND'):
       club_name=u'Lundamo Idrettslag'
    elif club_code == (u'LUNDH'):
       club_name=u'Lundehøgda IL'
    elif club_code == (u'LUST'):
       club_name=u'Luster Idrettslag'
    elif club_code == (u'LYE'):
       club_name=u'Lye Idrettslag'
    elif club_code == (u'LYN'):
       club_name=u'Lyn Ski'
    elif club_code == (u'LNGD'):
       club_name=u'Lyngdal Idrettslag'
    elif club_code == (u'LYKA'):
       club_name=u'Lyngen/ Karnes Il'
    elif club_code == (u'LYNGO'):
       club_name=u'Lyngstad og Omegn Idrettslag'
    elif club_code == (u'LRSKG'):
       club_name=u'Lørenskog Friidrettslag'
    elif club_code == (u'LFK'):
       club_name=u'Løten Friidrett'
    elif club_code == (u'LTN'):
       club_name=u'Løten Friidrett 2'
    elif club_code == (u'MALM'):
       club_name=u'Malm IL'
    elif club_code == (u'MLSEL'):
       club_name=u'Målselv Idrettslag'
    elif club_code == (u'MALV'):
       club_name=u'Malvik Idrettslag'
    elif club_code == (u'MAAL'):
       club_name=u'Måløy Idrettslag Hovedstyre'
    elif club_code == (u'MAHA'):
       club_name=u'Mandal & Halse I.l.'
    elif club_code == (u'MNDL'):
       club_name=u'Måndalen Idrettslag'
    elif club_code == (u'MABY'):
       club_name=u'Markabygda Idrettslag'
    elif club_code == (u'MARKA'):
       club_name=u'Markane IL'
    elif club_code == (u'MARNA'):
       club_name=u'Marnardal Idrettslag'
    elif club_code == (u'MEDKI'):
       club_name=u'Medkila Skilag'
    elif club_code == (u'MELD'):
       club_name=u'Meldal Idrettslag'
    elif club_code == (u'MELHU'):
       club_name=u'Melhus Idrettslag'
    elif club_code == (u'MDSND'):
       club_name=u'Midsund Idrettslag'
    elif club_code == (u'MJSD'):
       club_name=u'Mjøsdalen IL'
    elif club_code == (u'MOD'):
       club_name=u'Modum Friidrettsklubb'
    elif club_code == (u'MOELV'):
       club_name=u'Moelven IL'
    elif club_code == (u'MOI'):
       club_name=u'Moi Idrettslag'
    elif club_code == (u'MOLDE'):
       club_name=u'Molde og Omegn Idrettsforening'
    elif club_code == (u'OLYMP'):
       club_name=u'Molde Olymp'
    elif club_code == (u'MOITU'):
       club_name=u'Moltustranda Idrettslag'
    elif club_code == (u'MOSJ'):
       club_name=u'Mosjøen Friidrettsklubb'
    elif club_code == (u'MOSS'):
       club_name=u'Moss Idrettslag'
    elif club_code == (u'MOSV'):
       club_name=u'Mosvik Idrettslag'
    elif club_code == (u'MUIL'):
       club_name=u'MUIL - Mefjordvær Ungdoms- og Idrettslag'
    elif club_code == (u'NAML'):
       club_name=u'Namdal løpeklubb'
    elif club_code == (u'NAMDA'):
       club_name=u'Namdalseid Idrettslag'
    elif club_code == (u'NAMSE'):
       club_name=u'Namsen Fif'
    elif club_code == (u'NANN'):
       club_name=u'Nannestad Idrettslag'
    elif club_code == (u'NAR'):
       club_name=u'Narvik Idrettslag'
    elif club_code == (u'NESB'):
       club_name=u'Nesbyen Idrettslag'
    elif club_code == (u'NESO'):
       club_name=u'Nesodden IF'
    elif club_code == (u'NES'):
       club_name=u'Nesøya Idrettslag'
    elif club_code == (u'NID'):
       club_name=u'Nidelv Idrettslag'
    elif club_code == (u'NISS'):
       club_name=u'Nissedal Idrettslag'
    elif club_code == (u'NITT'):
       club_name=u'Nittedal Idrettslag'
    elif club_code == (u'NRDKJ'):
       club_name=u'Nordkjosbotn Idrettslag'
    elif club_code == (u'NEIDS'):
       club_name=u'Nordre Eidsvoll Idrettslag'
    elif club_code == (u'NFJEL'):
       club_name=u'Nordre Fjell Friidrett'
    elif club_code == (u'NLAND'):
       club_name=u'Nordre Land Idrettslag'
    elif club_code == (u'NTRY'):
       club_name=u'Nordre Trysil IL'
    elif club_code == (u'NORIL'):
       club_name=u'Nordøy Idrettslag'
    elif club_code == (u'NORR'):
       club_name=u'Norrøna IL'
    elif club_code == (u'NRUN'):
       club_name=u'Northern Runners'
    elif club_code == (u'NTNUI'):
       club_name=u'NTNUI - Norges Teknisk-Naturvitenskapelige Universitets Idrettsforening'
    elif club_code == (u'NYSK'):
       club_name=u'Nydalens Skiklub'
    elif club_code == (u'NYKIR'):
       club_name=u'Nykirke Idrettsforening'
    elif club_code == (u'NTTRY'):
       club_name=u'Nøtterøy Idrettsforening'
    elif club_code == (u'ODDA'):
       club_name=u'Odda Idrettslag'
    elif club_code == (u'OGND'):
       club_name=u'Ogndal Idrettslag Hovedlaget'
    elif club_code == (u'OLD'):
       club_name=u'Olden Idrettslag'
    elif club_code == (u'OLDA'):
       club_name=u'Olderdalen Idrettsklubb'
    elif club_code == (u'OPPD'):
       club_name=u'Oppdal IL Hovedlaget'
    elif club_code == (u'OPP'):
       club_name=u'Oppegård Idrettslag'
    elif club_code == (u'OPSL'):
       club_name=u'Oppsal Idrettsforening'
    elif club_code == (u'OPST'):
       club_name=u'Oppstad Idrettslag'
    elif club_code == (u'OPPST'):
       club_name=u'Oppstad Idrettslag 2'
    elif club_code == (u'OPSTR'):
       club_name=u'Oppstryn Idrettslag'
    elif club_code == (u'OPPT'):
       club_name=u'Opptur Motbakkeklubb'
    elif club_code == (u'ORKA'):
       club_name=u'Orkanger Idrettsforening'
    elif club_code == (u'ORKD'):
       club_name=u'Orkdal Idrettslag'
    elif club_code == (u'ORRE'):
       club_name=u'Orre Idrettslag'
    elif club_code == (u'OS'):
       club_name=u'Os Idrettslag'
    elif club_code == (u'OSTU'):
       club_name=u'Os Turnforening'
    elif club_code == (u'FRII'):
       club_name=u'OSI Friidrett'
    elif club_code == (u'POLIT'):
       club_name=u'Oslo Politis Idrettslag'
    elif club_code == (u'OSI'):
       club_name=u'Oslostudentenes Idrettsklubb'
    elif club_code == (u'OST'):
       club_name=u'Osterøy Idrottslag'
    elif club_code == (u'OTRA'):
       club_name=u'Otra IL'
    elif club_code == (u'OTTE'):
       club_name=u'Ottestad Idrettslag'
    elif club_code == (u'OTKS'):
       club_name=u'Ottestad Kast og Styrkeløft'
    elif club_code == (u'OVRH'):
       club_name=u'Overhalla Idrettslag'
    elif club_code == (u'PORS'):
       club_name=u'Porsanger Idrettslag'
    elif club_code == (u'RANA'):
       club_name=u'Rana Friidrettsklubb'
    elif club_code == (u'RAN'):
       club_name=u'Ranheim IL'
    elif club_code == (u'RAU'):
       club_name=u'Raufoss IL Friidrett'
    elif club_code == (u'RAUM'):
       club_name=u'Raumnes & Årnes Idrettslag'
    elif club_code == (u'RE'):
       club_name=u'Re Friidrettsklubb'
    elif club_code == (u'READY'):
       club_name=u'Ready Idrettsforeningen'
    elif club_code == (u'RENA'):
       club_name=u'Rena Idrettslag'
    elif club_code == (u'RENDA'):
       club_name=u'Rendalen Idrettslag'
    elif club_code == (u'RENB'):
       club_name=u'Rennebu Idrettslag'
    elif club_code == (u'RIND'):
       club_name=u'Rindal Idrettslag'
    elif club_code == (u'RING'):
       club_name=u'Ringerike Friidrettsklubb'
    elif club_code == (u'RIS'):
       club_name=u'Risør Idrettslag'
    elif club_code == (u'RJU'):
       club_name=u'Rjukan Idrettslag'
    elif club_code == (u'ROGNE'):
       club_name=u'Rogne Idrettslag'
    elif club_code == (u'ROMFR'):
       club_name=u'Romerike Friidrett'
    elif club_code == (u'ROMUL'):
       club_name=u'Romerike Ultraløperklubb'
    elif club_code == (u'ROMRA'):
       club_name=u'Romsdal Randoneklubb'
    elif club_code == (u'ROSEN'):
       club_name=u'Rosendal Turnlag'
    elif club_code == (u'ROYAL'):
       club_name=u'Royal Sport'
    elif club_code == (u'RUS'):
       club_name=u'Rustad Idrettslag'
    elif club_code == (u'RYGGE'):
       club_name=u'Rygge Idrettslag'
    elif club_code == (u'RIL'):
       club_name=u'Røa Allianseidrettslag'
    elif club_code == (u'RDIL'):
       club_name=u'Røldal Idrettslag'
    elif club_code == (u'ROSIL'):
       club_name=u'Røros Idrettslag'
    elif club_code == (u'RKEN'):
       club_name=u'Røyken UIL'
    elif club_code == (u'SALA'):
       club_name=u'Salangen IF Friidrett'
    elif club_code == (u'SAMN'):
       club_name=u'Samnanger Idrettslag'
    elif club_code == (u'SANTU'):
       club_name=u'Sandane Turn og Idrettslag'
    elif club_code == (u'STIF'):
       club_name=u'SANDEFJORD TURN & IDRETTSFORENING'
    elif club_code == (u'SAND'):
       club_name=u'Sandnes IL'
    elif club_code == (u'SNDI'):
       club_name=u'Sandnes Idrettslag 2'
    elif club_code == (u'SNDSJ'):
       club_name=u'Sandnessjøen Idrettslag'
    elif club_code == (u'SARP'):
       club_name=u'Sarpsborg IL'
    elif club_code == (u'SAUD'):
       club_name=u'Sauda Idrettslag'
    elif club_code == (u'SAUL'):
       club_name=u'Sauland Idrettslag'
    elif club_code == (u'SELB'):
       club_name=u'Selbu IL'
    elif club_code == (u'SELJE'):
       club_name=u'Selje Idrettslag'
    elif club_code == (u'SELJO'):
       club_name=u'Seljord Idrettslag'
    elif club_code == (u'SELS'):
       club_name=u'Selsbakk Idrettsforening'
    elif club_code == (u'SEM'):
       club_name=u'Sem Idrettsforening'
    elif club_code == (u'SIGFR'):
       club_name=u'Sigdal Friidrettsklubb'
    elif club_code == (u'SIGSK'):
       club_name=u'Sigdals Skiklub'
    elif club_code == (u'SILJ'):
       club_name=u'Siljan Idrettslag'
    elif club_code == (u'SIRMA'):
       club_name=u'Sirma Il'
    elif club_code == (u'SJET'):
       club_name=u'Sjetne Idrettslag'
    elif club_code == (u'VEDA'):
       club_name=u'Sk Vedavåg Karmøy'
    elif club_code == (u'VID'):
       club_name=u'SK Vidar'
    elif club_code == (u'SKAGE'):
       club_name=u'Skagerrak Sportsklubb'
    elif club_code == (u'SKLA'):
       club_name=u'Skåla Idrettslag'
    elif club_code == (u'SKRPH'):
       club_name=u'Skarphedin IL'
    elif club_code == (u'SKAU'):
       club_name=u'Skaubygda Il'
    elif club_code == (u'SKAUN'):
       club_name=u'Skaun Idrettslag'
    elif club_code == (u'SKI'):
       club_name=u'Ski IL Friidrett'
    elif club_code == (u'SKJK'):
       club_name=u'Skjåk IL'
    elif club_code == (u'SKJO'):
       club_name=u'Skjoldar Il'
    elif club_code == (u'SKOGN'):
       club_name=u'Skogn Idrettslag'
    elif club_code == (u'SKO'):
       club_name=u'Skotterud Idrettslag'
    elif club_code == (u'SKREIA'):
       club_name=u'Skreia Idrettslag'
    elif club_code == (u'SNSA'):
       club_name=u'Snåsa Idrettslag'
    elif club_code == (u'SNGG'):
       club_name=u'Snøgg Friidrett'
    elif club_code == (u'SIL'):
       club_name=u'Sogndal Idrettslag'
    elif club_code == (u'SOKND'):
       club_name=u'Sokndal Friidrettsklubb'
    elif club_code == (u'SOLA'):
       club_name=u'Sola Friidrett'
    elif club_code == (u'SOLN'):
       club_name=u'Solnut IL'
    elif club_code == (u'SORTL'):
       club_name=u'Sortland Friidrettsklubb'
    elif club_code == (u'SOT'):
       club_name=u'Sotra Sportsklubb'
    elif club_code == (u'SPILL'):
       club_name=u'Spillum Idrettslag'
    elif club_code == (u'SPRD'):
       club_name=u'Spiridon Langløperlag'
    elif club_code == (u'SPJVK'):
       club_name=u'Spjelkavik og Omegn Friidrettsklubb'
    elif club_code == (u'KRAFT'):
       club_name=u'Sportsklubben Kraft'
    elif club_code == (u'nan'):
       club_name=u'Sportsklubben Rye'
    elif club_code == (u'RYE'):
       club_name=u'Sportsklubben Rye 2'
    elif club_code == (u'VIDAR'):
       club_name=u'Sportsklubben Vidar'
    elif club_code == (u'SPYDE'):
       club_name=u'Spydeberg IL'
    elif club_code == (u'STJIL'):
       club_name=u'Staal Jørpeland IL'
    elif club_code == (u'STAD'):
       club_name=u'Stadsbygd IL'
    elif club_code == (u'STRHE'):
       club_name=u'Stårheim IL'
    elif club_code == (u'STDIF'):
       club_name=u'Stavanger Døve-Idrettsforening'
    elif club_code == (u'STAVA'):
       club_name=u'Stavanger Friidrettsklubb'
    elif club_code == (u'STAVIF'):
       club_name=u'Stavanger Idrettsforening Allianseidrettslag - Friidrett'
    elif club_code == (u'STEGA'):
       club_name=u'Stegaberg Idrettslag'
    elif club_code == (u'STEIN'):
       club_name=u'Stein Friidrettsklubb'
    elif club_code == (u'STEKJ'):
       club_name=u'Steinkjer Friidrettsklubb'
    elif club_code == (u'STSK'):
       club_name=u'Stettevik Sportsklubb'
    elif club_code == (u'STJF'):
       club_name=u'Stjørdal Fridrettsklubb'
    elif club_code == (u'STJP'):
       club_name=u'Stjørdal Paraidrettslag'
    elif club_code == (u'STJB'):
       club_name=u'Stjørdals-Blink IL'
    elif club_code == (u'STOKK'):
       club_name=u'Stokke Idrettslag'
    elif club_code == (u'STOKM'):
       club_name=u'Stokmarknes Idrettslag'
    elif club_code == (u'STO'):
       club_name=u'Stord Idrettslag'
    elif club_code == (u'STOFJ'):
       club_name=u'Storfjord idrettslag'
    elif club_code == (u'STRIL'):
       club_name=u'Stranda Idrottslag'
    elif club_code == (u'STRAB'):
       club_name=u'Strandebarm Idrettslag'
    elif club_code == (u'STRA'):
       club_name=u'Stranden Idrettslag'
    elif club_code == (u'STRAU'):
       club_name=u'Straumsnes Idrettslag'
    elif club_code == (u'STRI'):
       club_name=u'Strindheim Idrettslag'
    elif club_code == (u'STRY'):
       club_name=u'Stryn Turn og Idrettslag'
    elif club_code == (u'STREN'):
       club_name=u'Støren Sportsklubb'
    elif club_code == (u'SUNND'):
       club_name=u'Sunndal IL Friidrett'
    elif club_code == (u'SURN'):
       club_name=u'Surnadal Idrettslag'
    elif club_code == (u'SVTU'):
       club_name=u'Svalbard Turn Idrettslag'
    elif club_code == (u'SVARS'):
       club_name=u'Svarstad Idrettslag'
    elif club_code == (u'SVEIO'):
       club_name=u'Sveio Idrettslag'
    elif club_code == (u'SVEL'):
       club_name=u'Svelgen Turn og Idrettsforening'
    elif club_code == (u'SVINT'):
       club_name=u'Svint IL'
    elif club_code == (u'SVORK'):
       club_name=u'SVORKMO N.O.I.'
    elif club_code == (u'SYKK'):
       club_name=u'Sykkylven Idrottslag'
    elif club_code == (u'SYLL'):
       club_name=u'Sylling Idrettsforening'
    elif club_code == (u'SDAL'):
       club_name=u'Sædalen Idrettslag'
    elif club_code == (u'GRAA'):
       club_name=u'Sætre Idrætsforening Graabein'
    elif club_code == (u'STIL'):
       club_name=u'Søfteland Turn & Idrettslag'
    elif club_code == (u'SGNE'):
       club_name=u'Søgne Idrettslag'
    elif club_code == (u'SMNA'):
       club_name=u'Sømna Idrettslag'
    elif club_code == (u'SNDLA'):
       club_name=u'Søndre Land IL Friidrett'
    elif club_code == (u'SAAL'):
       club_name=u'Søre Ål Idrettslag'
    elif club_code == (u'SRILD'):
       club_name=u'Sørild Fridrettsklubb'
    elif club_code == (u'SRKDL'):
       club_name=u'Sørkedalens Idrettsforening'
    elif club_code == (u'HOVD'):
       club_name=u'T I L Hovding'
    elif club_code == (u'TAMBA'):
       club_name=u'IL Tambarskjelvar'
    elif club_code == (u'TAMSAN'):
       club_name=u'Tamil Sangam IL'
    elif club_code == (u'TING'):
       club_name=u'Tingvoll Friidrettsklubb'
    elif club_code == (u'TIST'):
        club_name = u'Tistedalen FL'
    elif club_code == (u'TJAL'):
       club_name=u'IK Tjalve'
    elif club_code == (u'TJØLL'):
       club_name=u'Tjølling Idrettsforening'
    elif club_code == (u'TJI'):
       club_name=u'Tjøme Idrettslag'
    elif club_code == (u'TJL'):
       club_name=u'Tjøme Løpeklubb'
    elif club_code == (u'TOL'):
       club_name=u'Tolga Idrettslag'
    elif club_code == (u'TOMR'):
       club_name=u'Tomrefjord Idrettslag'
    elif club_code == (u'TORO'):
       club_name=u'Torodd IF'
    elif club_code == (u'TORVI'):
       club_name=u'Torvikbukt Idrettslag'
    elif club_code == (u'TREU'):
       club_name=u'Treungen Idrettslag'
    elif club_code == (u'TRIO'):
       club_name=u'Trio idrettslag'
    elif club_code == (u'TRF'):
       club_name=u'Tromsø Friidrettsklubb'
    elif club_code == (u'TRL'):
       club_name=u'Tromsø Løpeklubb'
    elif club_code == (u'TRS'):
       club_name=u'Tromsø Svømmeklubb'
    elif club_code == (u'TROO'):
       club_name=u'Trondheim & Omegn Sportsklubb'
    elif club_code == (u'TROF'):
       club_name=u'Trondheim Friidrett'
    elif club_code == (u'TSK'):
       club_name=u'Trøgstad Skiklubb'
    elif club_code == (u'TUIL'):
       club_name=u'TUIL Tromsdalen Friidrett'
    elif club_code == (u'TVEDE'):
       club_name=u'Tvedestrand Turn & Idrettsforening'
    elif club_code == (u'TYR'):
       club_name=u'Tyrving IL'
    elif club_code == (u'TNSBF'):
       club_name=u'Tønsberg Friidrettsklubb'
    elif club_code == (u'TRBIL'):
       club_name=u'Tørvikbygd Idrettslag'
    elif club_code == (u'TYEN'):
       club_name=u'Tøyen Sportsklubb'
    elif club_code == (u'ULLK'):
       club_name=u'Ullensaker/Kisa IL Friidrett'
    elif club_code == (u'ULKI'):
       club_name=u'Ullensaker/Kisa IL Friidrett 2'
    elif club_code == (u'UND'):
       club_name=u'Undheim Idrettslag'
    elif club_code == (u'URFRI'):
       club_name=u'Urædd Friidrett'
    elif club_code == (u'UTL'):
       club_name=u'Utleira Idrettslag'
    elif club_code == (u'VAAL'):
       club_name=u'Vaaler Idrettsforening'
    elif club_code == (u'VA'):
       club_name=u'Vadsø Atletklubb'
    elif club_code == (u'VTF'):
       club_name=u'Vadsø Turnforening (Vtf)'
    elif club_code == (u'VGAA'):
       club_name=u'Vågå Idrettslag'
    elif club_code == (u'VIL'):
       club_name=u'Vågstranda Idrettslag'
    elif club_code == (u'VALK'):
       club_name=u'Valkyrien Idrettslag'
    elif club_code == (u'VALL'):
       club_name=u'Valldal Idrettslag'
    elif club_code == (u'VAL'):
       club_name=u'Vallset IL'
    elif club_code == (u'VAR'):
       club_name=u'Varegg Fleridrett'
    elif club_code == (u'VARH'):
       club_name=u'Varhaug Idrettslag'
    elif club_code == (u'VART'):
       club_name=u'Varteig Idrettslag'
    elif club_code == (u'VEG'):
       club_name=u'Vegårshei Idrettslag'
    elif club_code == (u'VELD'):
       club_name=u'Veldre Friidrett'
    elif club_code == (u'VELL'):
       club_name=u'Velledalen Idrettslag'
    elif club_code == (u'VERD'):
       club_name=u'Verdal Friidrettsklubb'
    elif club_code == (u'VESTB'):
       club_name=u'Vestby Idrettslag'
    elif club_code == (u'VESTB'):
       club_name=u'Vestby Idrettslag'
    elif club_code == (u'VESTF'):
       club_name=u'Vestfossen Idrettsforening'
    elif club_code == (u'VSPON'):
       club_name=u'Vestre Spone IF'
    elif club_code == (u'VIKIL'):
       club_name=u'Vik Idrettslag'
    elif club_code == (u'VIKAN'):
       club_name=u'Vikane IL'
    elif club_code == (u'VIK'):
       club_name=u'TIF Viking'
    elif club_code == (u'VIKSD'):
       club_name=u'Viksdalen Idrettslag'
    elif club_code == (u'VILJ'):
       club_name=u'Viljar IL'
    elif club_code == (u'VNDIL'):
       club_name=u'Vind Idrettslag'
    elif club_code == (u'VINDA'):
       club_name=u'Vindafjord Idrettslag'
    elif club_code == (u'VINJE'):
       club_name=u'Vinje Idrottslag'
    elif club_code == (u'VOLL'):
       club_name=u'Vollan Idrettsklubb'
    elif club_code == (u'VOSS'):
       club_name=u'Voss Idrottslag'
    elif club_code == (u'YTTER'):
       club_name=u'Ytterøy Idrettslag'
    elif club_code == (u'ORJIL'):
       club_name=u'Ørje Idrettslag'
    elif club_code == (u'ORSTA'):
       club_name=u'Ørsta Idrettslag'
    elif club_code == (u'OMARSJ'):
       club_name=u'Østmarka Marsjklubb'
    elif club_code == (u'OTRET'):
       club_name=u'Øyer/Tretten Idrettsforening'
    elif club_code == (u'OSLID'):
       club_name=u'Øystre Slidre Idrettslag'
    else:
       club_name=club_code

    return club_name
   
def get_stats(event,cat,season):
    event_id = {'100':'4', '200': '5', '400':'7', '800':'9', '1500':'11', '3000':'13', '5000':'14', '10000':'15',
            '100H':'35', '110H':'42', '400H':'59', '2000SC':'65', '3000SC' : '121'}
    catcodes = {'KS': '22', 'MS': '11'}

    #print(cat, event)
    #event = event_code(event)
    #print(type(event),event)
    if event in event_id.keys():
       url = 'https://www.minfriidrettsstatistikk.info/php/LandsStatistikk.php?showclass=' + cat + '&showevent=' + event_id[event] + '&outdoor=Y&showseason=' + season + '&showclub=0'
       #print(url)
       r = requests.get(url)

    
       soup = BeautifulSoup(r.text, 'html.parser')
       tables = [
           [
               [td.get_text(strip=True) for td in tr.find_all('td')]
               for tr in table.find_all('tr')
           ]
           for table in soup.find_all('table')
       ] 
   
       stats = []
       for table in tables:
       #for row in tables[0]:
           for row in table:
               if not row == [] and not row[0] == '-----':
                  #print(row)
                  #name, club  =  row[1].split(',')
                  name, club  =  row[1].rsplit(',', 1)
                  dob = row[2]
                  #dob = datetime.date.strptime(row[2],'%d.%m.%y')
                  perf = row[0]
                  stats.append( (name, dob, perf) )
       #print('x',stats)
       return stats

def get_seed_marks(name, dob, event, cat, season): 
    event_id = {'100':'4', '200': 5, '400':'7', 800:'9', '1500':'11', '3000':'13', '5000':'14', '10000':'15',
            '100H':'35', '110H':'42', '400H':'59', '3000SC' : '120' ,
            'HJ':'68', 'PV':'70', 'LJ':'71', 'TJ':'75', } 
    catcodes = {'KS': '22', 'MS': '11',
            'G15':'6', 'G16':'7', 'G17':'8', 'G18/19':'9',
            'J15':'17', 'J16':'18', 'J17':'19', 'J18/19':'20'}

    if cat not in ('MS', 'KS'):
        return ''

    event = event_code(event)
    cat    = catcodes[cat]

    global event_stats
    event_stats = {}

    #print(event)
    if event not in event_stats.keys():
        event_stats[event] = {}
    if cat not in event_stats[event]:
        event_stats[event][cat] = {}
    if season not in event_stats[event][cat]:
        event_stats[event][cat][season] = get_stats(event,cat,season)

    res = 'nm'
    print(event, cat, season)
    #print('a',event_stats)
    s = event_stats.get(event, None)
    #print(s)
    if not s==None:
        #for p in event_stats[event][cat][season]:
        #print(cat,season)
        #print(type(s[cat][season]))
        #print(s[cat][season])
        #print(s)
        if not s[cat][season]==None:
           for p in s[cat][season]:
               #print(p)
               nme = p[0]
               #dd = datetime.datetime.strptime(p[1], '%d.%m.%y')
               #res = p[2]

               ratio = fuzz.token_set_ratio(name, nme)
               #print(name,nme,pratio)
               if ratio>85:
                   if 'Magnus' in name:
                       print(name, nme, ratio, p[2])
                   res = p[2]
                   break

        minsecpat = '(\d?\d)[:.,](\d\d[,.]\d?\d)'
        match1 = re.match(minsecpat,res)
        reswindpat = "(\d?\d[,.]\d\d)[(]([+-]\d[,.]\d)[)]"
        match2 = re.match(reswindpat,res)
        if match1:
            mins = match1.group(1)
            secs = match1.group(2).replace(',','.')
            res = mins + ':' + secs
        elif match2:
            secs = match2.group(1)
            wind = match2.group(2)
            #print('wind=', wind)
            res = secs.replace(',','.')
        else:
            res = res.replace(',','.')
        #print(name,res)
        return res






# ...
if len(sys.argv) < 2:
   sys.exit("Usage: %s <infile>" % sys.argv[0])
   
infile = sys.argv[1]
print(infile)
tree = read_xml_into_tree(infile)
save_xml_copy(tree)


"""
events_crosstable = make_crosstable(tree)
for e1 in sorted( events_crosstable.keys() ):
    for e2 in sorted( events_crosstable[e1].keys() ):
        print e1 +'|'+ e2, events_crosstable[e1][e2]
"""

#write_xlsx_results_template(tree)
#write_start_lists_as_html(tree)


write_opentrack_import(tree)

"""
event = 'Lengde satssone'
classes = [ 'F6', 'F7', 'G8', 'J8', 'J9' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G10', 'G11', 'J10', 'J11' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G 12', 'G13', 'J12', 'J13' ]
make_horizontal_protocol(tree, event, classes)
event = 'Lengde'
classes = [ 'G14', 'J14', 'J15']
make_horizontal_protocol(tree, event, classes)

event = 'Kule'
classes = [ 'G10', 'G11', 'J10', 'J11' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G14', 'G17', 'J14', 'J15' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'J12', 'J13']
make_horizontal_protocol(tree, event, classes)


event = u'Høyde'
classes = [ 'J13', 'G14', 'J15']
make_vertical_protocol(tree, event, classes)
classes = [ 'J10', 'J11', 'J12']
make_vertical_protocol(tree, event, classes)
classes = [ 'G10', 'G11', 'G12']
make_vertical_protocol(tree, event, classes)
event = 'Lengde'
classes = [ 'J14' ]
make_horizontal_protocol(tree, event, classes)
event = 'Kule'
classes = [ 'J11', 'J12', 'J13', 'G11' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G13', 'J14' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G14' ]
make_horizontal_protocol(tree, event, classes)
event = 'Liten ball'
classes = [ 'J 9', 'J10', 'G 9', 'G10' ]
make_horizontal_protocol(tree, event, classes)
"""
