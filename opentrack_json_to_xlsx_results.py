# -*- coding: utf-8 -*-

# TODO: 
#       + clean up/more modular
#       + PARA categories
#       + sorting order
#       + ties
#       + foreign and non-default teams
#       + best valid attempt (wind)
#       + rounds on different days
#       + series in multi round field events
#
import sys
import json
import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import colors as xlcolors
from openpyxl.styles import Font, Color
import requests
import random
from collections import defaultdict
import argparse

from stevneinfo import clubs, events

import pprint

noplace = int(1.e10)-1
def get_category(birthdate, eventdate, gender):
    birthyear = birthdate.year
    eventyear = eventdate.year
    age = int(eventyear)-int(birthyear)

    g = {'F' : 'J', 'M' : 'G' }
    if age > 19:
        g = {'F' : 'K', 'M' : 'M' }
        if age < 35:
           a = 'S'
        else:
           a = 'V' + '%d'%(5*int(age/5))
    elif age in (18,19):
        a = '18/19'
    else:
        a = '%d'%(age)

    cat = g[gender]+a
    return cat

#def sort_results_by():
    #

def get_organiser_name(key):
    organisers = { "1376e260-82f7-4bf6-9da6-064fd76c6d87" : "IL Koll", 
                   "575153d6-7f1b-4795-9276-5f8d57414944" : 'IK Tjalve'
            }
    return organisers.get(key, key)

def is_relay(event_code):
    # from athlib/codes.py
    PAT_RELAYS = r"^(?:(\d{1,2})[xX](\d{2,5}[hH]?|[rR][eE][lL][aA][yY]|[dDsS][mM][rR]))$"  # 4x100, 4x400, 4xReLAy, 4xDMR, 4xSMR, 12x200H
    match = re.search(PAT_RELAYS, event_code)
    return match is not None


#---------------------------------------
#if len(sys.argv) < 2:
#   sys.exit("Usage: %s <url>" % sys.argv[0])
parser = argparse.ArgumentParser()
parser.add_argument('url')
parser.add_argument('--sort_by', default='age')
args = parser.parse_args()
   
#url = sys.argv[1]
#print(args)
url = args.url
print(url)
if args.sort_by in ('age', 'cat'):
    sort_by = args.sort_by
else:
    sys.exit("Wrong value for 'sort_by', must be in ('age', 'cat')")


r=requests.get(url+'json')
j = json.loads(r.text)
#with open('downloads.json', 'r') as f: 
#    j = json.load(f)

#print(type(j))
#print(j.keys())
#print(j['date'])
d  = j['date']
d2 = j['finishDate']
isodateformat = "%Y-%m-%d"
date0 = datetime.datetime.strptime(d, isodateformat)
date1 = datetime.datetime.strptime(d2, isodateformat)
#print(d, date0, date1)
#bdate = datetime.datetime.strptime('2005-06-24', isodateformat)
#print(get_category(bdate,date,'F'))
dates = []
d = date0
while d <= date1:
    dates.append(d)
    d += datetime.timedelta(days=1)
#print(dates)


if 'nameLocal' in j.keys():
    meetname = j['nameLocal']
meetname = j['fullName']

slug = j['slug']
outdoors = 'J'
if j.get('venue') == None: 
    venue = ''
else:
    venue = j['venue']['formalName']
    if j['venue']['indoor'] == 'true':
        outdoors = 'N'
if j['type']=="INDOOR":
    outdoors='N'


#print(meetname, venue)
#organiser_name =  j['organiser']['name']
#organiser_name = get_organiser_name(organiser_key)


ignore_bibs = []
competitors = {}
#print(j['competitors'][0])
for c in j['competitors']:
#   print(c.keys())
    fn = ''; ln = ''; dob= ''; t=''
    bib  = c['competitorId']
    if 'firstName' in c.keys():
        fn   = c['firstName']
    if 'lastName' in c.keys():
        ln   = c['lastName']
    if 'dateOfBirth' in c.keys():
        d    = c['dateOfBirth']
        dob  = datetime.datetime.strptime(d, isodateformat)
    if 'teamId' in c.keys():
        t    = c['teamId']
#   if 'teamName' in c.keys():
#       t    = c['teamName']
#       print (t)

    if 'gender' in c.keys():
        g = c['gender']
    if fn=='':
        ignore_bibs.append(bib)
    else:
        competitors[bib] = [fn, ln, dob, g, t]
#       print(bib, competitors[bib])
relay_teams = {}
for t in j['relayTeams']:
#   print( t.keys() )
    bib = t['bib']
    teamname = t['name']
    if 'runnerNames' in t.keys():
        runners = t['runnerNames']
    else:
        runners = t['runners']
    relay_teams[bib] = (teamname, runners)

#print(relay_teams)


#print(type(j['events']))
#print(type(j['events'][0]))
#print(j['events'][0].keys())
#print(j['events'][0]['units'][0])
outdoor = 'J'
if j['type'] == "INDOOR":
    outdoor = 'N'
"""
results = {}
for e in j['events']:
    eventcode = e['eventCode']
    if eventcode not in results:
        results[eventcode] = {}
    #print(eventcode)
    for u in e['units']:
        #print(u.keys())
        for r in u['results']:
            bib = r['bib']
            if bib not in ignore_bibs:
                #print(r)
                d = competitors[bib][2]
                g = competitors[bib][3]
                cat = get_category(d,date,g)
                #print(d, date.strftime('%d.%m.%Y'), cat)
                if cat not in results[eventcode].keys():
                    results[eventcode][cat] = []
                # might want to add a dict of performance attributes to this tuple
                # e.g. jump series, sort keys ...
                rfiesults[eventcode][cat].append( (bib, r['performance']) )    
 
    print( bib, (fn, ln, dob.strftime('%d.%m.%Y'), t) )
    competitors[bib] = (fn, ln, dob, t)
#print(competitors)
"""
#print(competitors)

multis = ( 'BI', 'TRI', 'QUAD', 'PEN', 'HEX', 'HEP', 'OCT', 'ENN', 'DEC', 'HEN', 'DOD', 'ICO')

poolnr = 0
results ={}
series = {}
for e in j["events"]:
#   pp = pprint.PrettyPrinter(indent=4)
#   pp.pprint(e)
    day = e["day"]
    event_code = e["eventCode"]
    category = e["category"]
#    print(category)
    event_key = (category, event_code)
#   print(event_code, event_key)
    """
    if is_relay(event_key[1]):
        continue
    """
    series[event_key] = {}
    if day not in results.keys():
        results[day] = {}
    if event_key not in e.keys():
        results[day][event_key] = {}
        trials = {}
        if event_code in multis:
            if multis.index(event_code) > 3:
                del results[day][event_key] 
                day +=1
                results[day][event_key] = {}
#           print(event_key, day)
#           print(event_key)
#           print( results[day][event_key] )
            for r in e['results']:
#               print(r)
                if "bib" in r.keys():
                    bib = r["bib"]
                
                if bib not in ignore_bibs:
                     bdate = competitors[bib][2]
                     g = competitors[bib][3]
#                    cat = get_category(bdate,date0,g)
                     if sort_by == 'cat':
                         cat = category
                     else:
                         cat = get_category(bdate,date0,g)

#                    cat = r['category']
                     pool = 0
                     if results[day][event_key].get(cat) == None:
                         results[day][event_key][cat] = { pool : { 'marks' : [] } }
#                    if results[day][event_key][cat].get(pool) == None:
#                    results[event_code][cat][pool] = []
#                        results[day][event_key][cat][pool] = {'marks' : []}
#                    if not wind == None:
#                        results[day][event_key][cat][pool]['wind'] = wind
                     if 'total' in r.keys():
                         res = f'{r["total"]}'
                     if "place" in r.keys():
                         pl = r["place"]
                         if pl is None:
                             pl = noplace

#                        t = r['teamName']
#                        competitors[bib][4] = t
                     else:
                         pl = noplace
#                    print(bib, cat, res, pl)
#                    series[event_key][bib] = []
                     if 'perfsByEvent' in r.keys():
                         sl = []
                         is_dnf = False
                         for pf,pt in zip(r['perfsByEvent'].items(), r['pointsByEvent'].items()):
                             sl.append(f'{pf[1]}({+pt[1]})')
                             is_dnf = is_dnf or pf[1] == 'DNS'
                         s = '/'.join(sl)
                         series[event_key][bib]=s
                         if is_dnf:
                             res = 'DNF'

#                    pool = 0
                     results[day][event_key][cat][pool]['marks'].append((bib, res, pl))
        elif is_relay(event_code):
#           print('relay')
            cat = category
            for pool, u in enumerate(e["units"]):
                heatname = 'Heat'
                if 'heatName' in u.keys():
                    if 'Final' in u['heatName']:
                        heatname = 'Finale'
                heatnumber = u['heat']
#               for r in u["results"]:
#                   print(r)


        else:
           for pool, u in enumerate(e["units"]):
               #results[event_code] ={}
               if "windAssistance" in u.keys():
                   wind = u["windAssistance"]
               else:
                   wind = None
               #print(wind)
               heatname = 'Heat'
               if 'heatName' in u.keys():
                   if 'Final' in u['heatName']:
                       heatname = 'Finale'
               heatnumber = u['heat']
#               print(heatname, heatnumber)
               for r in u["results"]:
                   print(r)
                   if "bib" in r.keys():
                       bib = r["bib"]
                   
                   if bib not in ignore_bibs:
                        print(bib)
                        bdate = competitors[bib][2]
                        g = competitors[bib][3]
                        if sort_by == 'cat':
                            cat = category
                        else:
                            cat = get_category(bdate,date0,g)
                        
#                        print(category, cat)
                        if results[day][event_key].get(cat) == None:
                            results[day][event_key][cat] = {}
                        if results[day][event_key][cat].get(pool) == None:
                            #results[event_code][cat][pool] = []
                            results[day][event_key][cat][pool] = {'marks' : [], 'heatname' : heatname, 'heatnumber' : heatnumber}
                        if not wind == None:
                            results[day][event_key][cat][pool]['wind'] = wind
                        #
                    
                        if 'performance' in r.keys():
                            res = r['performance']
                        else:
                            res = ''

                        if "place" in r.keys():
                            pl = r["place"]
                        else:
                            pl = noplace
                      
#                    if "order" in r.keys():
#                        pl = r["order"]
                       
                        results[day][event_key][cat][pool]['marks'].append((bib, res, pl))
                        
               if event_code in ('HJ', 'SHJ', 'PV'):
                   for t in u['trials']:
                       bib = t['bib']
                       if trials.get(bib)==None:
                           trials[bib] = {}
                       height = t['height']
                       if trials[bib].get(height)==None:
                           trials[bib][height] = []
                       trials[bib][height].append(t['result'])
                   for bib in trials.keys():
                       s = ''
                       for height in sorted(trials[bib].keys() ):
                           s += height + '(' + ''.join(trials[bib][height]) + ') ' 
                       s = s.replace('.',',')
#                    print(s)
                       i0 = i1 = len(s)
                       if 'x' in s:
                           i0 = s.index('x')
                       if 'o' in s:
                           i1 = s.index('o')
                       ij = min(i0,i1)
                       if ij < len(s):
                           series[event_key][bib] = s[ij-5:]
                       else:
                           series[event_key][bib] = ''
               elif event_code in ('LJ', 'TJ', 'SP', 'DT', 'HT', 'JT', 'OT', 'BT'):
                   for t in u['trials']:
                       bib = t['bib']
                       if trials.get(bib)==None:
                           trials[bib] = {}
#                   print(event_code, t)
                       rond = t['round']
                       if trials[bib].get(rond)==None:
                           trials[bib][rond] = {}
                       trials[bib][rond]['result'] = t['result']
                       if 'wind' in t.keys():
                           trials[bib][rond]['wind'] = t['wind']

                   for bib in trials.keys():
                       s = ''
                       for rond in sorted(trials[bib].keys() ):
                           s += trials[bib][rond]['result'] 
                           if 'wind' in trials[bib][rond].keys():
                               s += "(%3.1f)" % (trials[bib][rond]['wind'])
                           s += '/'    
                       s = s.replace('.',',')
                       series[event_key][bib] = s[:-1]

                     # add code to extract event results and scores here (... use series[event_code][bib] ...)
                     #series[even_code][bib] = ""



#pp = pprint.PrettyPrinter(indent=4)
#pp.pprint(results)

#... write template for Results to xlsx workbook
wb = Workbook()

ws = wb.active
    
greenfont = Font(name='Calibri', color="0000FF00")
#greenfont = Font(name='Calibri', color=xlcolors.GREEN)
boldfont = Font(name='Calibri', bold=True, underline="single")
    
ws.title = "Resultatliste"
    
ws['a1'] = 'Stevne:';         ws['b1'] = meetname
ws['a2'] = 'Stevnested:';     ws['b2'] = venue
ws['a3'] = 'Stevnedato:';     ws['b3'] = date0.strftime('%d.%m.%Y'); ws['c3'] = date1.strftime('%d.%m.%Y')
ws['a4'] = 'Arrangør:';       #ws['b4'] = organiser_name; #b4=ws['b4']; b4.font=greenfont
ws['a5'] = 'Kontaktperson:';  ws['b5'] = '<navn>'    ; b5=ws['b5']; b5.font=greenfont
ws['a6'] = 'Erklæring*: ';    ws['b6'] = 'J'     #; b6=ws['b6']; b6.font=greenfont
ws['a7'] = 'Telefon:';        ws['b7'] = '<tlf>'     ; b7=ws['b7']; b7.font=greenfont
ws['a8'] = 'Epost:';          ws['b8'] = j['contactDetails']  ; #b8=ws['b8']; b8.font=greenfont
ws['a9'] = 'Utendørs:';       ws['b9'] = outdoors #   ; b9=ws['b9']; b9.font=greenfont
ws['a10'] = 'Kommentar:';     ws['b10'] = url
ws['a11'] = 'Kommentar:';     


row_counter = 13 

day = 1
#for day,date in zip(range(1,len(dates)+1), dates):
for day,date in enumerate(dates):
    day +=1
    ws[f'A{row_counter}'] = 'Resultater';     ws[f'B{row_counter}'] = date.strftime('%d.%m.%Y')
    row_counter +=2
    for event_key in sorted(results[day].keys()):
#       print(event_key)
        event = event_key[1]
        for cat in sorted(results[day][event_key].keys() ):
            ws["A%(row_counter)d"%vars()] = cat; arc = ws["A%(row_counter)d"%vars()]; arc.font=boldfont
            ws["B%(row_counter)d"%vars()] = events.event_spec(event,cat) ; brc = ws["B%(row_counter)d"%vars()]; brc.font=boldfont
            row_counter +=1
            heats = sorted(results[day][event_key][cat].keys() )
            for h, heat in zip(range(len(heats)), heats):
#                print('h', h, heat)
#                print( results[day][event_key][cat][heat] )
                heatname = results[day][event_key][cat][heat].get('heatname','')
                heatnumber = results[day][event_key][cat][heat].get('heatnumber','')
                ws["A%(row_counter)d"%vars()] = heatname+':';  ws["B%(row_counter)d"%vars()] = heatnumber
                if 'wind' in results[day][event_key][cat][heat].keys():
                    ws["C%(row_counter)d"%vars()] = "Vind:";  ws["D%(row_counter)d"%vars()] = results[day][event_key][cat][heat]['wind']
                row_counter +=1
#               print( results[day][event_key][cat][heat]['marks'] )
                sorted_results = sorted(results[day][event_key][cat][heat]['marks'], key=lambda tup: tup[2])
                pat = "[GJ](\d?\d)"
                match = re.search(pat,event_key[0])
                has_age = False
                if match: 
                    age = int(match.group(1))
                    has_age = True
                    if age < 11:
                        sorted_results = results[day][event_key][cat][heat]['marks']
                        random.shuffle(sorted_results)
                for i,r in zip(range(len(sorted_results)),sorted_results):
                    bib = r[0]
#                   print(bib)
#                   print(cat, r)
                    perf = r[1].replace('.',',')
                    if has_age and age < 11:
                        place = noplace
                    else:
                        place = r[2]
    
                    fn  = competitors[bib][0]
                    ln  = competitors[bib][1]
                    dob = competitors[bib][2]
                    club = competitors[bib][4]
    
                    if place == noplace:
                        pl = ''
                    else:
                        pl = i+1
                    ws["A%(row_counter)d"%vars()] = pl
                    #ws["B%(row_counter)d"%vars()] = bib
                    ws["C%(row_counter)d"%vars()] = ' '.join((fn,ln))
                    print(type(dob))
                    if isinstance(dob, str):
                        d = dob
                    else:
                        d = dob.strftime('%Y')
                    ws["D%(row_counter)d"%vars()] = d
                    ws["E%(row_counter)d"%vars()] = clubs.club_name(club)
#                   ws["E%(row_counter)d"%vars()] = club
                    ws["F%(row_counter)d"%vars()] = perf
    
    #--- extract wind for best performance from series
                    s = series[event_key].get(bib, 'no_series')
                    if event in ('LJ', 'TJ') and not s == 'no_series':
                        pat = r'/?%(perf)s\(([+-]?\d,\d)\)/?' % vars()
                        match = re.search(pat,s)
                        if match:
                            ws["G%(row_counter)d"%vars()] = match.group(1)
    
                    if not s=='no_series':
                        row_counter +=1
                        ws["A%(row_counter)d"%vars()] = s
                    row_counter +=1
            row_counter +=1
        
print("done")

xlname = slug + '-' + date0.strftime(isodateformat) + '.xlsx'
wb.save(xlname)
