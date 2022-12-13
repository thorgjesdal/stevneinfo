# -*- coding: utf-8 -*-

# TODO: 
#       + combined events results
#       + load json directly from url
#       + clean up/more modular
#
import sys
import json
import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import colors as xlcolors
from openpyxl.styles import Font, Color
import requests
import random
from collections import defaultdict
from athlib import tyrving_score 

import pprint

noplace = int(1.e10)-1
def get_category(birthdate, eventdate, gender):
    birthyear = birthdate.year
    eventyear = eventdate.year
    age = int(eventyear)-int(birthyear)

    g = {'F' : 'J', 'M' : 'G' }
    if age > 19:
        g = {'F' : 'K', 'M' : 'M' }
        if age < 35:
           a = 'S'
        else:
           a = 'V' + '%d'%(5*int(age/5))
    elif age in (18,19):
        a = '18/19'
    else:
        a = '%d'%(age)

    cat = g[gender]+a
    return cat

track_events = ['60', '80', '100', '150', '200', '300', '400', '600', '800', '1500', '2000', '3000', '5000', '10000', '60H', '80H', '100H', '110H', '200H', '300H', '400H', '1500SC', '2000SC', '3000SC']
jump_events  = ['HJ', 'PV', 'LJ', 'TJ', 'SHJ', 'SLJ']
throw_events = ['SP', 'DT', 'JT', 'HT', 'BT', 'OT']
   
def event_name(code):
    event_names = {
            '60'     : '60 meter'          , 
            '80'     : '80 meter'          , 
            '100'    : '100 meter'         , 
            '150'    : '150 meter'         , 
            '200'    : '200 meter'         , 
            '300'    : '300 meter'         , 
            '400'    : '400 meter'         , 
            '600'    : '600 meter'         , 
            '800'    : '800 meter'         , 
            '1000'   : '1000 meter'        , 
            '1500'   : '1500 meter'        , 
            '3000'   : '3000 meter'        , 
            '5000'   : '5000 meter'        , 
            '10000'  : '10000 meter'       , 
            '60H'    : '60 meter hekk'     , 
            '80H'    : '80 meter hekk'     , 
            '100H'   : '100 meter hekk'    , 
            '110H'   : '110 meter hekk'    , 
            '200H'   : '200 meter hekk'    ,
            '300H'   : '300 meter hekk'    , 
            '400H'   : '400 meter hekk'    , 
            '3000SC' : '3000 meter hinder' , 
            '1000W'  : 'Kappgang 1000 meter'        , 
            '3000W'  : 'Kappgang 3000 meter'        , 
            'HJ'     : 'Høyde'             , 
            'PV'     : 'Stav'              , 
            'LJ'     : 'Lengde'            , 
            'TJ'     : 'Tresteg'           , 
            'SP'     : 'Kule'              , 
            'DT'     : 'Diskos'            , 
            'HT'     : 'Slegge'            , 
            'JT'     : 'Spyd'              , 
            'OT'     : 'Liten ball'              , 
            'BT'     : 'Liten ball'              , 
            'DEC'    : 'Tikamp'            , 
            'HEP'    : 'Sjukamp'           ,
            'SHJ'    : 'Høyde uten tilløp' ,
            'SLJ'    : 'Lengde uten tilløp'           ,
            'SLJ'    : 'Tresteg uten tilløp'           
            }
    return event_names[code]

def event_spec(event, klasse):
    # 18.05.2020 rewrite based om implements.py form athlib
    gender = 'F'
    if klasse[0] in ('M', 'G'):
        gender = 'M'

    weight = ''
    if event == 'SP' or event == 'HT':
        if gender == 'F':
           if klasse in ('J10', 'J11', 'J12', 'J13'):
              weight = '2,0kg'
           elif klasse in ('J14', 'J15', 'J15', 'J16', 'J17'):
               weight = '3,0kg'
           elif klasse in ('J18/19', 'KU20', 'KS', 'KV35', 'KV40', 'KV45'):
               weight = '4,0kg'
           elif klasse >= 'KV50':
               weight = '3,0kg'
        elif gender == 'M':
           if klasse in ('G10', 'G11' ):
              weight = '2,0kg'
           elif klasse in ('G12', 'G13' ):
               weight = '3,0kg'
           elif klasse in ('G14', 'G15', 'MV70', 'MV75' ):
               weight = '4,0kg'
           elif klasse in ('G16', 'G17', 'MV60', 'MV65' ):
               weight = '5,0kg'
           elif klasse in ('G18/19', 'MU20', 'MV50', 'MV55' ):
               weight = '6,0kg'
           elif klasse in ('MS', 'MU23', 'MV35', 'MV40', 'MV45' ):
               weight = '7,26kg'
           elif klasse >= 'MV80':
               weight = '3,0kg'
    elif event == 'DT' :
        if gender == 'F':
           if klasse in ('J10', 'J11', 'J12', 'J13'):
              weight = '0,6kg'
           elif klasse in ('J14', 'J15'):
               weight = '0,75kg'
           elif klasse >= 'KV80':
               weight = '0,75kg'
           else:
               weight = '1,0kg'
        elif gender == 'M':
           if klasse in ('G10', 'G11' ):
              weight = '0,6kg'
           elif klasse in ('G12', 'G13' ):
               weight = '0,75kg'
           elif klasse in ('G14', 'G15' ):
               weight = '1,0kg'
           elif klasse in ('G16', 'G17', 'MV50', 'MV55' ):
               weight = '1,5kg'
           elif klasse in ('G18/19', 'MU20' ):
               weight = '1,75kg'
           elif klasse in ('MS', 'MU23', 'MV35', 'MV40', 'MV45' ):
               weight = '2,0kg'
           elif klasse >= 'MV60':
               weight = '1,0kg'
    elif event == 'JT' :
        if gender == 'F':
           if klasse in ('J10', 'J11', 'J12', 'J13', 'J14'):
              weight = '400g'
           elif klasse in ('J15', 'J16', 'J17', 'KV50', 'KV55'):
               weight = '500g'
           elif klasse in ('J18/19', 'KU20', 'KU23', 'KS', 'KV35', 'KV40', 'KV45'):
               weight = '600g'
           elif klasse >= 'KV60':
               weight = '400g'
        elif gender == 'M':
           if klasse in ('G10', 'G11', 'G13', ):
              weight = '400g'
           elif klasse in ('G14', 'G15', 'MV60', 'MV65' ):
               weight = '600g'
           elif klasse in ('G16', 'G17', 'MV50', 'MV55' ):
               weight = '700g'
           elif klasse in ('G18/19', 'MU20', 'MS', 'MU23', 'MV35', 'MV40', 'MV45' ):
               weight = '800g'
           elif klasse in ('MV70', 'MV75' ):
               weight = '500g'
           elif klasse >= 'MV80':
               weight = '400g'
    elif event == 'OT' :
        weight='150g'


    throws = {}
    throws['SP'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'KV35' : '4,0kg', 'KV40' : '4,0kg', 'KV45' : '4,0kg', 
                       'KV50' : '3,0kg', 'KV55' : '3,0kg', 'KV60' : '3,0kg', 'KV65' : '3,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg'
                       } 
    throws['DT'] = { 'J10' : '0,6kg', 'J11' : '0,6kg', 'J12' : '0,6kg', 'J13' : '0,6kg', 
                       'J14' : '0,75kg', 'J15' : '0,75kg', 'J16' : '0,75kg', 'J17' : '0,75kg',
                       'J18/19' : '1,0kg', 'KU20' : '1,0kg', 'KU23' : '1,0kg', 'KS' : '1,0kg', 
                       'G10' : '0,6kg', 'G11' : '0,6kg', 'G12' : '0,75kg', 'G13' : '0,75kg', 
                       'G14' : '1,0kg', 'G15' : '1,0kg', 'G16' : '1,5kg', 'G17' : '1,5kg',
                       'G18/19' : '1,75kg', 'MU20' : '1,75kg', 'MU23' : '2,0kg', 'MS' : '2,0kg', 
                       'MV35' : '2,0kg', 'MV40' : '2,0kg', 'MV45' : '2,0kg',
                       'MV50' : '1,5kg', 'MV55' : '1,5kg', 
                       'MV60' : '1,0kg', 'MV65' : '1,0kg', 'MV70' : '1,0kg', 'MV75' : '1,0kg' 
                       } 
    throws['HT'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg'} 
    throws['JT'] = { 'J10' : '400g', 'J11' : '400g', 'J12' : '400g', 'J13' : '400g', 
                       'J14' : '400g', 'J15' : '500g', 'J16' : '500g', 'J17' : '500g',
                       'J18/19' : '600g', 'KU20' : '600g', 'KU23' : '600g', 'KS' : '600g', 
                       'G10' : '400g', 'G11' : '400g', 'G12' : '400g', 'G13' : '400g', 
                       'G14' : '600g', 'G15' : '600g', 'G16' : '700g', 'G17' : '700g',
                       'G18/19' : '800g', 'MU20' : '800g', 'MU23' : '800g', 'MS' : '800g'} 
    throws['OT'] = { 'J10' : '150g', 'J11' : '150g', 'J12' : '150g', 'J13' : '150g', 'J14' : '150g', 
                             'G10' : '150g', 'G11' : '150g', 'G12' : '150g', 'G13' : '150g', 'G14' : '150g' }
    hurdles = {}
    hurdles['60H'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '76,2cm', 'J13' : '76,2cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'KV50' : '76,2cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '76,2cm', 'G13' : '76,2cm', 'G14' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm', 'G17' : '91,4cm',
                                 'G18/19' : '100cm','MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm' ,
                                 'MV35':'100cm', 'MV40':'91,4cm', 'MV45':'91,4cm', 'MV50':'91,4cm', 'MV55':'91,4cm', 
                                 'MV60':'84cm', 'MV65':'84cm', 'MV70':'76,2cm', 'MV75':'76,2cm', 
                                 }
    hurdles['80H'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'G14' : '84,0cm' } 
    hurdles['100H'] = { 'J16' : '76,2cm', 'J17' : '76,2cm', 'J18/19' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm'}
    hurdles['110H'] = { 'G17' : '91,4cm', 'G18/19' : '100cm','MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm' }
    hurdles['200H'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '68,0cm', 'J13' : '68,0cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '68,0cm', 'G13' : '68,0cm', 'G14' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '76,2cm', 'G17' : '76,2cm',
                                 'G18/19' : '76,2cm','MU20' : '76,2cm', 'MU23' : '76,2cm', 'MS' : '76,2cm' }
    hurdles['300H'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm',
                                 'default':''}
    hurdles['400H'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm' }


    if event in ('SP', 'DT', 'JT', 'HT', 'OT'):
       #e = event + ' ' + throws[event][klasse]
#       e = event_name(event) + ' ' + throws[event].get(klasse,'')
       e = event_name(event) + ' ' + weight
    elif event in ('60H', '80H', '100H', '110H', '200H', '300H', '400H'): 
#      e = event_name(event) + ' ' + hurdles[event][klasse]
       e = event_name(event) + ' ' + hurdles[event].get(klasse,'')
    else:
       e = event_name(event)

    return e
#---------------------------------------
if len(sys.argv) < 2:
   sys.exit("Usage: %s <infile>" % sys.argv[0])
   
url = sys.argv[1]+'json'
print(url)

r=requests.get(url)
j = json.loads(r.text)

slug = j['slug']
   
d  = j['date']
d2 = j['finishDate']
isodateformat = "%Y-%m-%d"
date0 = datetime.datetime.strptime(d, isodateformat)
date1 = datetime.datetime.strptime(d2, isodateformat)
dates = []
d = date0
while d <= date1:
    dates.append(d)
    d += datetime.timedelta(days=1)


ignore_bibs = []
competitors = {}
for c in j['competitors']:
    fn = ''; ln = ''; dob= ''; t=''
    bib  = c['competitorId']
    if 'firstName' in c.keys():
        fn   = c['firstName']
    if 'lastName' in c.keys():
        ln   = c['lastName']
    if 'dateOfBirth' in c.keys():
        d    = c['dateOfBirth']
        dob  = datetime.datetime.strptime(d, isodateformat)
    if 'teamName' in c.keys():
        t    = c['teamName']
    if 'teamId' in c.keys():
        t2    = c['teamId']

    if 'gender' in c.keys():
        g = c['gender']
    if fn=='':
        ignore_bibs.append(bib)
    else:
        competitors[bib] = (fn, ln, dob, g, t2)


poolnr = 0
results ={}
series = {}
scores = []
for e in j["events"]:
    day = e["day"]
    event_code = e["eventCode"]

    for pool, u in enumerate(e["units"]):
        for r in u["results"]:
            if "bib" in r.keys():
                bib = r["bib"]
            
            age = int(date0.strftime('%Y'))-int(competitors[bib][2].strftime('%Y'))
            if age < 11:
                continue
            elif age == 18:
                age=19

            gender = competitors[bib][3]
            g='G'
            if gender=='F':
                g='J'
            category = f'{g}{age}'

            if category not in results.keys():
                results[category] = {}

            if bib not in results[category].keys():
                results[category][bib] = { 'runs' : [], 'jumps' : [], 'throws' : [], 'count' : 0, 'score' : 0 }

            res = r['performance']
            if res not in ['DNF', 'DNS', 'NM', 'NH', 'DQ', '']:
                if event_code == 'BT':
                    event_code='OT'
                tyrving = tyrving_score(gender,age,event_code,res)


            if event_code in track_events:
                results[category][bib]['runs'].append((event_code, res, tyrving))
            elif event_code in jump_events:
                results[category][bib]['jumps'].append((event_code, res, tyrving))
            elif event_code in throw_events:
                results[category][bib]['throws'].append((event_code, res, tyrving))

for cat in results.keys():
    for bib in results[cat].keys():
        count =   int( len(results[cat][bib]['runs']) > 0 ) + int( len(results[cat][bib]['jumps']) > 0 ) + int( len(results[cat][bib]['throws']) > 0 )   
        results[cat][bib]['count'] = count

        if count == 3:
            score = 0.
            for ev in ['runs', 'jumps', 'throws']:
                score += max([ results[cat][bib][ev][i][2] for i in range( len(results[cat][bib][ev]) ) ]  )
            results[cat][bib]['score'] = score

#pp = pprint.PrettyPrinter(indent=4)
#pp.pprint(results)
#exit()

#... write template for Results to xlsx workbook
wb = Workbook()

ws = wb.active
    
row_counter = 1 
columns = {'60' : 'D', '100' : 'D', '600':'E' , '800' : 'E', 'HJ':'F', 'LJ': 'G', 'SP' : 'H', 'OT' : 'I' }
for cat in results.keys():
    for bib in results[cat].keys():
        if competitors[bib][4] == 'KOLL':
            ws[f'A{row_counter}'] = cat
            ws[f'B{row_counter}'] = competitors[bib][0]
            ws[f'C{row_counter}'] = competitors[bib][1]
            for ev in ['runs', 'jumps', 'throws']:
                for i in results[cat][bib][ev]:
                    column = columns[i[0]]
                    ws[f'{column}{row_counter}'] = f'{i[1]} ({i[2]})'
            ws[f'J{row_counter}'] = results[cat][bib]['score']

            row_counter +=1
                


xlname = slug + '-' + date0.strftime(isodateformat) + '_score.xlsx'
print(xlname)
wb.save(xlname)
