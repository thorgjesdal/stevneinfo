#!/usr/bin/python
# -*- coding: utf-8 -*-
import sys
from xml.etree import ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import colors as xlcolors
from openpyxl.styles import Font, Color
#from tabulate import tabulate
from reportlab.lib import colors as rlcolors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from unidecode import unidecode
#import requests
#from bs4 import BeautifulSoup


def read_xml_into_tree(infile):
   with open(infile, 'rt') as f:
      tree = ET.parse(f)
   return tree

def extract_competition_data(tree):
   s = tree.find('.//Competition')
   mn = s.attrib['name']
   md = s.attrib['startDate']

   v = s.find('CompetitionVenue')
   mv = v.attrib['startingvenue']
   return{'meet_name' : mn, 'meet_date' : md, 'venue' : mv}

def output_file_name(tree):
   cdata = extract_competition_data(tree)
   mn = cdata['meet_name']
   md = cdata['meet_date']
   mv = cdata['venue']

   sn =  mn + ' ' + mv + ' ' + md
   sn = sn.replace(' ','_')
   sn = sn.replace('/','-')
   fname = unidecode( sn.replace(' ','_') )
   return fname

def save_xml_copy(tree):
   fname = output_file_name(tree)
   # save a copy of the element tree
   tree.write(fname+'.xml', encoding="utf-8")

def istrack(event):
    return 'meter' in event

def ishurdles(event):
    return istrack(event) and 'hekk' in event

def issteeple(event):
    return istrack(event) and 'hinder' in event

def isfield(event):
    return 'meter' not in event

def isvjump(event):
    return isfield(event.encode('utf-8')) and event in [u'Høyde', u'Stav', u'Høyde uten tilløp']

def ishjump(event):
    return isfield(event.encode('utf-8')) and event in [u'Lengde', u'Lengde satssone', u'Tresteg', u'Lengde uten tilløp', u'Tresteg uten tilløp']

def isthrow(event):
    return isfield(event.encode('utf-8')) and event in [u'Kule', u'Diskos', u'Slegge', u'Spyd', u'Vektkast', u'Liten ball']

def sort_athletes_by_class_by_event(tree):
    athlete_by_class_by_event = {}
    for c in tree.findall('.//Competitor'):
       p = c.find('Person')
       club = p.attrib['clubName']
       cs = club.split(',')
       if len(cs) > 1:
          club = cs[1].strip() + ' ' + cs[0].strip()
       n = p.find('./Name')
       name = "Name"
       fn = n.find('Given')
       en = n.find('Family')
       bd = p.find('BirthDate')
       name = fn.text + ' ' + en.text
       dd = int(bd.attrib['day'])
       mm = int(bd.attrib['month'])
       yyyy = int(bd.attrib['year'])
       dob = "%(dd)02d.%(mm)02d.%(yyyy)04d" % vars()
       ec = c.find('./Entry/EntryClass')
       klasse = ec.attrib['classCode']
       ec = c.find('./Entry/Exercise')
       event = ec.attrib['name'] 
       athlete_key = name + dob + club
      
       if event not in athlete_by_class_by_event.keys():
          athlete_by_class_by_event[event]={}
       if klasse not in athlete_by_class_by_event[event].keys():
          athlete_by_class_by_event[event][klasse]=[]
       athlete_by_class_by_event[event][klasse].append({'name': name, 'dob': dob, 'club' : club }) 

    return athlete_by_class_by_event


def list_entries(tree):
    entries_list = []
    for c in tree.findall('.//Competitor'):
       p = c.find('Person')
       club = p.attrib['clubName']
       cs = club.split(',')
       if len(cs) > 1:
          club = cs[1].strip() + ' ' + cs[0].strip()
       n = p.find('./Name')
       name = "Name"
       fn = n.find('Given')
       en = n.find('Family')
       bd = p.find('BirthDate')
       dd = int(bd.attrib['day'])
       mm = int(bd.attrib['month'])
       yyyy = int(bd.attrib['year'])
       dob = "%(dd)02d.%(mm)02d.%(yyyy)04d" % vars()
       ec = c.find('./Entry/EntryClass')
       klasse = ec.attrib['classCode']
       ec = c.find('./Entry/Exercise')
       event = ec.attrib['name'] 
       entry = {'first_name': fn.text, 'last_name' : en.text, 'birth_date' : dob, 'club' : club, 'event' : event}

       entries_list.append( entry )
    
    return entries_list

def sort_athletes_by_event_by_class(tree):
    athlete_by_event_by_class = {}
    for c in tree.findall('.//Competitor'):
       p = c.find('Person')
       club = p.attrib['clubName']
       cs = club.split(',')
       if len(cs) > 1:
          club = cs[1].strip() + ' ' + cs[0].strip()
       n = p.find('./Name')
       name = "Name"
       fn = n.find('Given')
       en = n.find('Family')
       bd = p.find('BirthDate')
       name = fn.text + ' ' + en.text
       dd = int(bd.attrib['day'])
       mm = int(bd.attrib['month'])
       yyyy = int(bd.attrib['year'])
       dob = "%(dd)02d.%(mm)02d.%(yyyy)04d" % vars()
       ec = c.find('./Entry/EntryClass')
       klasse = ec.attrib['classCode']
       ec = c.find('./Entry/Exercise')
       event = ec.attrib['name'] 
       athlete_key = name + dob + club

       if klasse not in athlete_by_event_by_class.keys():
           athlete_by_event_by_class[klasse] = {}
       if event not in athlete_by_event_by_class[klasse].keys():
           athlete_by_event_by_class[klasse][event] = [] 
       athlete_by_event_by_class[klasse][event].append({'name': name, 'dob': dob, 'club' : club }) 


    return athlete_by_event_by_class

def sort_events_by_athlete(tree):
    events_by_athlete= {}
    for c in tree.findall('.//Competitor'):
       p = c.find('Person')
       club = p.attrib['clubName']
       g = p.attrib['sex']
       cs = club.split(',')
       if len(cs) > 1:
          club = cs[1].strip() + ' ' + cs[0].strip()
       n = p.find('./Name')
       name = "Name"
       fn = n.find('Given')
       en = n.find('Family')
       bd = p.find('BirthDate')
       name = fn.text + ' ' + en.text
       dd = int(bd.attrib['day'])
       mm = int(bd.attrib['month'])
       yyyy = int(bd.attrib['year'])
       dob = "%(dd)02d.%(mm)02d.%(yyyy)04d" % vars()
       ec = c.find('./Entry/EntryClass')
       klasse = ec.attrib['classCode']
       ec = c.find('./Entry/Exercise')
       #event = klasse + ' ' +ec.attrib['name'] 
       event = ( klasse, ec.attrib['name'] )
       athlete_key = '+'.join((fn.text, en.text, dob, g, club))

       if athlete_key not in events_by_athlete.keys():
           events_by_athlete[athlete_key] = []
       if event not in events_by_athlete[athlete_key]:
           events_by_athlete[athlete_key].append(event)

    return events_by_athlete

def write_start_lists_as_html(tree):
    competition_data = extract_competition_data(tree)
    #mn = competition_data['meet_name']
    #md = competition_data['meet_date']
    #mv = competition_data['venue']
    mn = competition_data['meet_name'].encode('utf-8')
    md = competition_data['meet_date'].encode('utf-8')
    mv = competition_data['venue'].encode('utf-8')

    athlete_by_class_by_event = sort_athletes_by_class_by_event(tree)

    fname = output_file_name(tree)
    of = open (fname+"_deltagere_pr_ovelse.html".encode('utf-8'), 'w')
    of.write(""" <!DOCTYPE html>
    <meta charset="UTF-8">
    <html>
    <body>
    <title> %(mn)s </title>
    <h1> %(mn)s </h1>
    """ % vars() )
    #of.write("%s, %s" % ( md, mv.encode('utf-8') ) )
    of.write("%s, %s" % ( md, mv ) )
    event_keys = athlete_by_class_by_event.keys() 
    event_keys.sort()
    for event_key in event_keys:
        class_keys = athlete_by_class_by_event[event_key].keys()
        class_keys.sort()
        for class_key in  class_keys:
           of.write( "<h2>%s %s </h2>\n<ul style=\"list-style-type:none\">\n"% (event_key.encode('utf-8'), class_key) )
           for athlete in athlete_by_class_by_event[event_key][class_key]:
               of.write("<li>" +athlete['name'].encode('utf-8') + ' (' + athlete['dob'][-4:] +'), ' + athlete['club'].encode('utf-8') +  "</li>\n" )
           of.write("</ul>\n")
    
    of.write("""</body>
    </html>""")
    of.close()


def write_opentrack_import(tree):
    competition_data = extract_competition_data(tree)
    mn = competition_data['meet_name']
    md = competition_data['meet_date']
    mv = competition_data['venue']

    events = list_events(tree)
    events_by_athlete = sort_events_by_athlete(tree)

    #... write template for Results to xlsx workbook
    wb = Workbook()
    ws = wb.active
    
    row_counter = 1
    ws["A1"] = 'Competitor Id'
    ws["B1"] = 'National Id'
    ws["C1"] = 'First name'
    ws["D1"] = 'Last name'
    ws["E1"] = 'Gender'
    ws["F1"] = 'Date of birth'
    ws["G1"] = 'Team ID'
    ws["H1"] = 'Nationality'
    ws["I1"] = 'Event'
    ws["J1"] = 'Pb'
    ws["K1"] = 'Sb'
    ws["M1"] = 'Event selection'
    row_counter = 2

    jf = 0
    jt = 0
    full_events = {}
    for e in events:
        if isfield(e[1]):
            jf +=1
            event_ref = "F%02d"%jf
        else:
            jt +=1
            event_ref = "T%02d"%jt

        full_events[ ( class_code(e[0]) , e[1] ) ]  = event_ref + ' - ' + ' '.join(( class_code(e[0]), event_spec(e[1], class_code(e[0])) ))
        ws["M%d"%row_counter] = event_ref + ' - '  + ' '.join([e[0], event_spec(e[1], class_code(e[0]))])
        ws["O%d"%row_counter] = event_ref
        ws["P%d"%row_counter] = event_code(e[1])
        ws["Q%d"%row_counter] = age_group(class_code(e[0]))
        ws["R%d"%row_counter] = gender(class_code(e[0]))
        ws["S%d"%row_counter] = class_code(e[0])
#       ws["S%d"%row_counter] = age_group(class_code(e[0]))

        ws["U%d"%row_counter] = ' '.join(( class_code(e[0]), event_spec(e[1], class_code(e[0])) ))
        ws["V%d"%row_counter] = '1'
        ws["W%d"%row_counter] = '1'
        ws["X%d"%row_counter] = '12:00'
        
        row_counter +=1
    ws.insert_cols(13)

    row_counter = 2    
    bib = 0
    for key in events_by_athlete.keys():
        bib +=1
        k = key.split('+')
        fn = k[0]
        en = k[1]
        dob = '-'.join(( k[2][6:10], k[2][3:5], k[2][0:2] ))
        g = k[3]
        club = k[4]

        for e in events_by_athlete[key]:
            ws["A%d"%row_counter] = bib
            ws["C%d"%row_counter] = fn
            ws["D%d"%row_counter] = en
            ws["E%d"%row_counter] = gender(g)
            ws["F%d"%row_counter] = dob
            ws["G%d"%row_counter] = club_code(club)
            ws["I%d"%row_counter] = full_events[e]
            row_counter +=1

    fname = output_file_name(tree)
    xlname = fname+'_opentrack.xlsx'
    wb.save(xlname)

def write_xlsx_results_template(tree):
    competition_data = extract_competition_data(tree)
    mn = competition_data['meet_name']
    md = competition_data['meet_date']
    mv = competition_data['venue']

    athlete_by_event_by_class = sort_athletes_by_event_by_class(tree)

    #... write template for Results to xlsx workbook
    wb = Workbook()
    ws = wb.active
    
    greenfont = Font(name='Calibri', color=xlcolors.GREEN)
    boldfont = Font(name='Calibri', bold=True, underline="single")
    
    ws.title = "Resultatliste"
    
    ws['a1'] = 'Stevne:';         ws['b1'] = mn
    ws['a2'] = 'Stevnested:';     ws['b2'] = mv
    ws['a3'] = 'Stevnedato:';     ws['b3'] = md          ; ws['c3'] = '<til dato>'; c3=ws['c3']; c3.font=greenfont
    ws['a4'] = 'Arrangør:';       ws['b4'] = '<arrangør>'; b4=ws['b4']; b4.font=greenfont
    ws['a5'] = 'Kontaktperson:';  ws['b5'] = '<navn>'    ; b5=ws['b5']; b5.font=greenfont
    ws['a6'] = 'Erklæring*: ';    ws['b6'] = '<J/N>'     ; b6=ws['b6']; b6.font=greenfont
    ws['a7'] = 'Telefon:';        ws['b7'] = '<tlf>'     ; b7=ws['b7']; b7.font=greenfont
    ws['a8'] = 'Epost:';          ws['b8'] = '<e-post>'  ; b8=ws['b8']; b8.font=greenfont
    ws['a9'] = 'Utendørs:';       ws['b9'] = '<J/N>'     ; b9=ws['b9']; b9.font=greenfont
    ws['a10'] = 'Kommentar:'
    
    ws['a12'] = 'Resultater';     ws['b12'] = md
    
    row_counter = 14
    class_keys = athlete_by_event_by_class.keys()
    class_keys.sort()
    for klasse in class_keys:
       event_keys = athlete_by_event_by_class[klasse].keys()
       event_keys.sort()
       for event in event_keys:
           
           e = event_spec(event,klasse)
           ws["A%(row_counter)d"%vars()] = klasse; arc = ws["A%(row_counter)d"%vars()]; arc.font=boldfont
           ws["B%(row_counter)d"%vars()] = e     ; brc = ws["B%(row_counter)d"%vars()]; brc.font=boldfont
           ws["C%(row_counter)d"%vars()] = "<spesiell konkurransestatus>";  crc = ws["C%(row_counter)d"%vars()]; crc.font=greenfont
           row_counter +=1

           if istrack(event):
               ws["A%(row_counter)d"%vars()] = "<Heat | Finale:>";  arc = ws["A%(row_counter)d"%vars()]; arc.font=greenfont
               ws["C%(row_counter)d"%vars()] = "Vind:";  crc = ws["C%(row_counter)d"%vars()]; crc.font=greenfont
               row_counter +=1
          
           for athlete in athlete_by_event_by_class[klasse][event]:
              ws["C%(row_counter)d"%vars()] = athlete['name']
              ws["D%(row_counter)d"%vars()] = athlete['dob'][-4:]
              ws["E%(row_counter)d"%vars()] = athlete['club']
              ws["F%(row_counter)d"%vars()] = "<resultat>"
              if ishjump(event):
                 ws["G%(row_counter)d"%vars()] = "<vind>";  grc = ws["G%(row_counter)d"%vars()]; grc.font=greenfont
                 ws["H%(row_counter)d"%vars()] = "<resultat>";  hrc = ws["H%(row_counter)d"%vars()]; hrc.font=greenfont
                 ws["I%(row_counter)d"%vars()] = "<vind>";  irc = ws["I%(row_counter)d"%vars()]; irc.font=greenfont
              if isfield(event):
                 row_counter +=1 # add blank line for series
                 ws["A%(row_counter)d"%vars()] = "<hopp-/kastserie>";  arc = ws["A%(row_counter)d"%vars()]; arc.font=greenfont
    
              row_counter +=1
           row_counter +=1
           
    
    fname = output_file_name(tree)
    xlname = fname+'.xlsx'
    wb.save(xlname)

def make_horizontal_protocol(tree, event, classes):
    """make the event protocol sheet for a horizontal jump or throw event
    writes the protocol sheet to a pdf
    Input:
    	tree: ElementTree
    	event: the event (fails if event is not a hjump or throw)
 	classes: list of classes """
    if not ishjump(event) and not isthrow(event):
       sys.exit('make_horizontal_protocol: event is not hjump or throw')
    athlete_by_event_by_class = sort_athletes_by_event_by_class(tree)
    athlete_by_class_by_event = sort_athletes_by_class_by_event(tree)
 
    eventclass = event + ' ' + '+'.join(classes)
 
    fname = output_file_name(tree) + '_' + event + '-' + '+'.join(classes) +'.pdf'
    fname = fname.replace(' ', '_')
    fname = fname.replace('/', '-')
    print(fname)
    doc = SimpleDocTemplate(fname, pagesize=A4)
    doc.pagesize = landscape(A4)
 
    rows_on_page = 12
    # container for the 'Flowable' objects
    elements = []
    
    styles = getSampleStyleSheet()
 
    data= [ [event + ': ' + ', '.join(classes)],
            ['Klasse', 'Navn', 'F.år', 'Klubb', 'Forsøk 1', 'Forsøk 2', 'Forsøk 3', 'Forsøk 4', 'Forsøk 5', 'Forsøk 6', 'Resultat'] ]
    if ishjump(event):
       data[1].append('Vind')
 
    rows = 0
    for c in classes:
       if c in athlete_by_class_by_event[event].keys():
          for athlete in athlete_by_class_by_event[event][c]:
             data.append( [ c, athlete['name'], athlete['dob'][-4:], athlete['club'] ] )
             rows +=1
    pages = int(rows/(rows_on_page-1)) + 1
    print(pages)
 
    if rows%rows_on_page > 5:
       pages +=1
    rows_in_table = pages*rows_on_page 
 
    if rows < rows_in_table:
       for i in range(rows_in_table-rows-2):
          data.append([ ' ' ])
 
 
 
    t=Table(data, [1.9*cm, 6.0*cm, 2.1*cm, 2.6*cm , 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm], rows_in_table*[1.4*cm], repeatRows=2)
 
    t.setStyle(TableStyle([
                           ('SPAN',(0,0),(-1,0)),
                           ('ALIGN',(0,1),(0,-1),'CENTER'),
                           ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                           ('ALIGN',(1,1),(1,-1),'LEFT'),
                           ('INNERGRID', (0,0), (-1,-1), 0.25, rlcolors.black),
                           ('BOX', (0,0), (-1,-1), 0.25, rlcolors.black),
                           ]))
    elements.append(t)
    # write the document to disk
    doc.build(elements)

def make_vertical_protocol(tree, event, classes):
    """make the event protocol sheet for a horizontal jump or throw event
    writes the protocol sheet to a pdf
    Input:
    	tree: ElementTree
    	event: the event (fails if event is not a vjump )
 	classes: list of classes """
    if not isvjump(event):
       sys.exit('make_vertical_protocol: event is not vjump')
    athlete_by_event_by_class = sort_athletes_by_event_by_class(tree)
    athlete_by_class_by_event = sort_athletes_by_class_by_event(tree)
 
    eventclass = event + ' ' + '+'.join(classes)
 
    fname = output_file_name(tree) + '_' + event + '-' + '+'.join(classes) +'.pdf'
    fname = fname.replace(' ', '_')
    fname = fname.replace('/', '-')
    print(fname)
    doc = SimpleDocTemplate(fname, pagesize=A4)
    doc.pagesize = landscape(A4)
 
    rows_on_page = 22
    # container for the 'Flowable' objects
    elements = []
    
    styles = getSampleStyleSheet()
 
    data= [ [event + ': ' + ', '.join(classes) ],
            ['Klasse', 'Navn', 'F.år', 'Klubb', '', '', '', '', '', '','','','','','','', 'Res', 'Pl' ] ]
 
    rows = 0
    for c in classes:
       if c in athlete_by_class_by_event[event].keys():
          for athlete in athlete_by_class_by_event[event][c]:
             data.append( [ c, athlete['name'], athlete['dob'][-4:], athlete['club'][0:11] ] )
             rows +=1
    pages = int(rows/(rows_on_page-1)) + 1
    print(pages)
 
    if rows%rows_on_page > 15:
       pages +=1
    rows_in_table = pages*rows_on_page 
 
    if rows < rows_in_table:
       for i in range(rows_in_table-rows-2):
          data.append([ ' ' ])
 
 
 
    t=Table(data, [1.9*cm, 6.0*cm, 2.1*cm, 2.6*cm , 0.8*cm, 0.8*cm, 0.8*cm, 0.8*cm, 0.8*cm, 0.8*cm, 0.8*cm,0.8*cm,0.8*cm,0.8*cm,0.8*cm, 0.8*cm, 1.6*cm, 0.8*cm], rows_in_table*[0.7*cm], repeatRows=2)
 
    t.setStyle(TableStyle([
                           ('SPAN',(0,0),(-1,0)),
                           ('ALIGN',(0,1),(0,-1),'CENTER'),
                           ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                           ('ALIGN',(1,1),(1,-1),'LEFT'),
                           ('INNERGRID', (0,0), (-1,-1), 0.25, rlcolors.black),
                           ('BOX', (0,0), (-1,-1), 0.25, rlcolors.black),
                           ]))
    elements.append(t)
    # write the document to disk
    doc.build(elements)

def list_events(tree):
    event_list = []
    for c in tree.findall('.//Competitor'):
       ec = c.find('./Entry/EntryClass')
       kl= ec.attrib['shortName']
       ec = c.find('./Entry/Exercise')
       ev= ec.attrib['name'] 
       event = (kl,ev)
       if event not in event_list:
           event_list.append(event)
    event_list.sort()
    return event_list

def make_crosstable(tree):
    events_by_athlete = sort_events_by_athlete(tree)
    crosstable = {}

    for a in events_by_athlete.keys():
        for e1 in events_by_athlete[a]:
            if e1 not in crosstable.keys():
                crosstable[e1] = {}
            for e2 in events_by_athlete[a]:
                if e2 not in crosstable[e1].keys():
                    crosstable[e1][e2] = 0
                crosstable[e1][e2] +=1
    return crosstable

def event_code(event):
    event_codes = {
            u'60 meter'          : '60', 
            u'80 meter'          : '80', 
            u'100 meter'         : '100', 
            u'150 meter'         : '150', 
            u'200 meter'         : '200', 
            u'300 meter'         : '300', 
            u'400 meter'         : '400', 
            u'600 meter'         : '600', 
            u'800 meter'         : '800', 
            u'1000 meter'        : '1000', 
            u'1500 meter'        : '1500', 
            u'3000 meter'        : '3000', 
            u'5000 meter'        : '5000', 
            u'10000 meter'       : '10000', 
            u'60 meter hekk'     : '60H', 
            u'80 meter hekk'     : '80H', 
            u'100 meter hekk'    : '100H', 
            u'110 meter hekk'    : '110H', 
            u'200 meter hekk'    : '200H', 
            u'300 meter hekk'    : '300H', 
            u'400 meter hekk'    : '400H', 
            u'3000 meter hinder' : '3000SC', 
            u'Høyde'             : 'HJ', 
            u'Stav'              : 'PV', 
            u'Lengde'            : 'LJ', 
            u'Lengde satssone'   : 'LJ', 
            u'Tresteg'           : 'TJ', 
            u'Tresteg satssone'  : 'TJ', 
            u'Kule'              : 'SP', 
            u'Diskos'            : 'DT', 
            u'Slegge'            : 'HT', 
            u'Spyd'              : 'JT', 
            u'Tikamp'            : 'DEC', 
            u'Sjukamp'           : 'HEP' 
            }
    return event_codes[event]
 
def event_name(code):
    event_names = {
            '60'     : '60 meter'          , 
            '80'     : '80 meter'          , 
            '100'    : '100 meter'         , 
            '150'    : '150 meter'         , 
            '200'    : '200 meter'         , 
            '300'    : '300 meter'         , 
            '400'    : '400 meter'         , 
            '600'    : '600 meter'         , 
            '800'    : '800 meter'         , 
            '1000'   : '1000 meter'        , 
            '1500'   : '1500 meter'        , 
            '3000'   : '3000 meter'        , 
            '5000'   : '5000 meter'        , 
            '10000'  : '10000 meter'       , 
            '60H'    : '60 meter hekk'     , 
            '80H'    : '80 meter hekk'     , 
            '100H'   : '100 meter hekk'    , 
            '110H'   : '110 meter hekk'    , 
            '200H'   : '200 meter hekk'    ,
            '300H'   : '300 meter hekk'    , 
            '400H'   : '400 meter hekk'    , 
            '3000SC' : '3000 meter hinder' , 
            'HJ'     : 'Høyde'             , 
            'PV'     : 'Stav'              , 
            'LJ'     : 'Lengde'            , 
            'TJ'     : 'Tresteg'           , 
            'SP'     : 'Kule'              , 
            'DT'     : 'Diskos'            , 
            'HT'     : 'Slegge'            , 
            'JT'     : 'Spyd'              , 
            'DEC'    : 'Tikamp'            , 
            'HEP'    : 'Sjukamp'           
            }
    return event_names[code]

def class_code(name):
    class_codes = {
            u'6 år Fellesklasse' : 'F6' ,
            u'7 år Fellesklasse' : 'F7' ,
            'Gutter 8'     : 'G 8'          , 
            'Gutter 9'     : 'G 9'          , 
            'Gutter 10'    : 'G10'          , 
            'Gutter 11'    : 'G11'          , 
            'Gutter 12'    : 'G12'          , 
            'Gutter 13'    : 'G13'          , 
            'Gutter 14'    : 'G14'          , 
            'Gutter 15'    : 'G15'          , 
            'Gutter 16'    : 'G16'          , 
            'Gutter 17'    : 'G17'          , 
            'Gutter 18/19' : 'G18/19'       , 
            'Menn junior'  : 'MJ'           , 
            'Menn U20'     : 'MU20'         , 
            'Menn U23'     : 'MU23'         , 
            'Menn senior'  : 'MS'           , 
            'Menn veteraner' : 'MV'         , 
            'Jenter 8'     : 'J 8'          , 
            'Jenter 9'     : 'J 9'          , 
            'Jenter 10'    : 'J10'          , 
            'Jenter 11'    : 'J11'          , 
            'Jenter 12'    : 'J12'          , 
            'Jenter 13'    : 'J13'          , 
            'Jenter 14'    : 'J14'          , 
            'Jenter 15'    : 'J15'          , 
            'Jenter 16'    : 'J16'          , 
            'Jenter 17'    : 'J17'          , 
            'Jenter 18/19' : 'J18/19'       , 
            'Kvinner junior'  : 'KJ'        , 
            'Kvinner U20'     : 'KU20'      , 
            'Kvinner U23'     : 'KU23'      , 
            'Kvinner senior'  : 'KS'        , 
            'Kvinner veteraner' : 'KV'      ,
            'Ikke valgt klasse' : 'IVK'
            }
    return class_codes[name.strip()]

def age_group(class_code):
    age_groups = {
            'F6'    : 'U7',
            'F7'    : 'U8',
            'G 8'    : 'U9',
            'G 9'    : 'U10',
            'G10'    : 'U11',
            'G11'    : 'U12',
            'G12'    : 'U13',
            'G13'    : 'U14',
            'G14'    : 'U15',
            'G15'    : 'U16',
            'G16'    : 'U17',
            'G17'    : 'U18',
            'G18/19' : 'U20',
            'MJ'     : 'U20' ,
            'MS'     : 'S' ,
            'MV'     : 'V35' ,
            'MV35'   : 'V35' ,
            'J 8'    : 'U9',
            'J 9'    : 'U10',
            'J10'    : 'U11',
            'J11'    : 'U12',
            'J12'    : 'U13',
            'J13'    : 'U14',
            'J14'    : 'U15',
            'J15'    : 'U16',
            'J16'    : 'U17',
            'J17'    : 'U18',
            'J18/19' : 'U20',
            'KJ'     : 'U20' ,
            'KS'     : 'S'  ,
            'IVK'   : 'ALL'  
            }
    """
    age_groups = {
            'F6'    : 'U7MF',
            'F7'    : 'U8MF',
            'G 8'    : 'U9B',
            'G 9'    : 'U10B',
            'G10'    : 'U11B',
            'G11'    : 'U12B',
            'G12'    : 'U13B',
            'G13'    : 'U14B',
            'G14'    : 'U15B',
            'G15'    : 'U16B',
            'G16'    : 'U17B',
            'G17'    : 'U18B',
            'G18/19' : 'U20M',
            'MJ'     : 'U20M' ,
            'MS'     : 'SM' ,
            'MV'     : 'V35M' ,
            'MV35'   : 'V35M' ,
            'J 8'    : 'U9G',
            'J 9'    : 'U10G',
            'J10'    : 'U11G',
            'J11'    : 'U12G',
            'J12'    : 'U13G',
            'J13'    : 'U14G',
            'J14'    : 'U15G',
            'J15'    : 'U16G',
            'J16'    : 'U17G',
            'J17'    : 'U18G',
            'J18/19' : 'U20W',
            'KJ'     : 'U20W' ,
            'KS'     : 'SW' 
            }
    """

    return age_groups[class_code]

def gender(class_code):
    if class_code[0] in ('G', 'M'):
        g = 'M'
    elif class_code[0] in ('J', 'K'):
        g = 'F'
    else:
        g = 'MF'

    return g

def event_spec(event, klasse):
    throws = {}
    throws['Kule'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg'} 
    throws['Diskos'] = { 'J10' : '0,6kg', 'J11' : '0,6kg', 'J12' : '0,6kg', 'J13' : '0,6kg', 
                       'J14' : '0,75kg', 'J15' : '0,75kg', 'J16' : '0,75kg', 'J17' : '0,75kg',
                       'J18/19' : '1,0kg', 'KU20' : '1,0kg', 'KU23' : '1,0kg', 'KS' : '1,0kg', 
                       'G10' : '0,6kg', 'G11' : '0,6kg', 'G12' : '0,75kg', 'G13' : '0,75kg', 
                       'G14' : '1,0kg', 'G15' : '1,0kg', 'G16' : '1,5kg', 'G17' : '1,5kg',
                       'G18/19' : '1,75kg', 'MU20' : '1,75kg', 'MU23' : '2,0kg', 'MS' : '2,0kg'} 
    throws['Slegge'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg'} 
    throws['Spyd'] = { 'J10' : '400g', 'J11' : '400g', 'J12' : '400g', 'J13' : '400g', 
                       'J14' : '400g', 'J15' : '500g', 'J16' : '500g', 'J17' : '500g',
                       'J18/19' : '600g', 'KU20' : '600g', 'KU23' : '600g', 'KS' : '600g', 
                       'G10' : '400g', 'G11' : '400g', 'G12' : '400g', 'G13' : '400g', 
                       'G14' : '600g', 'G15' : '600g', 'G16' : '700g', 'G17' : '700g',
                       'G18/19' : '800g', 'MU20' : '800g', 'MU23' : '800g', 'MS' : '800g'} 
    throws['Liten Ball'] = { 'J10' : '150g', 'J11' : '150g', 'J12' : '150g', 'J13' : '150g', 'J14' : '150g', 
                             'G10' : '150g', 'G11' : '150g', 'G12' : '150g', 'G13' : '150g', 'G14' : '150g' }
    hurdles = {}
    hurdles['60 meter hekk'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '76,2cm', 'J13' : '76,2cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '76,2cm', 'G13' : '76,2cm', 'G14' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm', 'G17' : '91,4cm',
                                 'G18/19' : '100cm','MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm' }
    hurdles['80 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'G14' : '84,0cm' } 
    hurdles['100 meter hekk'] = { 'J16' : '76,2cm', 'J17' : '76,2cm', 'J18/19' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm'}
    hurdles['110 meter hekk'] = { 'G17' : '91,4cm', 'G18/19' : '100cm','MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm' }
    hurdles['200 meter hekk'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '68,0cm', 'J13' : '68,0cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '68,0cm', 'G13' : '68,0cm', 'G14' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '76,2cm', 'G17' : '76,2cm',
                                 'G18/19' : '76,2cm','MU20' : '76,2cm', 'MU23' : '76,2cm', 'MS' : '76,2cm' }
    hurdles['300 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm' }
    hurdles['400 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm' }

    if isthrow(event):
       e = event + ' ' + throws[event][klasse]
    elif ishurdles(event):
       e = event + ' ' + hurdles[event][klasse]
    else:
       e = event

    return e


def club_code(club_name):
    if club_name in ('IL Koll', 'Idrettslaget Koll', 'Koll, IL', 'Koll, Idrettslaget'):
        club_code = 'KOLL'
    elif club_name in ('IL i BUL', 'Idrottslaget i BUL', 'BUL, IL i', 'BUL, Idrottslaget i'):
        club_code = 'ILBUL'
    elif club_name in ( 'IK Tjalve', 'Idrettsklubben Tjalve', 'Tjalve, IK', 'Tjalve, Idrettsklubben', 'Tjalve Idrettsklubben' ):
        club_code = 'TJAL'
    elif club_name in ( 'Tyrving IL', 'Tyrving Idrettslag' ):
        club_code = 'TYR'
    elif club_name in ( 'Romerike Friidrett' ):
        club_code = 'ROMFR'
    elif club_name in ( u'Bækkelagets SK' ): 
        club_code = 'BSK'
    elif club_name in ( 'Nesodden IF' ): 
        club_code = 'NESO'
    elif club_name in ( 'Groruddalen Friidrettsklubb' ): 
        club_code = 'GRO'
    elif club_name in ( 'Idrettslaget Sandvin', 'IL Sandvin' ): 
        club_code = 'SANDV'
    elif club_name in ( 'Eidanger Idrettslag' ): 
        club_code = 'EIDA'
    elif club_name in ( 'Ski IL Friidrett' ): 
        club_code = 'SKI'
    elif club_name in ( 'Gui Sportsklubb' ): 
        club_code = 'GUI'
    elif club_name in ( 'Sturla IF', 'Idrettsforeningen Sturla'): 
        club_code = 'STUR'
    else:
        club_code = club_name

    return club_code

def club_name(club_code):
    if club_code == 'KOLL':
        club_name = 'IL Koll'
    elif club_code in ( 'ILBUL', 'ILIBUL'):
        club_name = 'IL i BUL'
    elif club_code ==  'TJAL':
        club_name = 'IK Tjalve'
    elif club_code ==  'TYR':
        club_name = 'Tyrving IL'
    elif club_code ==  'ROMFR':
        club_name = 'Romerike Friidrett'
    elif club_code ==  'BSK':
        club_name = u'Bækkelagets SK'
    elif club_code ==  'NESO':
        club_name = 'Nesodden IF'
    elif club_code ==  'GRO':
        club_name = 'Groruddalen FIK'
    elif club_code ==  'SANDV':
        club_name = 'IL Sandvin'
    elif club_code ==  'EIDA':
        club_name = 'Eidanger Idrettslag'
    elif club_code ==  'SKI':
        club_name = 'Ski IL Friidrett'
    elif club_code ==  'GUI':
        club_name = 'Gui Sportsklubb'
    elif club_code ==  'STUR':
        club_name = 'IF Sturla'
    else:
        club_name=club_code

    return club_name
   

 
 
# ...
if len(sys.argv) < 2:
   sys.exit("Usage: %s <infile>" % sys.argv[0])
   
infile = sys.argv[1]
print(infile)
tree = read_xml_into_tree(infile)
save_xml_copy(tree)


"""
events_crosstable = make_crosstable(tree)
for e1 in sorted( events_crosstable.keys() ):
    for e2 in sorted( events_crosstable[e1].keys() ):
        print e1 +'|'+ e2, events_crosstable[e1][e2]
"""

write_xlsx_results_template(tree)
write_start_lists_as_html(tree)


write_opentrack_import(tree)

"""
event = 'Lengde satssone'
classes = [ 'F6', 'F7', 'G8', 'J8', 'J9' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G10', 'G11', 'J10', 'J11' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G 12', 'G13', 'J12', 'J13' ]
make_horizontal_protocol(tree, event, classes)
event = 'Lengde'
classes = [ 'G14', 'J14', 'J15']
make_horizontal_protocol(tree, event, classes)

event = 'Kule'
classes = [ 'G10', 'G11', 'J10', 'J11' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G14', 'G17', 'J14', 'J15' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'J12', 'J13']
make_horizontal_protocol(tree, event, classes)


event = u'Høyde'
classes = [ 'J13', 'G14', 'J15']
make_vertical_protocol(tree, event, classes)
classes = [ 'J10', 'J11', 'J12']
make_vertical_protocol(tree, event, classes)
classes = [ 'G10', 'G11', 'G12']
make_vertical_protocol(tree, event, classes)
event = 'Lengde'
classes = [ 'J14' ]
make_horizontal_protocol(tree, event, classes)
event = 'Kule'
classes = [ 'J11', 'J12', 'J13', 'G11' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G13', 'J14' ]
make_horizontal_protocol(tree, event, classes)
classes = [ 'G14' ]
make_horizontal_protocol(tree, event, classes)
event = 'Liten ball'
classes = [ 'J 9', 'J10', 'G 9', 'G10' ]
make_horizontal_protocol(tree, event, classes)
"""
