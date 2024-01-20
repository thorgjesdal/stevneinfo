from openpyxl import Workbook

from stevneinfo import events, categories as cats

def get_gender(cat):
    gender = {'G' : 'M', 'M' : 'M', 'J' : 'F', 'K' : 'F' }
    return gender[cat[0]]

def get_age(cat):
    if cat[1] == 'S':
        age = 'SEN'
    elif cat[1] in ['U', 'V']:
        age = cat[1:]
    elif cat[0] in ['G', 'J']:
        age = cat[-2:]
    return age

def ce_fullname(event):
    if event=='HEX':
        name = '6-kamp'
    elif event == 'HEP':
        name = '7-kamp'
    elif event == 'ENN':
        name = '9-kamp'
    elif event == 'DEC':
        name = '10-kamp'
    elif event == 'TRI':
        name = '3-kamp'
    elif event == 'QUAD':
        name = '4-kamp'
    elif event == 'PEN':
        name = '5-kamp'
    return name

def event_name(code):
    event_names = {
            '60'     : '60 meter'          , 
            '80'     : '80 meter'          , 
            '100'    : '100 meter'         , 
            '150'    : '150 meter'         , 
            '200'    : '200 meter'         , 
            '300'    : '300 meter'         , 
            '400'    : '400 meter'         , 
            '600'    : '600 meter'         , 
            '800'    : '800 meter'         , 
            '1000'   : '1000 meter'        , 
            '2000'   : '2000 meter'        , 
            '1500'   : '1500 meter'        , 
            '3000'   : '3000 meter'        , 
            '5000'   : '5000 meter'        , 
            '10000'  : '10000 meter'       , 
            '60H'    : '60 meter hekk'     , 
            '80H'    : '80 meter hekk'     , 
            '100H'   : '100 meter hekk'    , 
            '110H'   : '110 meter hekk'    , 
            '200H'   : '200 meter hekk'    ,
            '300H'   : '300 meter hekk'    , 
            '400H'   : '400 meter hekk'    , 
            '1500SC' : '1500 meter hinder' , 
            '2000SC' : '2000 meter hinder' , 
            '3000SC' : '3000 meter hinder' , 
            '1000W'  : 'Kappgang 1000 meter'        , 
            '2000W'  : 'Kappgang 2000 meter'        , 
            '3000W'  : 'Kappgang 3000 meter'        , 
            'HJ'     : 'Høyde'             , 
            'PV'     : 'Stav'              , 
            'LJ'     : 'Lengde'            , 
            'TJ'     : 'Tresteg'           , 
            'SP'     : 'Kule'              , 
            'DT'     : 'Diskos'            , 
            'HT'     : 'Slegge'            , 
            'JT'     : 'Spyd'              , 
            'OT'     : 'Liten ball'              , 
            'BT'     : 'Liten ball'              , 
            'DEC'    : 'Tikamp'            , 
            'HEP'    : 'Sjukamp'           ,
            'PEN'    : 'Femkamp'           ,
            'SHJ'    : 'Høyde uten tilløp' ,
            'SLJ'    : 'Lengde uten tilløp'           ,
            'STJ'    : 'Tresteg uten tilløp'           
            }
    return event_names[code]


#multis = { }
#multis['G13'] =  {'ot_code' : 'M01', 'ce_code' : 'HEX', 'events' : [[ '60','LJ', 'SP' ],['60H', 'HJ', '600']]}
"""
multis = { 
        'G13':  {'ot_code' : 'M01', 'ce_code' : 'HEX', 'events' : [[ '60','LJ', 'SP' ],['60H', 'HJ', '600']]},
        'G14':  {'ot_code' : 'M02', 'ce_code' : 'HEX', 'events' : [[ '60','LJ', 'SP' ],['60H', 'HJ', '600']]},
        'G15':  {'ot_code' : 'M03', 'ce_code' : 'ENN', 'events' : [[ '100', 'LJ', 'SP', 'HJ' ],['100H', 'DT', 'PV', 'JT', '1000']]},
        'G16':  {'ot_code' : 'M04', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['100H', 'DT', 'PV', 'JT', '1500']]},
        'G17':  {'ot_code' : 'M05', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'J13': {'ot_code' : 'M06', 'ce_code' : 'HEX', 'events' : [[ '60','HJ', 'SP' ],['60H', 'LJ', '600']]},
        'J14': {'ot_code' : 'M07', 'ce_code' : 'HEX', 'events' : [[ '60','HJ', 'SP' ],['60H', 'LJ', '600']]},
        'J15': {'ot_code' : 'M08', 'ce_code' : 'HEP', 'events' : [[ '80H', 'HJ', 'SP', '200' ],['LJ', 'JT', '800']]},
        'J16': {'ot_code' : 'M09', 'ce_code' : 'HEP', 'events' : [[ '80H', 'HJ', 'SP', '200' ],['LJ', 'JT', '800']]},
        'J17': {'ot_code' : 'M10', 'ce_code' : 'HEP', 'events' : [[ '100H', 'HJ', 'SP', '200' ],['LJ', 'JT', '800']]},
        'MU20': {'ot_code' : 'M11', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'MS': {'ot_code' : 'M12', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'KU20': {'ot_code' : 'M13', 'ce_code' : 'HEP', 'events' : [[ '100H', 'HJ', 'SP', '200' ],['LJ', 'JT', '800']]},
        'KS': {'ot_code' : 'M14', 'ce_code' : 'HEP', 'events' : [[ '100H', 'HJ', 'SP', '200' ],['LJ', 'JT', '800']]},
        'MV35': {'ot_code' : 'M15', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'MV40': {'ot_code' : 'M16', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'MV45': {'ot_code' : 'M17', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'MV50': {'ot_code' : 'M18', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'MV55': {'ot_code' : 'M19', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'MV75': {'ot_code' : 'M20', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]},
        'KV50': {'ot_code' : 'M21', 'ce_code' : 'HEP', 'events' : [[ '100H', 'HJ', 'SP', '200' ],['LJ', 'JT', '800']]},
        'MV85': {'ot_code' : 'M22', 'ce_code' : 'DEC', 'events' : [[ '100', 'LJ', 'SP', 'HJ', '400' ],['110H', 'DT', 'PV', 'JT', '1500']]}
            }
"""
multis = { 
        'G13':  {'ot_code' : 'M01', 'ce_code' : 'QUAD', 'events' : [[ '60H','LJ', 'SP','600' ]]},
        'G14':  {'ot_code' : 'M02', 'ce_code' : 'QUAD', 'events' : [[ '60H','LJ', 'SP','600' ]]},
        'G15':  {'ot_code' : 'M05', 'ce_code' : 'HEP', 'events' : [[ '60', 'LJ', 'SP', 'HJ' ],['60H', 'PV','800']]},
        'G16':  {'ot_code' : 'M06', 'ce_code' : 'HEP', 'events' : [[ '60', 'LJ', 'SP', 'HJ' ],['60H', 'PV','800']]},
        'G17':  {'ot_code' : 'M07', 'ce_code' : 'HEP', 'events' : [[ '60', 'LJ', 'SP', 'HJ' ],['60H', 'PV','1000']]},
        'J13':  {'ot_code' : 'M03', 'ce_code' : 'QUAD', 'events' : [[ '60H','LJ', 'SP','600' ]]},
        'J14':  {'ot_code' : 'M04', 'ce_code' : 'QUAD', 'events' : [[ '60H','LJ', 'SP','600' ]]},
        'J15': {'ot_code' : 'M08', 'ce_code' : 'PEN', 'events' : [[],['60H', 'HJ', 'SP', 'LJ','600']]},
        'J16': {'ot_code' : 'M09', 'ce_code' : 'PEN', 'events' : [[],['60H', 'HJ', 'SP', 'LJ','600']]},
        'J17': {'ot_code' : 'M10', 'ce_code' : 'PEN', 'events' : [[],['60H', 'HJ', 'SP', 'LJ','800']]},
        'MU20':  {'ot_code' : 'M11', 'ce_code' : 'HEP', 'events' : [[ '60', 'LJ', 'SP', 'HJ' ],['60H', 'PV','1000']]},
        'MS':  {'ot_code' : 'M12', 'ce_code' : 'HEP', 'events' : [[ '60', 'LJ', 'SP', 'HJ' ],['60H', 'PV','1000']]},
        'KU20': {'ot_code' : 'M13', 'ce_code' : 'PEN', 'events' : [[],['60H', 'HJ', 'SP', 'LJ','800']]},
        'KS': {'ot_code' : 'M14', 'ce_code' : 'PEN', 'events' : [[],['60H', 'HJ', 'SP', 'LJ','800']]},
            }
#print(multis)
#print(multis.keys())

#... write to xlsx workbook
wb = Workbook()
ws = wb.active
ws.title = 'Events'
row_counter = 0

for k in multis.keys():
    #
    cat = k
    combined_events = multis[k]['ce_code']
    day = 1
    if not multis[k]['events'][0]:
        day = 2

    row_counter +=1
    ws["A%d"%row_counter] = multis[cat]['ot_code']
    ws["B%d"%row_counter] = combined_events
    ws["C%d"%row_counter] = get_age(cat)
    ws["D%d"%row_counter] = cats.get_gender(cat)
    ws["E%d"%row_counter] = k#+multis[k]['ce_code'][0]
    ws["G%d"%row_counter] = f'{cat} {events.event_spec(combined_events, cat)}'
    ws["H%d"%row_counter] = '1'
    ws["I%d"%row_counter] = day
    ws["J%d"%row_counter] = '12:00'


event_code = 0
for k in multis.keys():
    #print(multis[k])
    #print(multis[k]['events'])
    #print(multis[k]['events'][0])
    #print(multis[k]['events'][1])
    cat = k
    parent = multis[k]['ot_code']
    ce_name = ce_fullname(multis[k]['ce_code'])
    gender = cats.get_gender(cat)
    age = get_age(cat)
    for i, events_by_day in enumerate(multis[k]['events']):
        day = i+1
        for event in events_by_day:
            event_code +=1
            print(event_code, event, age, gender, cat, f'{cat} {event} ({ce_name})', '1', day, '12:00', parent)
            
            row_counter +=1
            ws["A%d"%row_counter] = f'{event_code:03d}'
            ws["B%d"%row_counter] = event
            ws["C%d"%row_counter] = age
            ws["D%d"%row_counter] = gender
            ws["E%d"%row_counter] = k#+multis[k]['ce_code'][0]
            ws["G%d"%row_counter] = f'{cat} {events.event_spec(event, cat)}'
            ws["H%d"%row_counter] = '1'
            ws["I%d"%row_counter] = day
            ws["J%d"%row_counter] = '12:00'
            ws["K%d"%row_counter] = parent


xlname = 'combined_events_table.xlsx'
wb.save(xlname)
   



