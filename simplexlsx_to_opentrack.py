import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict
import datetime
import requests
from bs4 import BeautifulSoup
#from fuzzywuzzy import fuzz
#from fuzzywuzzy import process
from stevneinfo import statistics as stats

import pprint

def get_gender(cat):
    gender_nor = {'M':'M', 'K':'F', 'G':'M', 'J':'F'}
    gender_eng = {'M':'M', 'W':'F', 'B':'M', 'G':'F'}
    if cat[0] in ('M', 'K', 'G', 'J'):
        g = gender_nor[cat[0]]
    elif cat[-1] in ('M', 'W', 'B', 'G'):
        g = gender_eng[cat[-1]]

    return g

def read_eventfile(f):
    wb = load_workbook(filename=f)
    ws = wb.active


    event_codes= {}
    for value in ws.iter_rows(min_row=1,min_col=1, max_col=10, values_only=True):
        cat = value[4]
        event = f'{value[1]}'
        code = value[0]

        key = (cat, event)
        print(key,code)
        event_codes[key] = code
        
    print(event_codes)
    return event_codes


def read_simplexlsx(f):
    wb = load_workbook(filename=f)
    ws = wb.active

    events_by_athlete= {}
    for value in ws.iter_rows(min_row=2,min_col=1, max_col=46, values_only=True):
        #print(value)
        if value[2] is not None:
            print(value)
            cat = value[0]
            print(cat)
            ev    = f'{value[1]}'.strip()
            maxname = 30
            fn = value[2]
            first_name   = fn[0:min(len(fn),maxname)]
            ln = value[3]
            last_name = ln[0:min(len(ln),maxname)]
            dob = value[4]
            team = value[9]
            sb = value[10]
            pb = value[11]
            if get_stats:
                #athlete_id = stats.get_athlete_id(first_name,last_name,datetime.datetime.strftime(dob,ddmmyyyyformat))
                athlete_id = stats.get_athlete_id(first_name,last_name,dob)
                i = -1
                print(cat, i)
               #if cat.endswith('X'):
               #    i = -2
                pb, sb =  stats.get_athlete_bests(athlete_id, ev, cat[0:i])
                #pb, sb =  stats.get_athlete_bests(athlete_id, ev, cat)
                qp = pb

            event = (cat, ev, sb, pb)
            athlete_key = (first_name, last_name, dob, team)
            print('++',athlete_key)

            if athlete_key not in events_by_athlete.keys():
                events_by_athlete[athlete_key] = []
            if event not in events_by_athlete[athlete_key]:
                events_by_athlete[athlete_key].append( event )

    print( events_by_athlete)
    return events_by_athlete




def write_opentrack_import(ef, cf):
#   print('Before')
    event_codes = read_eventfile(ef)
    print(event_codes)
    events_by_athlete = read_simplexlsx(cf)

    isodateformat = "%Y-%m-%d"
    ddmmyyyyformat = "%d.%m.%Y"
    #... write opentrack bulk import to xlsx workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Competitors'
    
    row_counter = 1
    ws["A1"] = 'Competitor Id'
    ws["B1"] = 'National Id'
    ws["C1"] = 'First name'
    ws["D1"] = 'Last name'
    ws["E1"] = 'Gender'
    ws["F1"] = 'Date of birth'
    ws["G1"] = 'Team ID'
#   ws["H1"] = 'Nationality'
    ws["H1"] = 'Event'
    ws["I1"] = 'Pb'
    ws["J1"] = 'Sb'
    row_counter = 2

    row_counter = 2    
    bib = 0
    pp = pprint.PrettyPrinter(indent=4)
    pp.pprint(event_codes)
#   pp.pprint(full_events)
    for key in events_by_athlete.keys():
        bib+=1
        fn   = key[0]
        ln   = key[1]
        dob  = key[2]
        #dob = datetime.datetime.strptime(dob,ddmmyyyyformat)
        team = key[3]
        #qp = key[4]
        print('+',key)
        for event in events_by_athlete[key]:
            print(event)
            e = (event[0], event[1])
            sb = event[2]
            pb = event[3]
            print(e)
#           e = ( event[2], event[1] )

            ws["A%d"%row_counter] = bib
#           ws["B%d"%row_counter] = ident
            ws["C%d"%row_counter] = fn
            ws["D%d"%row_counter] = ln
            #ws["E%d"%row_counter] = gender[event[0][0]]
            ws["E%d"%row_counter] = get_gender(event[0])
            print(dob)
            ws["F%d"%row_counter] = datetime.datetime.strftime(dob,isodateformat)
            ws["G%d"%row_counter] = team
            ws["H%d"%row_counter] = event_codes[e]
            ws["I%d"%row_counter] = pb
            ws["J%d"%row_counter] = sb

            row_counter +=1

    xlname = 'opentrack_input.xlsx'
    wb.save(xlname)
#-----

event_file = 'boysen_events.xlsx'
#event_file = 'ExtraEvents_NMMangekamp.xlsx'
#competitor_file = 'Etteranmeldinger_BassenSprint.xlsx'

if len(sys.argv) < 2:
   sys.exit("Usage: %s <infile>" % sys.argv[0])
   
infile = sys.argv[1]
print(infile)

get_stats=False

#pp = pprint.PrettyPrinter(indent=4)
#pp.pprint(event_codes)
#pp.pprint(events_by_athlete)

write_opentrack_import(event_file, infile)

