import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict
import datetime
import requests
from bs4 import BeautifulSoup
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from stevneinfo import clubs

import pprint

gender = {'M':'M', 'K':'F', 'G':'M', 'J':'F'}

def istrack(event):
    return 'meter' in event

def ishurdles(event):
    return istrack(event) and 'hekk' in event

def issteeple(event):
    return istrack(event) and 'hinder' in event

def isfield(event):
    #return event in ('HJ', 'PV', 'LJ', 'TJ', 'SHJ', 'SLJ', 'STJ',
    #                 'SP', 'DT', 'HT', 'JT', 'BT', 'OT' )
    return isvjump(event) or ishjump(event) or isthrow(event)

def isvjump(event):
    return event in [u'Høyde', u'Stav', u'Høyde uten tilløp']

def ishjump(event):
    return event in [u'Lengde', u'Lengde satssone', u'Tresteg', u'Lengde uten tilløp', u'Tresteg uten tilløp']

def isthrow(event):
    return event in [u'Kule', u'Diskos', u'Slegge', u'Spyd', u'Vektkast', u'Liten ball']

def ismulti(event):
    return 'kamp' in event

def event_code(event):
    event_codes = {
            u'60 meter'          : '60', 
            u'80 meter'          : '80', 
            u'100 meter'         : '100', 
            u'150 meter'         : '150', 
            u'200 meter'         : '200', 
            u'300 meter'         : '300', 
            u'400 meter'         : '400', 
            u'600 meter'         : '600', 
            u'800 meter'         : '800', 
            u'1000 meter'        : '1000', 
            u'1500 meter'        : '1500', 
            u'3000 meter'        : '3000', 
            u'5000 meter'        : '5000', 
            u'10 000 meter'       : '10000', 
            u'60 meter hekk'     : '60H', 
            u'80 meter hekk'     : '80H', 
            u'100 meter hekk'    : '100H', 
            u'110 meter hekk'    : '110H', 
            u'200 meter hekk'    : '200H', 
            u'300 meter hekk'    : '300H', 
            u'400 meter hekk'    : '400H', 
            u'2000 meter hinder' : '2000SC', 
            u'3000 meter hinder' : '3000SC', 
            u'Kappgang 1000 meter' : '1000W', 
            u'Kappgang 3000 meter' : '3000W', 
            u'Kappgang 5000 meter' : '5000W', 
            u'Kappgang 5 km'     : '5000W', 
            u'Kappgang'          : '1500W', 
            u'Høyde'             : 'HJ', 
            u'Høyde uten tilløp' : 'SHJ', 
            u'Stav'              : 'PV', 
            u'Lengde'            : 'LJ', 
            u'Lengde satssone'   : 'LJ', 
            u'Lengde uten tilløp': 'SLJ', 
            u'Tresteg'           : 'TJ', 
            u'Tresteg satssone'  : 'TJ', 
            u'Tresteg uten tilløp'  : 'STJ', 
            u'Kule'              : 'SP', 
            u'Diskos'            : 'DT', 
            u'Slegge'            : 'HT', 
            u'Spyd'              : 'JT', 
            u'Liten ball'        : 'OT', 
            u'Tikamp'            : 'DEC', 
            u'Sjukamp'           : 'HEP' ,
            u'10-kamp'           : 'DEC' ,
            u'9-kamp'           : 'ENN' ,
            u'7-kamp'           : 'HEP' ,
            u'6-kamp'           : 'HEX' ,
            u'5-kamp'           : 'PEN' ,
            u'4x200 meter stafett' : '4x200' 
            }
    return event_codes.get(event, '')

def cat_code(name):
    cat_codes = {
            u'6 år Fellesklasse' : u'F6' ,
            u'7 år Fellesklasse' : u'F7' ,
            u'Gutter Rekrutt'    : u'G6-7'          , 
            u'Gutter 6 - 7'      : u'G6-7'          , 
            u'Gutter 8'     : u'G8'          , 
            u'Gutter 9'     : u'G9'          , 
            u'Gutter  9-10' : u'G9-10'          , 
            u'Gutter 10'    : u'G10'          , 
            u'Gutter 11'    : u'G11'          , 
            u'Gutter 11-12' : u'G11-12'       , 
            u'Gutter 11-14' : u'G11-14'       , 
            u'Gutter 12'    : u'G12'          , 
            u'Gutter 13'    : u'G13'          , 
            u'Gutter 14'    : u'G14'          , 
            u'Gutter 15'    : u'G15'          , 
            u'Gutter 15-17' : u'G15-17'          , 
            u'Gutter 16'    : u'G16'          , 
            u'Gutter 17'    : u'G17'          , 
            u'Gutter 18/19' : u'G18/19'       , 
            u'Gutter 18-19' : u'G18-19'       , 
            u'Gutter alle klasser' : u'GALLE'       , 
            u'Menn junior'  : u'MJ'           , 
            u'Menn U20'     : u'MU20'         , 
            u'Menn u20'     : u'MU20'         , 
            u'Menn U23'     : u'MU23'         , 
            u'Menn senior'  : u'MS'           , 
            u'Menn Senior Para'  : u'MSPARA'           , 
            u'Menn alle klasser'  : u'MALLE'           , 
            u'Menn veteraner' : u'MV'         , 
            u'Menn veteran 35-39' : u'MV35'         , 
            u'Menn veteran 40-44' : u'MV40'         , 
            u'Menn veteran 45-49' : u'MV45'         , 
            u'Menn veteran 50-54' : u'MV50'         , 
            u'Menn veteran 55-59' : u'MV55'         , 
            u'Menn veteran 60-64' : u'MV60'         , 
            u'Menn veteran 65-69' : u'MV65'         , 
            u'Menn veteran 70-75' : u'MV70'         , 
            u'Menn veteran 75-79' : u'MV75'         , 
            u'Jenter Rekrutt'    : u'J6-7'          , 
            u'Jenter 6 - 7'     : u'J 6-7'          , 
            u'Jenter 8'     : u'J8'          , 
            u'Jenter 9'     : u'J9'          , 
            u'Jenter  9-10'     : u'J9-10'          , 
            u'Jenter 10'    : u'J10'          , 
            u'Jenter 11'    : u'J11'          , 
            u'Jenter 11-12' : u'J11-12'          , 
            u'Jenter 11-14' : u'J11-14'          , 
            u'Jenter 12'    : u'J12'          , 
            u'Jenter 13'    : u'J13'          , 
            u'Jenter 14'    : u'J14'          , 
            u'Jenter 15'    : u'J15'          , 
            u'Jenter 15-17' : u'J15-17'    , 
            u'Jenter 16'    : u'J16'          , 
            u'Jenter 17'    : u'J17'          , 
            u'Jenter 18/19' : u'J18/19'       , 
            u'Jenter 18-19' : u'J18-19'       , 
            u'Jenter alle klasser' : u'JALLE'       , 
            u'Kvinner junior'  : u'KJ'        , 
            u'Kvinner U20'     : u'KU20'      , 
            u'Kvinner u20'     : u'KU20'      , 
            u'Kvinner U23'     : u'KU23'      , 
            u'Kvinner u23'     : u'KU23'      , 
            u'Kvinner senior'  : u'KS'        , 
            u'Kvinner Senior'  : u'KS'        , 
            u'Kvinner alle klasser'  : u'KALLE'           , 
            u'Kvinner veteraner' : u'KV'      ,
            u'Kvinner veteran 35-39' : u'KV35'         , 
            u'Kvinner veteran 40-44' : u'KV40'         , 
            u'Kvinner veteran 45-49' : u'KV45'         , 
            u'Kvinner veteran 50-54' : u'KV50'         , 
            u'Kvinner veteran 55-59' : u'KV55'         , 
            u'Kvinner veteran 60-64' : u'KV60'         , 
            u'Kvinner veteran 65-69' : u'KV65'         , 
            u'Kvinner veteran 70-75' : u'KV70'         , 
            u'Kvinner veteran 75-79' : u'KV75'         , 
            u'Funksjonshemmede' : u'FH'      ,
            u'Ikke valgt klasse' : u'IVK'
            }
#   print(name)
    code = ''
    if name is not None:
        code = cat_codes.get(name.strip(), name.strip())
    return code

def get_gender(cat):
    if cat[0] in ('G', 'M'):
        g = 'M'
    elif cat[0] in ('J', 'K'):
        g = 'F'
    else:
        g = 'MF'
    return g

def age_group(cat):
    age_groups = {
            'F6'    : '6',
            'F7'    : '7',
            'G6-7'    : '6-7',
            'G8'    : '8',
            'G9'    : '9',
            'G10'    : '10',
            'G11'    : '11',
            'G11-14'    : '11-14',
            'G12'    : '12',
            'G13'    : '13',
            'G14'    : '14',
            'G15'    : '15',
            'G16'    : '16',
            'G17'    : '17',
            'G18/19' : '18-19',
            'G18-19' : '18-19',
            'GALLE' : 'ALL',
            'MJ'     : 'U20' ,
            'MU20'     : 'U20' ,
            'MU23'     : 'U23' ,
            'MS'     : 'SEN' ,
            'MALLE'     : 'ALL' ,
            'MV'     : 'V35' ,
            'MV35'   : 'V35' ,
            'MV40'   : 'V40' ,
            'MV45'   : 'V45' ,
            'MV50'   : 'V50' ,
            'MV55'   : 'V55' ,
            'MV60'   : 'V60' ,
            'MV65'   : 'V65' ,
            'MV70'   : 'V70' ,
            'MV75'   : 'V75' ,
            'J6-7'    : '6-7',
            'J8'    : '8',
            'J9'    : '9',
            'J10'    : '10',
            'J11'    : '11',
            'J11-14'    : '11-14',
            'J12'    : '12',
            'J13'    : '13',
            'J14'    : '14',
            'J15'    : '15',
            'J16'    : '16',
            'J17'    : '17',
            'J18/19' : '18-19',
            'J18-19' : '18-19',
            'JALLE' : 'ALL',
            'KJ'     : 'U20',
            'KU20'     : 'U20' ,
            'KU23'     : 'U23' ,
            'KS'     : 'SEN' ,
            'KV'     : 'V35' ,
            'KV35'   : 'V35' ,
            'KV40'   : 'V40' ,
            'KV45'   : 'V45' ,
            'KV50'   : 'V50' ,
            'KV55'   : 'V55' ,
            'KV60'   : 'V60' ,
            'KV65'   : 'V65' ,
            'KV70'   : 'V70' ,
            'KV75'   : 'V75' ,
            'KALLE'     : 'ALL' ,
            'FH'   : 'ALL' ,
            'IVK'    : 'ALL'  
            }
    return age_groups.get(cat, cat)

def event_spec(event, klasse):
    throws = {}
    throws['Kule'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg', 
                       'default' : ''} 
    throws['Diskos'] = { 'J10' : '0,6kg', 'J11' : '0,6kg', 'J12' : '0,6kg', 'J13' : '0,6kg', 
                       'J14' : '0,75kg', 'J15' : '0,75kg', 'J16' : '0,75kg', 'J17' : '0,75kg',
                       'J18/19' : '1,0kg', 'KU20' : '1,0kg', 'KU23' : '1,0kg', 'KS' : '1,0kg', 
                       'G10' : '0,6kg', 'G11' : '0,6kg', 'G12' : '0,75kg', 'G13' : '0,75kg', 
                       'G14' : '1,0kg', 'G15' : '1,0kg', 'G16' : '1,5kg', 'G17' : '1,5kg',
                       'G18/19' : '1,75kg', 'MU20' : '1,75kg', 'MU23' : '2,0kg', 'MS' : '2,0kg',
                       'default': ''} 
    throws['Slegge'] = { 'J10' : '2,0kg', 'J11' : '2,0kg', 'J12' : '2,0kg', 'J13' : '2,0kg', 
                       'J14' : '3,0kg', 'J15' : '3,0kg', 'J16' : '3,0kg', 'J17' : '3,0kg',
                       'J18/19' : '4,0kg', 'KU20' : '4,0kg', 'KU23' : '4,0kg', 'KS' : '4,0kg', 
                       'G10' : '2,0kg', 'G11' : '2,0kg', 'G12' : '3,0kg', 'G13' : '3,0kg', 
                       'G14' : '4,0kg', 'G15' : '4,0kg', 'G16' : '5,0kg', 'G17' : '5,0kg',
                       'G18/19' : '6,0kg', 'MU20' : '6,0kg', 'MU23' : '7,26kg', 'MS' : '7,26kg',
                       'default': ''} 
    throws['Spyd'] = { 'J10' : '400g', 'J11' : '400g', 'J12' : '400g', 'J13' : '400g', 
                       'J14' : '400g', 'J15' : '500g', 'J16' : '500g', 'J17' : '500g',
                       'J18/19' : '600g', 'KU20' : '600g', 'KU23' : '600g', 'KS' : '600g', 
                       'G10' : '400g', 'G11' : '400g', 'G12' : '400g', 'G13' : '400g', 
                       'G14' : '600g', 'G15' : '600g', 'G16' : '700g', 'G17' : '700g',
                       'G18/19' : '800g', 'MU20' : '800g', 'MU23' : '800g', 'MS' : '800g',
                       'default': ''} 
#    throws['Liten ball'] = { 'J10' : '150g', 'J11' : '150g', 'J12' : '150g', 'J13' : '150g', 'J14' : '150g', 
#                             'G10' : '150g', 'G11' : '150g', 'G12' : '150g', 'G13' : '150g', 'G14' : '150g' 
#                             }
    throws['Liten ball'] = defaultdict(lambda : '150g') 
    hurdles = {}
    hurdles['60 meter hekk'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '76,2cm', 'J13' : '76,2cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '84,0cm','KJ' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '76,2cm', 'G13' : '76,2cm', 'G14' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm', 'G17' : '91,4cm',
                                 'G18/19' : '100cm','MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm', 'default':'' }
    hurdles['80 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'G14' : '84,0cm' } 
    hurdles['100 meter hekk'] = { 'J16' : '76,2cm', 'J17' : '76,2cm', 'J18/19' : '84,0cm','KJ' : '84,0cm','KU20' : '84,0cm', 'KU23' : '84,0cm', 'KS' : '84,0cm',
                                 'G15' : '84,0cm', 'G16' : '91,4cm'}

    hurdles['110 meter hekk'] = { 'G17' : '91,4cm', 'G18/19' : '100cm','MJ' : '100cm', 'MU20' : '100cm', 'MU23' : '106,7cm', 'MS' : '106,7cm' }
    hurdles['200 meter hekk'] = { 'J10' : '68,0cm', 'J11' : '68,0cm', 'J12' : '68,0cm', 'J13' : '68,0cm', 'J14' : '76,2cm',
                                 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KJ' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G10' : '68,0cm', 'G11' : '68,0cm', 'G12' : '68,0cm', 'G13' : '68,0cm', 'G14' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '76,2cm', 'G17' : '76,2cm',
                                 'G18/19' : '76,2cm','MJ' : '76,2cm', 'MU20' : '76,2cm', 'MU23' : '76,2cm', 'MS' : '76,2cm' }
    hurdles['300 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KJ' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MJ' : '91,4cm', 'MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm' }
    hurdles['400 meter hekk'] = { 'J15' : '76,2cm', 'J16' : '76,2cm', 'J17' : '76,2cm',
                                 'J18/19' : '76,2cm','KJ' : '76,2cm','KU20' : '76,2cm', 'KU23' : '76,2cm', 'KS' : '76,2cm',
                                 'G15' : '76,2cm', 'G16' : '84,0cm', 'G17' : '84,0cm',
                                 'G18/19' : '91,4cm','MJ' : '91,4cm','MU20' : '91,4cm', 'MU23' : '91,4cm', 'MS' : '91,4cm' }

    if isthrow(event):
       #e = event + ' ' + throws[event][klasse]
       t = throws[event].get(klasse,None)
       if t == None:
           t = throws[event]['default']
       e = event + ' ' + t
    elif ishurdles(event):
       h = hurdles[event].get(klasse,None)
       if h == None:
           h = hurdles[event]['default']
       e = event + ' ' + h
    else:
       e = event

    return e



def sort_event_list(events):
    def sort_fcn(e):
        #print(e)
        catsort = ['G6-7', 'G8', 'G9', 'G9-10', 'G10', 'G11', 'G11-12', 'G11-14', 'G12', 'G13', 'G14', 'G15', 'G15-17', 'G16', 'G17', 'G18-19', 
                   'J6-7', 'J8', 'J9', 'J9-10', 'J10', 'J11', 'J11-12', 'J11-14', 'J12', 'J13', 'J14', 'J15', 'J15-17', 'J16', 'J17', 'J18-19', 
                   'MU20', 'MU23', 'MS', 'KU20', 'KU23', 'KS', 
                   'MSPARA',
                   'MV', 'MV35', 'MV40', 'MV45', 'MV50', 'MV55', 'MV60', 'MV65', 'MV70', 'MV75',
                   'KV', 'KV35', 'KV40', 'KV45', 'KV50', 'KV55', 'KV60', 'KV65', 'KV70', 'KV75']

        evsort = ['60', '100', '200', '400', '600', '800', '1000', '1500', 'MILE', '3000', '5000', '10000', 
                '60H', '80H', '100H', '110H', '200H', '300H', '400H', '2000SC', '3000SC', 
                'HJ', 'PV', 'LJ', 'TJ', 'SP', 'DT', 'JT', 'HT', 'OT', 'PEN', 'HEX', 'HEP', 'ENN', 'DEC']
        return 100*catsort.index(e[2]) + evsort.index(e[1])

    print(events)
    events.sort(key=sort_fcn)
    return events

def read_isonenxls(f):
    wb = load_workbook(filename=f)
    ws = wb.active

#   columns = ws[1]
#   print(list(columns))
#   sys.exit()
    events =  []
    events_by_athlete= {}
    i=1
    for value in ws.iter_rows(min_row=1,min_col=1, max_col=46, values_only=True):
        if i==1:
            columns = value
            i+=1
            continue
        if value[0] is None:
            continue
        if 'Avmeldt' not in value[columns.index('Påmeldingsstatus')] or value[columns.index('Påmeldingsstatus')] is None:
            first_name = value[columns.index('Fornavn')]
            last_name = value[columns.index('Etternavn')]
            dob = value[columns.index('Fødselsdato')]
            g = gender[value[columns.index('Kjønn')]]
            club = value[columns.index('Klubb')]
            ev = value[columns.index('Øvelse')]
            cat = cat_code(value[columns.index('Klasse')])
            nat = cat_code(value[columns.index('Landskode')])
            athlete_key = (first_name, last_name, dob, g, club, nat)
            event = (ev, event_code(ev),  cat)
            print(athlete_key)
            print(event)

            if event[0] is None:
                continue
            if event not in events:
                events.append(event)
    
            if athlete_key not in events_by_athlete.keys():
                events_by_athlete[athlete_key] = []
            if event not in events_by_athlete[athlete_key]:
                events_by_athlete[athlete_key].append( event )

#   events = sort_event_list(events)
    return events, events_by_athlete

def get_stats(event,cat,season):
    event_id = {'100':'4', '200': '5', '400':'7', '800':'9', '1500':'11', '3000':'13', '5000':'14', '10000':'15',
            '100H':'35', '110H':'42', '400H':'59', '2000SC':'65', '3000SC' : '121'}
    catcodes = {'KS': '22', 'MS': '11'}

    if event in event_id.keys():
       url = 'https://www.minfriidrettsstatistikk.info/php/LandsStatistikk.php?showclass=' + cat + '&showevent=' + event_id[event] + '&outdoor=Y&showseason=' + season + '&showclub=0'
       r = requests.get(url)

    
       soup = BeautifulSoup(r.text, 'html.parser')
       tables = [
           [
               [td.get_text(strip=True) for td in tr.find_all('td')]
               for tr in table.find_all('tr')
           ]
           for table in soup.find_all('table')
       ] 
   
       stats = []
       for table in tables:
           for row in table:
               if not row == [] and not row[0] == '-----':
                  name, club  =  row[1].rsplit(',', 1)
                  dob = row[2]
                  perf = row[0]
                  stats.append( (name, dob, perf) )
       return stats



def get_seed_marks(name, dob, event, cat, season): 
    event_id = {'100':'4', '200': 5, '400':'7', 800:'9', '1500':'11', '3000':'13', '5000':'14', '10000':'15',
            '100H':'35', '110H':'42', '400H':'59', '3000SC' : '120' ,
            'HJ':'68', 'PV':'70', 'LJ':'71', 'TJ':'75', } 
    catcodes = {'KS': '22', 'MS': '11',
            'G15':'6', 'G16':'7', 'G17':'8', 'G18/19':'9',
            'J15':'17', 'J16':'18', 'J17':'19', 'J18/19':'20'}

    print(cat,event)
    if cat not in ('MS', 'KS'):
        return ''
    #event = event_code(event)
    cat    = catcodes[cat]

    global event_stats
    event_stats = {}

    if event not in event_stats.keys():
        event_stats[event] = {}
    if cat not in event_stats[event]:
        event_stats[event][cat] = {}
    if season not in event_stats[event][cat]:
        event_stats[event][cat][season] = get_stats(event,cat,season)

    res = 'nm'
    s = event_stats.get(event, None)
    if not s is None:
        if not s[cat][season] is None:
           for p in s[cat][season]:
               nme = p[0]

               ratio = fuzz.token_set_ratio(name, nme)
               if ratio>85:
                   if 'Magnus' in name:
                       print(name, nme, ratio, p[2])
                   res = p[2]
                   break

        minsecpat = '(\d?\d)[:.,](\d\d[,.]\d?\d)'
        match1 = re.match(minsecpat,res)
        reswindpat = "(\d?\d[,.]\d\d)[(]([+-]\d[,.]\d)[)]"
        match2 = re.match(reswindpat,res)
        if match1:
            mins = match1.group(1)
            secs = match1.group(2).replace(',','.')
            res = mins + ':' + secs
        elif match2:
            secs = match2.group(1)
            wind = match2.group(2)
            res = secs.replace(',','.')
        else:
            res = res.replace(',','.')
        return res



def write_opentrack_import(f):
    events, events_by_athlete = read_isonenxls(f)

    isodateformat = "%Y-%m-%d"
    ddmmyyyyformat = "%d.%m.%Y"
    #... write opentrack bulk import to xlsx workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Competitors'
    ws1 = wb.create_sheet('Events')
    
    row_counter = 1
    ws["A1"] = 'Competitor Id'
    ws["B1"] = 'National Id'
    ws["C1"] = 'First name'
    ws["D1"] = 'Last name'
    ws["E1"] = 'Gender'
    ws["F1"] = 'Date of birth'
    ws["G1"] = 'Team ID'
    ws["H1"] = 'Event'
    ws["J1"] = 'Pb'
    ws["K1"] = 'Sb'
#   ws1["A1"] = 'Event selection'
    row_counter = 2

    jf = 0
    jt = 0
    jm = 0
    full_events = {}
    for e in events:
        event  = e[0]
        evcode = e[1]
        cat    = e[2]
        if isfield(event):
            jf +=1
            event_ref = "F%02d"%jf
        elif ismulti(event):
            jm +=1
            event_ref = "M%02d"%jm
        else:
            jt +=1
            event_ref = "T%02d"%jt

        #print(e)
        full_events[ ( cat , evcode ) ]  = event_ref + ' - ' + ' '.join(( cat, event_spec(event, cat) ))
        ws1["A%d"%row_counter] = event_ref + ' - '  + ' '.join([e[0], event_spec(e[1], cat)])
        ws1["B%d"%row_counter] = event_ref
        ws1["C%d"%row_counter] = evcode
        ws1["D%d"%row_counter] = age_group(cat)
        ws1["E%d"%row_counter] = get_gender(cat)
        ws1["F%d"%row_counter] = cat
#       ws1["G%d"%row_counter] = age_group(class_code(e[0]))

        ws1["H%d"%row_counter] = ' '.join(( cat, event_spec(event, cat) ))
        ws1["I%d"%row_counter] = '1'
        ws1["J%d"%row_counter] = '1'
        ws1["K%d"%row_counter] = '12:00'
        
        row_counter +=1
    ws1.delete_cols(1,1)
#   ws.insert_cols(13)
#   print (full_events)
    row_counter = 2    
    bib = 0
    pp = pprint.PrettyPrinter(indent=4)
#   pp.pprint(events_by_athlete)
#   pp.pprint(full_events)
    for key in events_by_athlete.keys():
        bib+=1
        maxname = 30
        fn   = key[0]
        fn   = fn[0:min(len(fn),maxname)]
        ln   = key[1]
        ln   = ln[0:min(len(ln),maxname)]
        dob  = key[2]
        dob = datetime.datetime.strptime(dob,ddmmyyyyformat)
        g    = key[3]
        club = key[4]
        nat  = key[5]
        for e in events_by_athlete[key]:
            #print(e)
            event     = e[0]
            eventcode = e[1]
            cat       = e[2]
            #e = ( event[2], event[1] )

            ws["A%d"%row_counter] = bib
#           ws["B%d"%row_counter] = ident
            ws["C%d"%row_counter] = fn
            ws["D%d"%row_counter] = ln
            ws["E%d"%row_counter] = g
            ws["F%d"%row_counter] = datetime.datetime.strftime(dob,isodateformat)
            ws["G%d"%row_counter] = clubs.club_code(club)
            ws["H%d"%row_counter] = full_events[ (e[2], e[1]) ]

            #event = e[1]
            if not isfield(event):
                if eventcode == "60": # for Bassen sprint
                    eventcode = "100"
                res1 = get_seed_marks(' '.join((fn, ln)), dob, eventcode, cat, '2022' )
                res = res1
                if res=='nm':
                    res=''
                if e[1] == '10 000 meter':
                    res3 = get_seed_marks(' '.join((fn, en)), dob, '5000 meter', e[0], '2021' )
                    res4 = get_seed_marks(' '.join((fn, en)), dob, '5000 meter', e[0], '2020' )
                    res3 = min(res3,res4)
                    if res3 == 'nm':
                        res3 = ''
                    ws["M%d"%row_counter] = res3
            else:
                res = ''
            ws["K%d"%row_counter] = res
            row_counter +=1

    xlname = 'opentrack_input.xlsx'
    wb.save(xlname)
#-----

if len(sys.argv) < 2:
   sys.exit("Usage: %s <infile>" % sys.argv[0])
   
infile = sys.argv[1]
print(infile)
#events, events_by_athlete = read_isonenxls(infile)
write_opentrack_import(infile)
#print(events)
#pp = pprint.PrettyPrinter(indent=4)
#pp.pprint(events_by_athlete)
