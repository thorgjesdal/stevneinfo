import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict
import datetime
import requests
from bs4 import BeautifulSoup
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import json
import argparse

from stevneinfo import clubs, categories as cats, events, statistics as stats

import pprint

gender = {'M':'M', 'K':'F', 'G':'M', 'J':'F'}

isodateformat = "%Y-%m-%d"
ddmmyyyyformat = "%d.%m.%Y"

def read_isonenxls(f):
    wb = load_workbook(filename=f)
    ws = wb.active

#   columns = ws[1]
#   print(list(columns))
#   sys.exit()
    event_list =  []
    events_by_athlete= {}
    days = []
    i=1
    for value in ws.iter_rows(min_row=1,min_col=1, max_col=46, values_only=True):
        if i==1:
            columns = value
            i+=1
            continue
        if value[0] is None:
            continue
        if 'Avmeldt' not in value[columns.index('Påmeldingsstatus')] or value[columns.index('Påmeldingsstatus')] is None:
            first_name = value[columns.index('Fornavn')]
            last_name = value[columns.index('Etternavn')]
            dob = value[columns.index('Fødselsdato')]
            g = gender[value[columns.index('Kjønn')]]
            club = value[columns.index('Klubb')]
            if value[columns.index('Øvelse')]:
                ev = value[columns.index('Øvelse')].strip()
            else:
                ev = ""
            cat = cats.cat_code(value[columns.index('Klasse')])
            #nat = cats.cat_code(value[columns.index('Landskode')])
            nat = value[columns.index('Landskode')]
            day = value[columns.index('Dato')]
            day = datetime.datetime.strptime(day,ddmmyyyyformat)
            if day not in days:
                days.append(day)
            athlete_key = (first_name, last_name, dob, g, club, nat)
            event = (events.event_code(ev),  cat, day)

            if event[0] is None:
                continue
            if event not in event_list:
                event_list.append(event)
    
            if athlete_key not in events_by_athlete.keys():
                events_by_athlete[athlete_key] = []
            if event not in events_by_athlete[athlete_key]:
                events_by_athlete[athlete_key].append( event )

#   events = sort_event_list(events)
    days.sort()
    for i,e in enumerate(event_list):
        event_list[i] = ( e[0], e[1], days.index(e[2])+1 )
    return event_list, events_by_athlete, days

def get_stats(event,cat,season):
    event_id = {'100':'4', '200': '5', '400':'7', '800':'9', '1500':'11', '3000':'13', '5000':'14', '10000':'15',
            '100H':'35', '110H':'42', '400H':'59', '2000SC':'65', '3000SC' : '121'}
    catcodes = {'KS': '22', 'MS': '11'}

    if event in event_id.keys():
       url = 'https://www.minfriidrettsstatistikk.info/php/LandsStatistikk.php?showclass=' + cat + '&showevent=' + event_id[event] + '&outdoor=Y&showseason=' + season + '&showclub=0'
       r = requests.get(url)

    
       soup = BeautifulSoup(r.text, 'html.parser')
       tables = [
           [
               [td.get_text(strip=True) for td in tr.find_all('td')]
               for tr in table.find_all('tr')
           ]
           for table in soup.find_all('table')
       ] 
   
       stats = []
       for table in tables:
           for row in table:
               if not row == [] and not row[0] == '-----':
                  name, club  =  row[1].rsplit(',', 1)
                  dob = row[2]
                  perf = row[0]
                  stats.append( (name, dob, perf) )
       return stats



def get_seed_marks(name, dob, event, cat, season): 
    event_id = {'100':'4', '200': 5, '400':'7', 800:'9', '1500':'11', '3000':'13', '5000':'14', '10000':'15',
            '100H':'35', '110H':'42', '400H':'59', '3000SC' : '120' ,
            'HJ':'68', 'PV':'70', 'LJ':'71', 'TJ':'75', } 
    catcodes = {'KS': '22', 'MS': '11',
            'G15':'6', 'G16':'7', 'G17':'8', 'G18/19':'9',
            'J15':'17', 'J16':'18', 'J17':'19', 'J18/19':'20'}

    cat    = catcodes[cat]

    global event_stats
    event_stats = {}

    if event not in event_stats.keys():
        event_stats[event] = {}
    if cat not in event_stats[event]:
        event_stats[event][cat] = {}
    if season not in event_stats[event][cat]:
        event_stats[event][cat][season] = get_stats(event,cat,season)

    res = 'nm'
    s = event_stats.get(event, None)
    if not s is None:
        if not s[cat][season] is None:
           for p in s[cat][season]:
               nme = p[0]

               ratio = fuzz.token_set_ratio(name, nme)
               if ratio>85:
                   if 'Magnus' in name:
                       print(name, nme, ratio, p[2])
                   res = p[2]
                   break

        minsecpat = '(\d?\d)[:.,](\d\d[,.]\d?\d)'
        match1 = re.match(minsecpat,res)
        reswindpat = "(\d?\d[,.]\d\d)[(]([+-]\d[,.]\d)[)]"
        match2 = re.match(reswindpat,res)
        if match1:
            mins = match1.group(1)
            secs = match1.group(2).replace(',','.')
            res = mins + ':' + secs
        elif match2:
            secs = match2.group(1)
            wind = match2.group(2)
            res = secs.replace(',','.')
        else:
            res = res.replace(',','.')
        return res

#def build_event_table_from_input():
    #

def build_event_table_from_json(url):
    #
    print(url)
    r = requests.get(url+'json')
    j = json.loads(r.text)

    event_list = []

    for e in j['events']:
        cat = e['category']
        eventcode = e['eventCode']
        day = e['day']

        event_list.append((eventcode, cat, day))
    return event_list

def write_opentrack_import(f):
#   event_list, events_by_athlete = read_isonenxls(f)

    isodateformat = "%Y-%m-%d"
    ddmmyyyyformat = "%d.%m.%Y"
    #... write opentrack bulk import to xlsx workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Competitors'
    ws1 = wb.create_sheet('Events')
    
    row_counter = 1
    ws["A1"] = 'Competitor Id'
    ws["B1"] = 'National Id'
    ws["C1"] = 'First name'
    ws["D1"] = 'Last name'
    ws["E1"] = 'Gender'
    ws["F1"] = 'Date of birth'
    ws["H1"] = 'Team ID'
    ws["I1"] = 'Event'
    ws["J1"] = 'Pb'
    ws["K1"] = 'Sb'
#   ws1["A1"] = 'Event selection'
    row_counter = 2

    multis = {}
    jm = 0
    for e in event_list:
        if events.ismulti(e[0]):
            jm +=1
            event_ref = "M%02d"%jm
            multis[event_ref] = ev[0]


    jf = 0
    jt = 0
    print(event_list)
    full_events = {}
    for e in event_list:
        print(e)
        evcode = e[0]
        event  = events.event_name(evcode)
        cat    = e[1]
        day    = e[2]
        #day    = days.index(e[2]) + 1
        if events.isfield(event):
            jf +=1
            event_ref = "F%02d"%jf
        elif events.ismulti(event):
            jm +=1
            event_ref = "M%02d"%jm
        else:
            jt +=1
            event_ref = "T%02d"%jt

        full_events[ ( cat , evcode ) ]  = event_ref + ' - ' + ' '.join(( cat, events.event_spec(evcode, cat) ))
#       full_events[ ( cat , event) ]  = event_ref + ' - ' + ' '.join(( cat, events.event_spec(event, cat) ))
        ws1["A%d"%row_counter] = event_ref + ' - '  + ' '.join([e[0], events.event_spec(evcode, cat)])
        ws1["B%d"%row_counter] = event_ref
        ws1["C%d"%row_counter] = evcode
        ws1["D%d"%row_counter] = cats.age_group(cat)
        ws1["E%d"%row_counter] = cats.get_gender(cat)
        ws1["F%d"%row_counter] = cat
#       ws1["G%d"%row_counter] = age_group(class_code(e[0]))

        ws1["H%d"%row_counter] = ' '.join(( cat, events.event_spec(evcode, cat) ))
        ws1["I%d"%row_counter] = '1'
        ws1["J%d"%row_counter] = f'{day}'
        ws1["K%d"%row_counter] = ''
        
        row_counter +=1
    ws1.delete_cols(1,1)
#   ws.insert_cols(13)
    row_counter = 2    
    bib = 0
    pp = pprint.PrettyPrinter(indent=4)
#   pp.pprint(events_by_athlete)
#   pp.pprint(full_events)
    for key in events_by_athlete.keys():
        bib+=1
        maxname = 30
        fn   = key[0]
        #fn   = fn[0:min(len(fn),maxname)]
        ln   = key[1]
        #ln   = ln[0:min(len(ln),maxname)]
        dob  = key[2]
        dob = datetime.datetime.strptime(dob,ddmmyyyyformat)
        g    = key[3]
        club = key[4]
        nat  = key[5]
        print(fn, ln)
        for e in events_by_athlete[key]:
            print(e)
            eventcode = e[0]
            event     = events.event_name(eventcode)
            cat       = e[1]
            if events.ismulti(eventcode):
                cat = cat+eventcode[0]

            #e = ( event[2], event[1] )

            ws["A%d"%row_counter] = bib
#           ws["B%d"%row_counter] = ident
            ws["C%d"%row_counter] = fn[0:min(len(fn),maxname)]
            ws["D%d"%row_counter] = ln[0:min(len(ln),maxname)]
            ws["E%d"%row_counter] = g
            ws["F%d"%row_counter] = datetime.datetime.strftime(dob,isodateformat)
            ws["H%d"%row_counter] = clubs.club_code(club)
            ws["I%d"%row_counter] = full_events[ (cat, eventcode) ]
            #ws["J%d"%row_counter] = full_events[ (e[1], e[0]) ]

            #event = e[1]
            """
            if not events.isfield(event):
                if eventcode == "60": # for Bassen sprint
                    eventcode = "100"
                res1 = get_seed_marks(' '.join((fn, ln)), dob, eventcode, cat, '2022' )
                res = res1
                if res=='nm':
                    res=''
                if e[1] == '10 000 meter':
                    res3 = get_seed_marks(' '.join((fn, en)), dob, '5000 meter', e[0], '2021' )
                    res4 = get_seed_marks(' '.join((fn, en)), dob, '5000 meter', e[0], '2020' )
                    res3 = min(res3,res4)
                    if res3 == 'nm':
                        res3 = ''
                    ws["M%d"%row_counter] = res3
            else:
                res = ''
                """
            pb = ''
            sb = ''
            if args.get_stats:
                athlete_id = stats.get_athlete_id(fn,ln,datetime.datetime.strftime(dob,ddmmyyyyformat))
                """
                if eventcode == "60": # for Bassen sprint
                    eventcode = "100"
                """
                athlete_bests =  stats.get_athlete_bests(athlete_id, eventcode, cat)

                pb = athlete_bests[0]
                sb = athlete_bests[1]

            ws["J%d"%row_counter] = pb
            ws["K%d"%row_counter] = sb
            row_counter +=1

    xlname = 'opentrack_input.xlsx'
    wb.save(xlname)
#-----

#if len(sys.argv) < 2:
#   sys.exit("Usage: %s <infile>" % sys.argv[0])
parser = argparse.ArgumentParser()
parser.add_argument('infile')
parser.add_argument('--url', help='opentrack competition url', default=None)
parser.add_argument('--get_stats', action='store_true', default=False)
args = parser.parse_args()

infile = args.infile
   
#infile = sys.argv[1]
print(infile)
event_list, events_by_athlete, days = read_isonenxls(infile)
if args.url:
    event_list = build_event_table_from_json(args.url)
    

write_opentrack_import(infile)
#print(events)
#pp = pprint.PrettyPrinter(indent=4)
#pp.pprint(events_by_athlete)
